from __future__ import annotations

import math
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Iterable

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook import Workbook


class FormulaError(Exception):
    pass


@dataclass(slots=True)
class _Token:
    type: str
    value: str


@dataclass(slots=True)
class _EvalContext:
    workbook: Workbook
    sheet_name: str
    current_cell: str | None
    cache: dict[tuple[str, str], Any]
    stack: set[tuple[str, str]]
    locals: dict[str, Any]

    def with_locals(self, locals_map: dict[str, Any]) -> "_EvalContext":
        return _EvalContext(
            workbook=self.workbook,
            sheet_name=self.sheet_name,
            current_cell=self.current_cell,
            cache=self.cache,
            stack=self.stack,
            locals=dict(locals_map),
        )


class FormulaEngine:
    def evaluate_formula(
        self,
        workbook: Workbook,
        sheet_name: str,
        formula: str,
        current_cell: str | None = None,
        _cache: dict[tuple[str, str], Any] | None = None,
        _stack: set[tuple[str, str]] | None = None,
    ) -> Any:
        if not isinstance(formula, str) or not formula.startswith("="):
            return formula
        tokens = self._tokenize(formula[1:])
        parser = _Parser(tokens)
        node = parser.parse_expression()
        if parser.peek() is not None:
            raise FormulaError("Unexpected trailing tokens.")
        context = _EvalContext(
            workbook=workbook,
            sheet_name=sheet_name,
            current_cell=current_cell,
            cache=_cache if _cache is not None else {},
            stack=_stack if _stack is not None else set(),
            locals={},
        )
        return self._eval(node, context)

    def evaluate_cell(
        self,
        workbook: Workbook,
        sheet_name: str,
        cell_ref: str,
        _cache: dict[tuple[str, str], Any] | None = None,
        _stack: set[tuple[str, str]] | None = None,
    ) -> Any:
        worksheet = workbook[sheet_name]
        value = worksheet[cell_ref].value
        if isinstance(value, str) and value.startswith("="):
            key = (sheet_name, cell_ref.upper())
            cache = _cache if _cache is not None else {}
            if key in cache:
                return cache[key]
            stack = _stack if _stack is not None else set()
            if key in stack:
                raise FormulaError(f"Circular reference detected at {sheet_name}!{cell_ref}.")
            stack.add(key)
            try:
                result = self.evaluate_formula(workbook, sheet_name, value, current_cell=cell_ref, _cache=cache, _stack=stack)
                cache[key] = result
                return result
            finally:
                stack.discard(key)
        return value

    def evaluate_range(
        self,
        workbook: Workbook,
        sheet_name: str,
        start_cell: str,
        end_cell: str,
        _cache: dict[tuple[str, str], Any] | None = None,
        _stack: set[tuple[str, str]] | None = None,
    ) -> list[list[Any]]:
        start_col, start_row = self._split_cell(start_cell)
        end_col, end_row = self._split_cell(end_cell)
        cache = _cache if _cache is not None else {}
        stack = _stack if _stack is not None else set()
        rows: list[list[Any]] = []
        for row_idx in range(start_row, end_row + 1):
            row: list[Any] = []
            for col_idx in range(start_col, end_col + 1):
                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                row.append(self.evaluate_cell(workbook, sheet_name, cell_ref, cache, stack))
            rows.append(row)
        return rows

    def _eval(self, node: Any, context: _EvalContext) -> Any:
        kind = node[0]
        if kind == "number":
            return node[1]
        if kind == "string":
            return node[1]
        if kind == "bool":
            return node[1]
        if kind == "name":
            local_name = node[1].upper()
            if local_name in context.locals:
                return context.locals[local_name]
            resolved = self._resolve_named_value(context.workbook, local_name)
            if resolved is not None:
                return resolved
            return node[1]
        if kind == "ref":
            return self._resolve_ref(context, node[1], node[2], node[3])
        if kind == "structured_ref":
            return self._resolve_structured_ref(context.workbook, node[1], node[2])
        if kind == "unary":
            value = self._coerce_scalar(self._eval(node[2], context))
            if node[1] == "+":
                return self._to_number(value)
            if node[1] == "-":
                return -self._to_number(value)
            raise FormulaError(f"Unsupported unary operator {node[1]}")
        if kind == "binary":
            left = self._eval(node[2], context)
            right = self._eval(node[3], context)
            return self._apply_binary(node[1], left, right)
        if kind == "func":
            return self._call_function(node[1], node[2], context)
        raise FormulaError(f"Unsupported node {kind}")

    def _resolve_ref(self, context: _EvalContext, sheet_name: str | None, start_cell: str, end_cell: str | None) -> Any:
        target_sheet = sheet_name or context.sheet_name
        if target_sheet not in context.workbook.sheetnames:
            raise FormulaError(f"Sheet not found: {target_sheet}")
        if end_cell is None:
            return self.evaluate_cell(context.workbook, target_sheet, self._normalize_cell(start_cell), context.cache, context.stack)
        return self.evaluate_range(
            context.workbook,
            target_sheet,
            self._normalize_cell(start_cell),
            self._normalize_cell(end_cell),
            context.cache,
            context.stack,
        )

    def _resolve_named_value(self, workbook: Workbook, name: str) -> Any | None:
        defined_names = getattr(workbook, "defined_names", None)
        if defined_names is None:
            return None
        try:
            defined_name = defined_names.get(name)
        except Exception:
            defined_name = None
        if defined_name is None:
            try:
                for candidate_name in defined_names.keys():
                    if str(candidate_name).upper() == name.upper():
                        defined_name = defined_names.get(candidate_name)
                        break
            except Exception:
                defined_name = None
        if defined_name is None:
            return None
        destinations = list(getattr(defined_name, "destinations", []) or [])
        if destinations:
            sheet_name, reference = destinations[0]
            if not sheet_name or not reference:
                return None
            if ":" in reference:
                start_cell, end_cell = reference.split(":", 1)
                return self.evaluate_range(workbook, sheet_name, start_cell, end_cell)
            return self.evaluate_cell(workbook, sheet_name, reference)
        attr_text = getattr(defined_name, "attr_text", None)
        if isinstance(attr_text, str) and attr_text:
            return attr_text
        return None

    def _resolve_structured_ref(self, workbook: Workbook, table_name: str, spec: list[str]) -> Any:
        table = None
        table_sheet = None
        for worksheet in workbook.worksheets:
            if table_name in worksheet.tables:
                table = worksheet.tables[table_name]
                table_sheet = worksheet
                break
        if table is None or table_sheet is None:
            raise FormulaError(f"Table not found: {table_name}")

        ref = table.ref
        start_cell, end_cell = ref.split(":", 1)
        start_col, start_row = self._split_cell(start_cell)
        end_col, end_row = self._split_cell(end_cell)
        header_row = start_row
        data_start_row = start_row + 1
        headers = [table_sheet.cell(row=header_row, column=col_idx).value for col_idx in range(start_col, end_col + 1)]

        normalized_spec = [part.strip() for part in spec if part.strip()]
        if not normalized_spec or any(part.upper() == "#ALL" for part in normalized_spec):
            return self.evaluate_range(workbook, table_sheet.title, start_cell, end_cell)
        if any(part.upper() == "#HEADERS" for part in normalized_spec):
            return headers
        if any(part.upper() == "#DATA" for part in normalized_spec):
            return self.evaluate_range(
                workbook,
                table_sheet.title,
                f"{get_column_letter(start_col)}{data_start_row}",
                f"{get_column_letter(end_col)}{end_row}",
            )

        column_name = next((part for part in normalized_spec if not part.startswith("#")), None)
        if column_name is None:
            return self.evaluate_range(workbook, table_sheet.title, start_cell, end_cell)
        column_index = None
        for offset, header in enumerate(headers, start=start_col):
            if self._values_equal(header, column_name):
                column_index = offset
                break
        if column_index is None:
            raise FormulaError(f"Column '{column_name}' not found in table '{table_name}'.")
        return self.evaluate_range(
            workbook,
            table_sheet.title,
            f"{get_column_letter(column_index)}{data_start_row}",
            f"{get_column_letter(column_index)}{end_row}",
        )

    def _call_function(self, name: str, args: list[Any], context: _EvalContext) -> Any:
        upper = name.upper()
        if upper == "LET":
            return self._let(args, context)
        if upper == "IF":
            condition = self._eval(args[0], context)
            if self._truthy(condition):
                return self._eval(args[1], context) if len(args) > 1 else True
            return self._eval(args[2], context) if len(args) > 2 else False
        if upper == "IFERROR":
            try:
                return self._eval(args[0], context)
            except Exception:
                return self._eval(args[1], context) if len(args) > 1 else ""
        if upper == "IFNA":
            try:
                return self._eval(args[0], context)
            except Exception:
                return self._eval(args[1], context) if len(args) > 1 else ""
        if upper == "OFFSET":
            return self._offset(context, args)
        if upper == "INDIRECT":
            return self._indirect(context, args)
        if upper == "CHOOSE":
            return self._choose(context, args)
        if upper == "SWITCH":
            return self._switch(context, args)
        values = [self._eval(arg, context) for arg in args]
        if upper == "SUM":
            return sum(self._numbers(values))
        if upper == "AVERAGE":
            nums = self._numbers(values)
            return sum(nums) / len(nums) if nums else 0
        if upper == "MIN":
            nums = self._numbers(values)
            return min(nums) if nums else 0
        if upper == "MAX":
            nums = self._numbers(values)
            return max(nums) if nums else 0
        if upper == "COUNT":
            return sum(1 for value in self._flatten(values) if self._is_number(value))
        if upper == "COUNTA":
            return sum(1 for value in self._flatten(values) if value not in (None, ""))
        if upper == "ROUND":
            return round(self._to_number(values[0]), int(self._to_number(values[1]) if len(values) > 1 else 0))
        if upper == "ABS":
            return abs(self._to_number(values[0]))
        if upper == "INT":
            return math.floor(self._to_number(values[0]))
        if upper == "LEN":
            return len(self._to_text(values[0]))
        if upper == "LOWER":
            return self._to_text(values[0]).lower()
        if upper == "UPPER":
            return self._to_text(values[0]).upper()
        if upper == "TRIM":
            return self._to_text(values[0]).strip()
        if upper == "LEFT":
            return self._to_text(values[0])[: int(self._to_number(values[1]) if len(values) > 1 else 1)]
        if upper == "RIGHT":
            count = int(self._to_number(values[1]) if len(values) > 1 else 1)
            return self._to_text(values[0])[-count:]
        if upper == "MID":
            start = int(self._to_number(values[1]) if len(values) > 1 else 1) - 1
            count = int(self._to_number(values[2]) if len(values) > 2 else 1)
            return self._to_text(values[0])[start : start + count]
        if upper == "CONCAT":
            return "".join(self._to_text(value) for value in self._flatten(values) if value is not None)
        if upper == "TEXTJOIN":
            delimiter = self._to_text(values[0])
            ignore_empty = self._truthy(values[1]) if len(values) > 1 else False
            parts = [self._to_text(value) for value in self._flatten(values[2:]) if not (ignore_empty and value in (None, ""))]
            return delimiter.join(parts)
        if upper == "AND":
            return all(self._truthy(value) for value in self._flatten(values))
        if upper == "OR":
            return any(self._truthy(value) for value in self._flatten(values))
        if upper == "NOT":
            return not self._truthy(values[0])
        if upper == "SUMIF":
            return self._sumif(values)
        if upper == "COUNTIF":
            return self._countif(values)
        if upper == "AVERAGEIF":
            result = self._sumif(values)
            count = self._countif(values)
            return result / count if count else 0
        if upper == "SUMIFS":
            return self._sumifs(values)
        if upper == "COUNTIFS":
            return self._countifs(values)
        if upper == "AVERAGEIFS":
            total = self._sumifs(values)
            count = self._countifs(values)
            return total / count if count else 0
        if upper == "MATCH":
            return self._match(values)
        if upper == "INDEX":
            return self._index(values)
        if upper == "VLOOKUP":
            return self._vlookup(values)
        if upper == "XLOOKUP":
            return self._xlookup(values)
        if upper == "UNIQUE":
            return self._unique(values)
        if upper == "SORT":
            return self._sort(values)
        if upper == "FILTER":
            return self._filter(values)
        if upper == "SUMPRODUCT":
            return self._sumproduct(values)
        raise FormulaError(f"Function not implemented yet: {name}")

    def _let(self, args: list[Any], context: _EvalContext) -> Any:
        if len(args) < 3:
            raise FormulaError("LET requires at least one binding and a final expression.")
        bindings = args[:-1]
        if len(bindings) % 2 != 0:
            raise FormulaError("LET bindings must be name/value pairs.")
        locals_map = dict(context.locals)
        for idx in range(0, len(bindings), 2):
            name_node = bindings[idx]
            if name_node[0] != "name":
                raise FormulaError("LET binding name must be an identifier.")
            value = self._eval(bindings[idx + 1], context.with_locals(locals_map))
            locals_map[name_node[1].upper()] = value
        return self._eval(args[-1], context.with_locals(locals_map))

    def _sumif(self, values: list[Any]) -> float:
        if len(values) < 2:
            return 0.0
        criteria_range = self._flatten([values[0]])
        criteria = self._to_text(values[1])
        sum_range = self._flatten([values[2]]) if len(values) > 2 else criteria_range
        total = 0.0
        for idx, candidate in enumerate(criteria_range):
            if self._matches_criteria(candidate, criteria) and idx < len(sum_range):
                total += self._to_number(sum_range[idx])
        return total

    def _countif(self, values: list[Any]) -> int:
        if len(values) < 2:
            return 0
        criteria_range = self._flatten([values[0]])
        criteria = self._to_text(values[1])
        return sum(1 for candidate in criteria_range if self._matches_criteria(candidate, criteria))

    def _sumifs(self, values: list[Any]) -> float:
        if len(values) < 3:
            return 0.0
        sum_range = self._flatten([values[0]])
        criteria_pairs = list(zip(values[1::2], values[2::2]))
        total = 0.0
        for idx, candidate in enumerate(sum_range):
            matched = True
            for criteria_range, criteria in criteria_pairs:
                flat = self._flatten([criteria_range])
                if idx >= len(flat) or not self._matches_criteria(flat[idx], self._to_text(criteria)):
                    matched = False
                    break
            if matched:
                total += self._to_number(candidate)
        return total

    def _countifs(self, values: list[Any]) -> int:
        if len(values) < 2:
            return 0
        criteria_pairs = list(zip(values[0::2], values[1::2]))
        first_range = self._flatten([criteria_pairs[0][0]])
        count = 0
        for idx in range(len(first_range)):
            matched = True
            for criteria_range, criteria in criteria_pairs:
                flat = self._flatten([criteria_range])
                if idx >= len(flat) or not self._matches_criteria(flat[idx], self._to_text(criteria)):
                    matched = False
                    break
            if matched:
                count += 1
        return count

    def _match(self, values: list[Any]) -> int:
        if len(values) < 2:
            raise FormulaError("MATCH requires lookup value and lookup array.")
        lookup_value = self._coerce_scalar(values[0])
        lookup_array = self._flatten([values[1]])
        for idx, candidate in enumerate(lookup_array, start=1):
            if self._values_equal(candidate, lookup_value):
                return idx
        raise FormulaError("MATCH not found.")

    def _index(self, values: list[Any]) -> Any:
        if not values:
            raise FormulaError("INDEX requires an array.")
        array = self._normalize_array(values[0])
        row_num = int(self._to_number(values[1])) if len(values) > 1 else 1
        col_num = int(self._to_number(values[2])) if len(values) > 2 else 1
        if row_num < 1 or col_num < 1:
            raise FormulaError("INDEX row and column numbers must be positive.")
        if row_num > len(array) or col_num > len(array[0]):
            raise FormulaError("INDEX out of range.")
        return array[row_num - 1][col_num - 1]

    def _vlookup(self, values: list[Any]) -> Any:
        if len(values) < 3:
            raise FormulaError("VLOOKUP requires lookup value, table array, and column index.")
        lookup_value = self._coerce_scalar(values[0])
        table = self._normalize_array(values[1])
        col_index = int(self._to_number(values[2]))
        for row in table:
            if row and self._values_equal(row[0], lookup_value):
                if col_index - 1 < len(row):
                    return row[col_index - 1]
                return None
        if len(values) > 3:
            return values[3]
        raise FormulaError("VLOOKUP not found.")

    def _xlookup(self, values: list[Any]) -> Any:
        if len(values) < 3:
            raise FormulaError("XLOOKUP requires lookup value, lookup array, and return array.")
        lookup_value = self._coerce_scalar(values[0])
        lookup_array = self._flatten([values[1]])
        return_array = self._flatten([values[2]])
        fallback = values[3] if len(values) > 3 else ""
        for idx, candidate in enumerate(lookup_array):
            if self._values_equal(candidate, lookup_value):
                return return_array[idx] if idx < len(return_array) else fallback
        return fallback

    def _unique(self, values: list[Any]) -> list[Any]:
        flattened = self._flatten(values)
        seen: set[Any] = set()
        result: list[Any] = []
        for value in flattened:
            key = self._hashable(value)
            if key in seen:
                continue
            seen.add(key)
            result.append(value)
        return result

    def _sort(self, values: list[Any]) -> list[Any]:
        array = self._normalize_array(values[0] if values else [])
        if not array:
            return []
        sort_index = int(self._to_number(values[1])) - 1 if len(values) > 1 else 0
        sort_order = self._to_number(values[2]) if len(values) > 2 else 1
        descending = sort_order < 0
        return sorted(array, key=lambda row: row[sort_index] if sort_index < len(row) else None, reverse=descending)

    def _filter(self, values: list[Any]) -> list[Any]:
        array = self._normalize_array(values[0] if values else [])
        if len(values) < 2:
            return array
        include = self._flatten([values[1]])
        result: list[list[Any]] = []
        for idx, row in enumerate(array):
            flag = include[idx] if idx < len(include) else False
            if self._truthy(flag):
                result.append(row)
        return result

    def _sumproduct(self, values: list[Any]) -> float:
        arrays = [self._flatten([value]) for value in values]
        if not arrays:
            return 0.0
        length = min(len(array) for array in arrays)
        total = 0.0
        for idx in range(length):
            product = 1.0
            for array in arrays:
                product *= self._to_number(array[idx])
            total += product
        return total

    def _offset(self, context: _EvalContext, args: list[Any]) -> Any:
        if not args:
            raise FormulaError("OFFSET requires a reference.")
        base_node = args[0]
        base = base_node if isinstance(base_node, tuple) and base_node[0] in {"ref", "structured_ref"} else self._eval(base_node, context)
        rows = int(self._to_number(self._eval(args[1], context))) if len(args) > 1 else 0
        cols = int(self._to_number(self._eval(args[2], context))) if len(args) > 2 else 0
        height = int(self._to_number(self._eval(args[3], context))) if len(args) > 3 else None
        width = int(self._to_number(self._eval(args[4], context))) if len(args) > 4 else None

        if isinstance(base, list):
            flat = self._flatten([base])
            if not flat:
                raise FormulaError("OFFSET received an empty reference.")
            base = flat[0]

        if isinstance(base, tuple) and base[0] == "ref":
            start_col, start_row = self._split_cell(base[2])
            end_col, end_row = self._split_cell(base[3] or base[2])
            sheet_name = base[1] or context.sheet_name
        elif isinstance(base, tuple) and base[0] == "structured_ref":
            resolved = self._resolve_structured_ref(context.workbook, base[1], base[2])
            if isinstance(resolved, list) and resolved:
                return resolved
            raise FormulaError("OFFSET cannot offset a structured reference directly.")
        elif isinstance(base, str) and re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", base.upper()):
            start_col, start_row = self._split_cell(base)
            end_col, end_row = start_col, start_row
            sheet_name = context.sheet_name
        else:
            raise FormulaError("OFFSET requires a cell or range reference.")

        start_col += cols
        start_row += rows
        height = height or (end_row - start_row + 1)
        width = width or (end_col - start_col + 1)
        return self.evaluate_range(
            context.workbook,
            sheet_name,
            f"{get_column_letter(start_col)}{start_row}",
            f"{get_column_letter(start_col + max(width, 1) - 1)}{start_row + max(height, 1) - 1}",
        )

    def _indirect(self, context: _EvalContext, args: list[Any]) -> Any:
        if not args:
            raise FormulaError("INDIRECT requires a text reference.")
        ref_text = self._to_text(self._eval(args[0], context)).strip()
        if not ref_text:
            raise FormulaError("INDIRECT received an empty reference.")
        if "!" in ref_text:
            sheet_name, ref_text = ref_text.split("!", 1)
            sheet_name = sheet_name.strip("'")
        else:
            sheet_name = context.sheet_name
        if ":" in ref_text:
            start_cell, end_cell = ref_text.split(":", 1)
            return self.evaluate_range(context.workbook, sheet_name, start_cell, end_cell)
        if re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", ref_text.upper()):
            return self.evaluate_cell(context.workbook, sheet_name, ref_text)
        named = self._resolve_named_value(context.workbook, ref_text.upper())
        if named is not None:
            return named
        raise FormulaError(f"INDIRECT could not resolve {ref_text}.")

    def _choose(self, context: _EvalContext, args: list[Any]) -> Any:
        if len(args) < 2:
            raise FormulaError("CHOOSE requires an index and at least one option.")
        index = int(self._to_number(self._eval(args[0], context)))
        if index < 1 or index >= len(args):
            raise FormulaError("CHOOSE index out of range.")
        return self._eval(args[index], context)

    def _switch(self, context: _EvalContext, args: list[Any]) -> Any:
        if len(args) < 3:
            raise FormulaError("SWITCH requires an expression and at least one case.")
        expression = self._eval(args[0], context)
        remaining = args[1:]
        default_expr = None
        if len(remaining) % 2 == 1:
            default_expr = remaining[-1]
            remaining = remaining[:-1]
        for idx in range(0, len(remaining), 2):
            if self._values_equal(expression, self._eval(remaining[idx], context)):
                return self._eval(remaining[idx + 1], context)
        if default_expr is not None:
            return self._eval(default_expr, context)
        raise FormulaError("SWITCH found no matching case.")

    def _matches_criteria(self, value: Any, criterion: str) -> bool:
        criterion = criterion.strip()
        if criterion in {"", "*"}:
            return value not in (None, "")
        if "*" in criterion:
            pattern = re.escape(criterion).replace(r"\*", ".*")
            return re.fullmatch(pattern, self._to_text(value), re.IGNORECASE) is not None
        for prefix in (">=", "<=", ">", "<", "="):
            if criterion.startswith(prefix):
                target = criterion[len(prefix) :]
                left = self._coerce_scalar(value)
                right = self._coerce_scalar(target)
                left_num = self._to_number(left) if self._is_number(left) else self._to_text(left)
                right_num = self._to_number(right) if self._is_number(right) else self._to_text(right)
                if prefix == ">=":
                    return left_num >= right_num
                if prefix == "<=":
                    return left_num <= right_num
                if prefix == ">":
                    return left_num > right_num
                if prefix == "<":
                    return left_num < right_num
                return self._values_equal(left_num, right_num)
        return self._to_text(value).lower() == criterion.lower()

    def _normalize_array(self, value: Any) -> list[list[Any]]:
        if isinstance(value, list) and value and isinstance(value[0], list):
            return value
        if isinstance(value, list):
            return [value]
        return [[value]]

    def _hashable(self, value: Any) -> Any:
        if isinstance(value, list):
            return tuple(self._hashable(item) for item in value)
        if isinstance(value, dict):
            return tuple(sorted((k, self._hashable(v)) for k, v in value.items()))
        return value

    @staticmethod
    def _flatten(values: Iterable[Any]) -> list[Any]:
        flattened: list[Any] = []
        for value in values:
            if isinstance(value, list):
                flattened.extend(FormulaEngine._flatten(value))
            else:
                flattened.append(value)
        return flattened

    @staticmethod
    def _numbers(values: Iterable[Any]) -> list[float]:
        return [FormulaEngine._to_number(value) for value in FormulaEngine._flatten(values) if value not in (None, "")]

    @staticmethod
    def _is_number(value: Any) -> bool:
        try:
            FormulaEngine._to_number(value)
            return True
        except Exception:
            return False

    @staticmethod
    def _coerce_scalar(value: Any) -> Any:
        if isinstance(value, list):
            flat = FormulaEngine._flatten(value)
            return flat[0] if flat else None
        return value

    @staticmethod
    def _to_number(value: Any) -> float:
        value = FormulaEngine._coerce_scalar(value)
        if value in (None, ""):
            return 0.0
        if isinstance(value, bool):
            return 1.0 if value else 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(",", "").strip())
        except Exception as exc:
            raise FormulaError(f"Cannot convert {value!r} to number.") from exc

    @staticmethod
    def _to_text(value: Any) -> str:
        value = FormulaEngine._coerce_scalar(value)
        if value is None:
            return ""
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        return str(value)

    @staticmethod
    def _truthy(value: Any) -> bool:
        value = FormulaEngine._coerce_scalar(value)
        if isinstance(value, bool):
            return value
        if value in (None, "", 0):
            return False
        return bool(value)

    @staticmethod
    def _values_equal(left: Any, right: Any) -> bool:
        left = FormulaEngine._coerce_scalar(left)
        right = FormulaEngine._coerce_scalar(right)
        if isinstance(left, (int, float)) and isinstance(right, (int, float)):
            return float(left) == float(right)
        return str(left).strip().lower() == str(right).strip().lower()

    @staticmethod
    def _apply_binary(operator: str, left: Any, right: Any) -> Any:
        if operator == "&":
            return f"{FormulaEngine._to_text(left)}{FormulaEngine._to_text(right)}"
        left_value = FormulaEngine._to_number(left) if operator in {"+", "-", "*", "/", "^"} else FormulaEngine._coerce_scalar(left)
        right_value = FormulaEngine._to_number(right) if operator in {"+", "-", "*", "/", "^"} else FormulaEngine._coerce_scalar(right)
        if operator == "+":
            return left_value + right_value
        if operator == "-":
            return left_value - right_value
        if operator == "*":
            return left_value * right_value
        if operator == "/":
            return left_value / right_value
        if operator == "^":
            return left_value**right_value
        if operator == "=":
            return FormulaEngine._values_equal(left_value, right_value)
        if operator == "<>":
            return not FormulaEngine._values_equal(left_value, right_value)
        if operator == "<":
            return left_value < right_value
        if operator == "<=":
            return left_value <= right_value
        if operator == ">":
            return left_value > right_value
        if operator == ">=":
            return left_value >= right_value
        raise FormulaError(f"Unsupported operator: {operator}")

    def _split_cell(self, cell_ref: str) -> tuple[int, int]:
        normalized = self._normalize_cell(cell_ref)
        match = re.fullmatch(r"([A-Z]{1,3})(\d+)", normalized)
        if not match:
            raise FormulaError(f"Invalid cell reference: {cell_ref}")
        return column_index_from_string(match.group(1)), int(match.group(2))

    def _normalize_cell(self, cell_ref: str) -> str:
        match = re.fullmatch(r"\$?([A-Z]{1,3})\$?(\d+)", cell_ref.upper())
        if not match:
            raise FormulaError(f"Invalid cell reference: {cell_ref}")
        return f"{match.group(1)}{match.group(2)}"

    def _tokenize(self, text: str) -> list[_Token]:
        tokens: list[_Token] = []
        idx = 0
        while idx < len(text):
            char = text[idx]
            if char.isspace():
                idx += 1
                continue
            if char == '"':
                end = idx + 1
                literal: list[str] = []
                while end < len(text):
                    if text[end] == '"' and (end + 1 >= len(text) or text[end + 1] != '"'):
                        break
                    if text[end] == '"' and end + 1 < len(text) and text[end + 1] == '"':
                        literal.append('"')
                        end += 2
                        continue
                    literal.append(text[end])
                    end += 1
                if end >= len(text):
                    raise FormulaError("Unterminated string literal.")
                tokens.append(_Token("STRING", "".join(literal)))
                idx = end + 1
                continue
            if text.startswith(">=", idx) or text.startswith("<=", idx) or text.startswith("<>", idx):
                tokens.append(_Token("OP", text[idx : idx + 2]))
                idx += 2
                continue
            if char in "+-*/^&=<>:,()![]#":
                tokens.append(_Token("OP", char))
                idx += 1
                continue
            if char == "'" and "!" in text[idx:]:
                sheet_match = re.match(r"'([^']+)'", text[idx:])
                if sheet_match:
                    tokens.append(_Token("SHEET", sheet_match.group(1)))
                    idx += len(sheet_match.group(0))
                    continue
            number_match = re.match(r"\d+(?:\.\d+)?", text[idx:])
            if number_match:
                tokens.append(_Token("NUMBER", number_match.group(0)))
                idx += len(number_match.group(0))
                continue
            ident_match = re.match(r"\$?[A-Za-z_][A-Za-z0-9_.\$]*", text[idx:])
            if ident_match:
                value = ident_match.group(0)
                upper = value.upper()
                if upper in {"TRUE", "FALSE"}:
                    tokens.append(_Token("BOOL", upper))
                elif re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", upper):
                    tokens.append(_Token("CELL", upper))
                else:
                    tokens.append(_Token("IDENT", value))
                idx += len(value)
                continue
            raise FormulaError(f"Unexpected character in formula: {char}")
        return tokens


class _Parser:
    def __init__(self, tokens: list[_Token]) -> None:
        self.tokens = tokens
        self.index = 0

    def peek(self) -> _Token | None:
        return self.tokens[self.index] if self.index < len(self.tokens) else None

    def pop(self) -> _Token:
        token = self.peek()
        if token is None:
            raise FormulaError("Unexpected end of formula.")
        self.index += 1
        return token

    def parse_expression(self) -> Any:
        return self.parse_comparison()

    def parse_comparison(self) -> Any:
        node = self.parse_concat()
        while self.peek() and self.peek().type == "OP" and self.peek().value in {"=", "<>", "<", "<=", ">", ">="}:
            op = self.pop().value
            right = self.parse_concat()
            node = ("binary", op, node, right)
        return node

    def parse_concat(self) -> Any:
        node = self.parse_additive()
        while self.peek() and self.peek().type == "OP" and self.peek().value == "&":
            op = self.pop().value
            right = self.parse_additive()
            node = ("binary", op, node, right)
        return node

    def parse_additive(self) -> Any:
        node = self.parse_term()
        while self.peek() and self.peek().type == "OP" and self.peek().value in {"+", "-"}:
            op = self.pop().value
            right = self.parse_term()
            node = ("binary", op, node, right)
        return node

    def parse_term(self) -> Any:
        node = self.parse_power()
        while self.peek() and self.peek().type == "OP" and self.peek().value in {"*", "/"}:
            op = self.pop().value
            right = self.parse_power()
            node = ("binary", op, node, right)
        return node

    def parse_power(self) -> Any:
        node = self.parse_unary()
        while self.peek() and self.peek().type == "OP" and self.peek().value == "^":
            op = self.pop().value
            right = self.parse_unary()
            node = ("binary", op, node, right)
        return node

    def parse_unary(self) -> Any:
        token = self.peek()
        if token and token.type == "OP" and token.value in {"+", "-"}:
            op = self.pop().value
            return ("unary", op, self.parse_unary())
        return self.parse_primary()

    def parse_primary(self) -> Any:
        token = self.pop()
        if token.type == "NUMBER":
            return ("number", float(token.value) if "." in token.value else int(token.value))
        if token.type == "STRING":
            return ("string", token.value)
        if token.type == "BOOL":
            return ("bool", token.value == "TRUE")
        if token.type == "IDENT":
            if self.peek() and self.peek().type == "OP" and self.peek().value == "!":
                self.pop()
                ref_token = self.pop()
                if ref_token.type != "CELL":
                    raise FormulaError("Expected cell reference after sheet name.")
                return self._parse_reference(token.value, ref_token.value)
            if self.peek() and self.peek().type == "OP" and self.peek().value == "[":
                return self._parse_structured_reference(token.value)
            if self.peek() and self.peek().type == "OP" and self.peek().value == "(":
                self.pop()
                args: list[Any] = []
                if not (self.peek() and self.peek().type == "OP" and self.peek().value == ")"):
                    while True:
                        args.append(self.parse_expression())
                        if self.peek() and self.peek().type == "OP" and self.peek().value == ",":
                            self.pop()
                            continue
                        break
                self._expect("OP", ")")
                return ("func", token.value, args)
            return ("name", token.value)
        if token.type == "CELL":
            return self._parse_reference(None, token.value)
        if token.type == "SHEET":
            self._expect("OP", "!")
            ref_token = self.pop()
            if ref_token.type != "CELL":
                raise FormulaError("Expected cell reference after sheet name.")
            return self._parse_reference(token.value, ref_token.value)
        if token.type == "OP" and token.value == "(":
            expr = self.parse_expression()
            self._expect("OP", ")")
            return expr
        raise FormulaError(f"Unexpected token: {token.type} {token.value}")

    def _parse_reference(self, sheet_name: str | None, start_cell: str) -> Any:
        end_cell = None
        if self.peek() and self.peek().type == "OP" and self.peek().value == ":":
            self.pop()
            ref_token = self.pop()
            if ref_token.type != "CELL":
                raise FormulaError("Expected cell reference after ':'.")
            end_cell = ref_token.value
        return ("ref", sheet_name, start_cell, end_cell)

    def _parse_structured_reference(self, table_name: str) -> Any:
        self._expect("OP", "[")
        parts: list[str] = []
        current: list[str] = []
        depth = 1
        while True:
            token = self.pop()
            if token.type == "OP" and token.value == "[":
                depth += 1
                if depth > 1:
                    current.append("[")
                continue
            if token.type == "OP" and token.value == "]":
                depth -= 1
                if depth == 0:
                    if current:
                        parts.append("".join(current).strip())
                    break
                current.append("]")
                continue
            if token.type == "OP" and token.value == "," and depth == 1:
                parts.append("".join(current).strip())
                current = []
                continue
            current.append(token.value)
        normalized_parts: list[str] = []
        for part in parts:
            cleaned = part.strip()
            cleaned = cleaned.strip("[]")
            cleaned = cleaned.replace("# ", "#")
            normalized_parts.append(cleaned)
        return ("structured_ref", table_name, normalized_parts)

    def _expect(self, type_name: str, value: str | None = None) -> _Token:
        token = self.pop()
        if token.type != type_name or (value is not None and token.value != value):
            raise FormulaError(f"Expected {type_name} {value or ''}".strip())
        return token
