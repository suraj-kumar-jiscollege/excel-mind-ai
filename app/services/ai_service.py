from __future__ import annotations

import json
import re
import sys
from typing import Any

import httpx
from fastapi import HTTPException
from openpyxl.utils import column_index_from_string

from app.config import settings
from app.models import ActionImpact, ActionPlan, ChartRecommendation, TaskStep, WorkbookSnapshot
from app.services.workbook_service import workbook_service


class AIService:
    async def preview_command(
        self,
        session_id: str,
        command: str,
        selected_cell: str | None = None,
        selected_value: Any = None,
    ) -> ActionPlan:
        snapshot = workbook_service.get_snapshot(session_id)
        command_lower = command.lower().strip()

        # Undo/Redo are still faster via heuristics
        if command_lower in ["undo", "undo last action", "revert"]:
            return self._undo_plan(snapshot)
        if command_lower in ["redo", "redo action", "repeat"]:
            return self._redo_plan(snapshot)

        # 1. Try Gemini first for everything else
        if settings.gemini_api_key:
            try:
                gemini_plan = await self._preview_with_gemini(command, snapshot, selected_cell, selected_value)
                if gemini_plan.action != "noop":
                    workbook_service.remember_plan(session_id, command, gemini_plan)
                    return gemini_plan
                else:
                    # Debug: Gemini returned noop
                    print(f"DEBUG: Gemini returned NOOP for command: {command}", file=sys.stderr)
            except Exception as e:
                print(f"Gemini API Error: {repr(e)}", file=sys.stderr)
        else:
            print("DEBUG: GEMINI_API_KEY is NOT SET in environment", file=sys.stderr)

        # 2. Fallback to heuristics
        heuristic_plan = self._preview_with_heuristics(command, snapshot, selected_cell, selected_value)
        
        # If both fail, give a more helpful 'Autonomous' error
        if heuristic_plan.action == "noop":
            if not settings.gemini_api_key:
                heuristic_plan.explanation = "Bhai, Gemini AI Key set nahi hai backend me. Please env variables check karein."
            else:
                heuristic_plan.explanation = "Bhai, Gemini ne is baar koi solid plan nahi diya. Thoda aur detail me puchiye?"
        
        if heuristic_plan.action != "noop":
            workbook_service.remember_plan(session_id, command, heuristic_plan)
        return heuristic_plan

    def _undo_plan(self, snapshot: WorkbookSnapshot) -> ActionPlan:
        return ActionPlan(
            action="undo",
            target_sheet=snapshot.active_sheet,
            preview_title="Undo last action",
            explanation="Revert the workbook to its previous state.",
            risk_level="low",
            requires_confirmation=False,
            parameters={},
            impact=ActionImpact(summary="The last change will be undone.", estimated_rows=0, estimated_cells=0)
        )

    def _redo_plan(self, snapshot: WorkbookSnapshot) -> ActionPlan:
        return ActionPlan(
            action="redo",
            target_sheet=snapshot.active_sheet,
            preview_title="Redo action",
            explanation="Re-apply the change that was just undone.",
            risk_level="low",
            requires_confirmation=False,
            parameters={},
            impact=ActionImpact(summary="The undone change will be re-applied.", estimated_rows=0, estimated_cells=0)
        )

    async def _preview_with_gemini(
        self,
        command: str,
        snapshot: WorkbookSnapshot,
        selected_cell: str | None = None,
        selected_value: Any = None,
    ) -> ActionPlan:
        prompt = self._build_prompt(command, snapshot, selected_cell, selected_value)
        url = (
            "https://generativelanguage.googleapis.com/v1beta/models/"
            f"{settings.gemini_model}:generateContent?key={settings.gemini_api_key}"
        )
        payload = {
            "contents": [{"role": "user", "parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": 0.1, "responseMimeType": "application/json"},
        }
        async with httpx.AsyncClient(timeout=25.0) as client:
            response = await client.post(url, json=payload)
        response.raise_for_status()
        data = response.json()
        candidates = data.get("candidates") or []
        if not candidates:
            raise HTTPException(status_code=502, detail="Gemini returned no candidates.")
        text = candidates[0]["content"]["parts"][0].get("text", "")
        
        # Robust JSON extraction
        try:
            start = text.find('{')
            end = text.rfind('}')
            if start != -1 and end != -1:
                json_str = text[start:end+1]
                return ActionPlan.model_validate_json(json_str)
            else:
                # Try raw text as last resort
                return ActionPlan.model_validate_json(text)
        except Exception as e:
            import sys
            print(f"Pydantic Validation Error: {repr(e)}\nRaw Text: {text}", file=sys.stderr)
            raise HTTPException(status_code=500, detail=f"AI Agent failed to format response: {repr(e)}")

    def _preview_with_heuristics(
        self,
        command: str,
        snapshot: WorkbookSnapshot,
        selected_cell: str | None = None,
        selected_value: Any = None,
        allow_batch: bool = True,
    ) -> ActionPlan:
        command_lower = command.lower().strip()
        target_sheet = self._find_target_sheet(command_lower, snapshot)
        sheet = self._sheet_by_name(snapshot, target_sheet)
        selected_header = self._selected_header_for_cell(sheet, selected_cell)
        if selected_header is None and selected_value not in (None, "") and selected_cell:
            selected_header = self._selected_header_from_value(sheet, selected_value)
        formula_text = self._extract_formula_text(selected_value)
        matched_headers = self._find_matching_headers(command_lower, sheet["headers"])
        matched_header = self._find_header_in_command(command_lower, sheet["headers"]) or selected_header
        numeric_header = self._pick_numeric_header(sheet, matched_header)
        text_header = self._pick_text_header(sheet, matched_header)

        if allow_batch:
            batch_plan = self._preview_batch_workflow(command_lower, snapshot, selected_cell, selected_value)
            if batch_plan is not None:
                return batch_plan

        follow_up_plan = self._preview_follow_up_workflow(command_lower, snapshot, selected_cell, selected_value)
        if follow_up_plan is not None:
            return follow_up_plan

        formula_plan = self._preview_formula_workflow(
            command_lower,
            snapshot,
            target_sheet,
            sheet,
            selected_cell,
            selected_header,
            formula_text,
        )
        if formula_plan is not None:
            return formula_plan

        table_plan = self._preview_formatting_workflow(
            command_lower,
            snapshot,
            target_sheet,
            sheet,
            matched_header,
            numeric_header,
            text_header,
        )
        if table_plan is not None:
            return table_plan

        editing_plan = self._preview_editing_workflow(
            command_lower,
            snapshot,
            target_sheet,
            sheet,
            selected_cell,
            matched_header,
            numeric_header,
            text_header,
        )
        if editing_plan is not None:
            return editing_plan

        analysis_plan = self._preview_analysis_workflow(command_lower, snapshot, target_sheet)
        if analysis_plan is not None:
            return analysis_plan

        if any(token in command_lower for token in ["clear filter", "reset filter", "show all"]):
            return ActionPlan(
                action="clear_filter",
                target_sheet=target_sheet,
                preview_title="Clear active filter",
                explanation=f"I will unhide all rows and clear the active filter on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=f"A1:{self._col_letter(sheet['max_column'])}{sheet['max_row']}",
                parameters={},
                impact=ActionImpact(
                    summary=f"All {max(0, sheet['max_row'] - 1)} data rows will become visible again.",
                    estimated_rows=max(0, sheet["max_row"] - 1),
                    estimated_cells=max(0, sheet["max_row"] - 1) * max(1, sheet["max_column"]),
                ),
            )

        if any(
            token in command_lower
            for token in ["convert", "as number", "as text", "as date", "date format", "number format", "numeric"]
        ):
            target_type = self._extract_target_type(command_lower)
            conversion_header = matched_header or text_header or numeric_header
            if target_type and conversion_header:
                convertible, skipped = workbook_service.profile_column_conversion(
                    snapshot.session_id,
                    target_sheet,
                    conversion_header,
                    target_type,
                )
                return ActionPlan(
                    action="convert_column_type",
                    target_sheet=target_sheet,
                    target_column=conversion_header,
                    preview_title="Convert column type",
                    explanation=f"I will convert {conversion_header} on {target_sheet} into {target_type} values where possible.",
                    risk_level="medium" if skipped else "low",
                    requires_confirmation=bool(skipped),
                    impacted_range=f"{workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, conversion_header)}2:{workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, conversion_header)}{sheet['max_row']}",
                    parameters={
                        "target_type": target_type,
                        "date_format": "DD-MMM-YYYY" if target_type == "date" else "",
                    },
                    impact=ActionImpact(
                        summary=f"{convertible} cells can be converted and {skipped} may stay unchanged.",
                        estimated_rows=convertible,
                        estimated_cells=convertible,
                        warnings=(
                            ["Some cells do not match the requested type and will stay as-is."]
                            if skipped
                            else []
                        ),
                    ),
                )

        if any(token in command_lower for token in ["duplicate", "dedup"]):
            duplicate_count = workbook_service.count_duplicate_rows(
                snapshot.session_id,
                target_sheet,
                [matched_header] if matched_header else None,
            )
            return ActionPlan(
                action="delete_duplicates",
                target_sheet=target_sheet,
                target_column=matched_header,
                preview_title="Remove duplicate rows",
                explanation=f"I will remove duplicate rows from {target_sheet} using {matched_header or 'the full row'} as the uniqueness key.",
                risk_level="high",
                requires_confirmation=True,
                impacted_range=f"A2:{self._col_letter(sheet['max_column'])}{sheet['max_row']}",
                parameters={"columns": [matched_header] if matched_header else []},
                impact=ActionImpact(
                    summary=f"{duplicate_count} duplicate rows are likely to be removed.",
                    estimated_rows=duplicate_count,
                    estimated_cells=duplicate_count * max(1, sheet["max_column"]),
                    warnings=["Duplicate removal rewrites the visible table and should be confirmed before applying."],
                ),
            )

        if "replace" in command_lower and "with" in command_lower:
            match = re.search(r"replace\s+(.+?)\s+with\s+(.+)", command_lower)
            if match:
                find_text = match.group(1).strip()
                replace_text = match.group(2).strip()
                hits = workbook_service.count_find_replace_hits(snapshot.session_id, target_sheet, find_text, matched_header)
                return ActionPlan(
                    action="find_replace",
                    target_sheet=target_sheet,
                    target_column=matched_header,
                    preview_title="Find and replace values",
                    explanation=f"I will replace '{find_text}' with '{replace_text}' on {target_sheet}.",
                    risk_level="medium",
                    requires_confirmation=True,
                    impacted_range=f"A1:{self._col_letter(sheet['max_column'])}{sheet['max_row']}",
                    parameters={"find": find_text, "replace": replace_text},
                    impact=ActionImpact(
                        summary=f"About {hits} text matches will be updated.",
                        estimated_rows=hits,
                        estimated_cells=hits,
                        warnings=["Text replacements are applied directly to matching cells."],
                    ),
                )

        if any(token in command_lower for token in ["sort", "ascending", "descending"]):
            descending = "desc" in command_lower or "descending" in command_lower or "z to a" in command_lower
            sort_header = matched_header or sheet["headers"][0]
            return ActionPlan(
                action="sort",
                target_sheet=target_sheet,
                target_column=sort_header,
                preview_title="Sort sheet",
                explanation=f"I will sort {target_sheet} by {sort_header} in {'descending' if descending else 'ascending'} order.",
                risk_level="medium",
                requires_confirmation=False,
                impacted_range=f"A2:{self._col_letter(sheet['max_column'])}{sheet['max_row']}",
                parameters={"descending": descending},
                impact=ActionImpact(
                    summary=f"{max(0, sheet['max_row'] - 1)} rows will be reordered.",
                    estimated_rows=max(0, sheet["max_row"] - 1),
                    estimated_cells=max(0, sheet["max_row"] - 1) * max(1, sheet["max_column"]),
                ),
            )

        if any(token in command_lower for token in ["filter", "show only"]) or (
            "where" in command_lower
            and not any(token in command_lower for token in ["sum", "total", "average", "avg", "count", "sumif", "averageif", "avgif", "countif"])
        ):
            filter_header = numeric_header or matched_header or text_header
            operator, criterion = self._extract_filter_criterion(command_lower)
            matched, total = workbook_service.count_filter_matches(snapshot.session_id, target_sheet, filter_header, operator, criterion)
            return ActionPlan(
                action="apply_filter",
                target_sheet=target_sheet,
                target_column=filter_header,
                preview_title="Apply smart filter",
                explanation=f"I will filter {target_sheet} so only rows matching {filter_header} {operator} {criterion} remain visible.",
                risk_level="medium",
                requires_confirmation=False,
                impacted_range=f"A1:{self._col_letter(sheet['max_column'])}{sheet['max_row']}",
                parameters={"operator": operator, "value": criterion},
                impact=ActionImpact(
                    summary=f"{matched} of {total} rows will remain visible after filtering.",
                    estimated_rows=matched,
                    estimated_cells=matched * max(1, sheet["max_column"]),
                    warnings=["Filtered rows are hidden, not deleted."],
                ),
            )

        if any(token in command_lower for token in ["chart", "graph"]):
            chart_recommendation = self._best_chart_recommendation(snapshot, target_sheet, matched_header, numeric_header, text_header, command_lower)
            chart_header = chart_recommendation.value_column if chart_recommendation else numeric_header
            if chart_header is None:
                return self._unsupported_plan(target_sheet)
            chart_type = self._infer_chart_type(command_lower, chart_recommendation)
            target_cell = f"{workbook_service.next_available_column_letter(snapshot.session_id, target_sheet, 2)}2"
            category_header = (
                chart_recommendation.category_column
                if chart_recommendation and chart_recommendation.category_column
                else text_header or sheet["headers"][0]
            )
            return ActionPlan(
                action="create_chart",
                target_sheet=target_sheet,
                target_cell=target_cell,
                target_column=chart_header,
                preview_title="Create chart",
                explanation=(
                    f"I will create a {chart_type} chart for {chart_header} on {target_sheet}."
                    if chart_recommendation is None
                    else f"I will create a {chart_type} chart for {chart_recommendation.value_column} using {chart_recommendation.category_column}."
                ),
                risk_level="low",
                requires_confirmation=False,
                impacted_range=target_cell,
                parameters={
                    "chart_type": chart_type,
                    "category_column": category_header,
                    "output_sheet": f"{target_sheet}_Insights",
                    "title": f"{chart_header} {chart_type.title()} Chart",
                },
                impact=ActionImpact(
                    summary=f"A new chart will be added using {max(0, sheet['max_row'] - 1)} data rows.",
                    estimated_rows=max(0, sheet["max_row"] - 1),
                    estimated_cells=max(0, sheet["max_row"] - 1),
                ),
            )

        if any(token in command_lower for token in ["recommend chart", "best chart", "suggest chart", "which chart", "chart recommendation"]):
            recommendation = self._best_chart_recommendation(snapshot, target_sheet, matched_header, numeric_header, text_header, command_lower)
            if recommendation is None:
                return ActionPlan(
                    action="recommend_chart",
                    target_sheet=target_sheet,
                    preview_title="Recommend chart",
                    explanation="I need at least one numeric column to recommend a useful chart.",
                    risk_level="low",
                    requires_confirmation=False,
                    impacted_range=None,
                    parameters={"recommendations": []},
                    impact=ActionImpact(summary="No chart recommendation could be generated."),
                )
            return ActionPlan(
                action="recommend_chart",
                target_sheet=target_sheet,
                preview_title="Recommend chart",
                explanation=(
                    f"Best fit: {recommendation.chart_type} chart for {recommendation.value_column} by {recommendation.category_column}."
                ),
                risk_level="low",
                requires_confirmation=False,
                impacted_range=None,
                parameters={"recommendations": [recommendation.model_dump()]},
                impact=ActionImpact(
                    summary=recommendation.detail,
                    estimated_rows=0,
                    estimated_cells=0,
                ),
            )

        threshold_match = re.search(r"(?:>=|>|above|greater than)\s*(\d+(?:\.\d+)?)", command_lower)
        if any(token in command_lower for token in ["highlight", "green", "format"]) and threshold_match and numeric_header:
            threshold = float(threshold_match.group(1))
            hits = workbook_service.count_threshold_hits(snapshot.session_id, target_sheet, numeric_header, threshold)
            column_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, numeric_header)
            return ActionPlan(
                action="highlight_threshold",
                target_sheet=target_sheet,
                target_column=numeric_header,
                preview_title="Highlight important values",
                explanation=f"I will highlight {numeric_header} values greater than {threshold} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=f"{column_letter}2:{column_letter}{sheet['max_row']}",
                parameters={"threshold": threshold, "color": "C7F9CC"},
                impact=ActionImpact(
                    summary=f"{hits} cells are expected to match the threshold rule.",
                    estimated_rows=hits,
                    estimated_cells=hits,
                ),
            )

        if (
            any(token in command_lower for token in ["sumif", "averageif", "avgif", "countif"])
            or any(token in command_lower for token in [" where ", " jahan ", " when ", " contains ", " includes "])
        ):
            operator, criterion = self._extract_filter_criterion(command_lower)
            formula_kind = "count" if any(token in command_lower for token in ["countif", "count "]) else (
                "average" if any(token in command_lower for token in ["averageif", "avgif", "average", "avg"]) else "sum"
            )
            value_header = next(
                (header for header in matched_headers if header in sheet.get("numeric_headers", [])),
                numeric_header,
            )
            condition_header = self._pick_condition_header(sheet, matched_headers, value_header, operator)
            condition_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, condition_header)
            if condition_letter is None:
                return self._unsupported_plan(target_sheet)
            value_letter = (
                workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, value_header)
                if value_header
                else None
            )
            if formula_kind in {"sum", "average"} and value_letter is None:
                return self._unsupported_plan(target_sheet)

            target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
            target_cell = workbook_service.find_first_empty_cell(snapshot.session_id, target_sheet, target_column_letter)
            criteria = self._excel_criteria(operator, criterion).replace('"', '""')
            condition_range = f"{condition_letter}2:{condition_letter}{sheet['max_row']}"
            if formula_kind == "count":
                formula = f'=COUNTIF({condition_range},"{criteria}")'
                preview_title = "Insert COUNTIF formula"
            elif formula_kind == "average":
                formula = f'=AVERAGEIF({condition_range},"{criteria}",{value_letter}2:{value_letter}{sheet["max_row"]})'
                preview_title = "Insert AVERAGEIF formula"
            else:
                formula = f'=SUMIF({condition_range},"{criteria}",{value_letter}2:{value_letter}{sheet["max_row"]})'
                preview_title = "Insert SUMIF formula"
            return ActionPlan(
                action="insert_formula",
                target_sheet=target_sheet,
                target_cell=target_cell,
                target_column=value_header or condition_header,
                formula=formula,
                preview_title=preview_title,
                explanation=(
                    f"I will insert a {formula_kind.upper()} formula for {target_sheet} using "
                    f"{condition_header} {operator} {criterion}."
                ),
                risk_level="low",
                requires_confirmation=False,
                impacted_range=target_cell,
                parameters={},
                impact=ActionImpact(summary="One conditional formula cell will be inserted.", estimated_rows=1, estimated_cells=1),
            )

        if any(token in command_lower for token in ["sum", "total"]):
            if numeric_header is None:
                return self._unsupported_plan(target_sheet)
            column_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, numeric_header)
            target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
            target_cell = workbook_service.find_first_empty_cell(snapshot.session_id, target_sheet, target_column_letter)
            return ActionPlan(
                action="insert_formula",
                target_sheet=target_sheet,
                target_cell=target_cell,
                target_column=numeric_header,
                formula=f"=SUM({column_letter}2:{column_letter}{sheet['max_row']})",
                preview_title="Insert total formula",
                explanation=f"I will insert a SUM formula for {numeric_header} into {target_cell}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=target_cell,
                parameters={},
                impact=ActionImpact(summary="One formula cell will be inserted.", estimated_rows=1, estimated_cells=1),
            )

        if "average" in command_lower or "avg" in command_lower:
            if numeric_header is None:
                return self._unsupported_plan(target_sheet)
            column_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, numeric_header)
            target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
            target_cell = workbook_service.find_first_empty_cell(snapshot.session_id, target_sheet, target_column_letter)
            return ActionPlan(
                action="insert_formula",
                target_sheet=target_sheet,
                target_cell=target_cell,
                target_column=numeric_header,
                formula=f"=AVERAGE({column_letter}2:{column_letter}{sheet['max_row']})",
                preview_title="Insert average formula",
                explanation=f"I will insert an AVERAGE formula for {numeric_header} into {target_cell}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=target_cell,
                parameters={},
                impact=ActionImpact(summary="One formula cell will be inserted.", estimated_rows=1, estimated_cells=1),
            )

        if "vlookup" in command_lower and len(snapshot.sheets) > 1:
            source_sheet = next(sheet_item for sheet_item in snapshot.sheets if sheet_item.name != target_sheet)
            target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
            return ActionPlan(
                action="fill_formula_down",
                target_sheet=target_sheet,
                target_column=target_column_letter,
                formula=f'=IFERROR(VLOOKUP(A{{row}},\'{source_sheet.name}\'!A:B,2,FALSE),"")',
                preview_title="Fill VLOOKUP column",
                explanation=f"I will fill a VLOOKUP formula on {target_sheet} using {source_sheet.name} as the source table.",
                risk_level="medium",
                requires_confirmation=True,
                impacted_range=f"{target_column_letter}2:{target_column_letter}{sheet['max_row']}",
                parameters={
                    "target_column_letter": target_column_letter,
                    "start_row": 2,
                    "end_row": max(2, sheet["max_row"]),
                },
                impact=ActionImpact(
                    summary=f"A lookup formula will be filled across {max(0, sheet['max_row'] - 1)} rows.",
                    estimated_rows=max(0, sheet["max_row"] - 1),
                    estimated_cells=max(0, sheet["max_row"] - 1),
                    warnings=["Cross-sheet formulas should be reviewed before applying on production files."],
                ),
            )

        if any(token in command_lower for token in ["verify", "match"]) and len(snapshot.sheets) > 1:
            source_sheet = next(sheet_item for sheet_item in snapshot.sheets if sheet_item.name != target_sheet)
            target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
            return ActionPlan(
                action="fill_formula_down",
                target_sheet=target_sheet,
                target_column=target_column_letter,
                formula=f'=IF(A{{row}}=\'{source_sheet.name}\'!A{{row}},"Match","No Match")',
                preview_title="Verify row matches",
                explanation=f"I will fill a match verification formula on {target_sheet} against {source_sheet.name}.",
                risk_level="medium",
                requires_confirmation=True,
                impacted_range=f"{target_column_letter}2:{target_column_letter}{sheet['max_row']}",
                parameters={
                    "target_column_letter": target_column_letter,
                    "start_row": 2,
                    "end_row": max(2, sheet["max_row"]),
                },
                impact=ActionImpact(
                    summary=f"A verification formula will be filled across {max(0, sheet['max_row'] - 1)} rows.",
                    estimated_rows=max(0, sheet["max_row"] - 1),
                    estimated_cells=max(0, sheet["max_row"] - 1),
                ),
            )

        if any(token in command_lower for token in ["pivot", "summary by", "group by", "summarize"]):
            group_header = text_header or sheet["headers"][0]
            val_header = numeric_header or (sheet["headers"][1] if len(sheet["headers"]) > 1 else sheet["headers"][0])
            return ActionPlan(
                action="create_pivot",
                target_sheet=target_sheet,
                target_column=group_header,
                preview_title="Create Pivot Table",
                explanation=f"I will create a pivot table summarizing {val_header} grouped by {group_header}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range="A1:B10",
                parameters={"group_by": group_header, "value_col": val_header, "agg": "sum"},
                impact=ActionImpact(
                    summary="A new sheet will be created with the summarized pivot data.",
                    estimated_rows=10,
                    estimated_cells=20,
                )
            )

        if any(token in command_lower for token in ["add sheet", "create sheet", "new sheet"]):
            new_sheet_name = "New_Sheet"
            match = re.search(r"(?:called|named|as) ([a-z0-9 _-]+)", command_lower)
            if match:
                new_sheet_name = match.group(1).title().strip()
            
            return ActionPlan(
                action="add_sheet",
                target_sheet=target_sheet,
                preview_title="Add new sheet",
                explanation=f"I will create a new sheet named {new_sheet_name}.",
                risk_level="low",
                requires_confirmation=False,
                parameters={"new_sheet_name": new_sheet_name},
                impact=ActionImpact(
                    summary=f"A new blank sheet '{new_sheet_name}' will be added.",
                )
            )

        if any(token in command_lower for token in ["join", "merge", "combine"]) and len(snapshot.sheets) > 1:
            source_sheet = next(sheet_item for sheet_item in snapshot.sheets if sheet_item.name != target_sheet)
            join_header = text_header or sheet["headers"][0]
            
            return ActionPlan(
                action="join_sheets",
                target_sheet=target_sheet,
                target_column=join_header,
                preview_title="Join Sheets",
                explanation=f"I will join data from {source_sheet.name} into {target_sheet} using {join_header}.",
                risk_level="medium",
                requires_confirmation=True,
                parameters={"secondary_sheet": source_sheet.name, "join_column": join_header},
                impact=ActionImpact(
                    summary=f"Data will be merged. This could add several new columns from {source_sheet.name}.",
                    estimated_rows=sheet["max_row"],
                    estimated_cells=sheet["max_row"] * 5,
                    warnings=["Merging sheets can be heavy. Review before applying."]
                )
            )

        return self._unsupported_plan(target_sheet)

    def _preview_formula_workflow(
        self,
        command_lower: str,
        snapshot: WorkbookSnapshot,
        target_sheet: str,
        sheet: dict[str, Any],
        selected_cell: str | None,
        selected_header: str | None,
        formula_text: str | None,
    ) -> ActionPlan | None:
        explain_intent = any(
            token in command_lower
            for token in [
                "explain formula",
                "what does this formula do",
                "explain this formula",
                "samjhao",
                "formula explain",
                "formula ka matlab",
            ]
        )
        fix_intent = any(
            token in command_lower
            for token in [
                "fix formula",
                "correct formula",
                "repair formula",
                "formula error",
                "broken formula",
                "formula ko fix",
                "formula sudharo",
            ]
        )
        generate_intent = any(
            token in command_lower
            for token in [
                "generate formula",
                "create formula",
                "write formula",
                "make formula",
                "formula banao",
                "formula do",
                "formula suggest",
            ]
        )

        if not (explain_intent or fix_intent or generate_intent):
            return None

        current_formula = formula_text if formula_text and formula_text.startswith("=") else None
        current_cell = selected_cell or f"{target_sheet}!A1"

        if explain_intent:
            if not current_formula:
                return ActionPlan(
                    action="explain_formula",
                    target_sheet=target_sheet,
                    target_cell=selected_cell,
                    target_column=selected_header,
                    preview_title="Explain formula",
                    explanation="Select a cell containing a formula so I can explain it step by step.",
                    risk_level="low",
                    requires_confirmation=False,
                    impacted_range=selected_cell,
                    parameters={"workflow": "explain", "formula": None},
                    impact=ActionImpact(summary="No workbook changes will be made."),
                )
            return ActionPlan(
                action="explain_formula",
                target_sheet=target_sheet,
                target_cell=selected_cell,
                target_column=selected_header,
                formula=current_formula,
                preview_title="Explain formula",
                explanation=self._explain_formula(current_formula, sheet, selected_header, selected_cell),
                risk_level="low",
                requires_confirmation=False,
                impacted_range=selected_cell,
                parameters={
                    "workflow": "explain",
                    "formula": current_formula,
                    "header": selected_header,
                },
                impact=ActionImpact(summary="Formula explanation ready.", estimated_rows=0, estimated_cells=0),
            )

        if fix_intent:
            if not current_formula:
                return ActionPlan(
                    action="fix_formula",
                    target_sheet=target_sheet,
                    target_cell=selected_cell,
                    target_column=selected_header,
                    preview_title="Fix formula",
                    explanation="Select a formula cell first so I can fix it.",
                    risk_level="low",
                    requires_confirmation=False,
                    impacted_range=selected_cell,
                    parameters={"workflow": "fix", "formula": None},
                    impact=ActionImpact(summary="No workbook changes will be made."),
                )
            suggested_formula = self._fix_formula(current_formula, sheet, selected_header)
            return ActionPlan(
                action="fix_formula",
                target_sheet=target_sheet,
                target_cell=selected_cell,
                target_column=selected_header,
                formula=suggested_formula,
                preview_title="Fix formula",
                explanation=(
                    f"I will replace the formula in {current_cell} with a safer version: {suggested_formula}"
                    if suggested_formula != current_formula
                    else f"I checked {current_cell} and did not find an obvious syntax issue."
                ),
                risk_level="medium" if suggested_formula != current_formula else "low",
                requires_confirmation=suggested_formula != current_formula,
                impacted_range=selected_cell,
                parameters={
                    "workflow": "fix",
                    "original_formula": current_formula,
                    "suggested_formula": suggested_formula,
                    "header": selected_header,
                },
                impact=ActionImpact(
                    summary=(
                        "A corrected formula will replace the selected cell."
                        if suggested_formula != current_formula
                        else "No change required."
                    ),
                    estimated_rows=1 if suggested_formula != current_formula else 0,
                    estimated_cells=1 if suggested_formula != current_formula else 0,
                ),
            )

        if generate_intent:
            generated = self._generate_formula_plan(
                command_lower,
                snapshot,
                target_sheet,
                sheet,
                selected_cell,
                selected_header,
                current_formula,
            )
            if generated is not None:
                return generated

        return ActionPlan(
            action="noop",
            target_sheet=target_sheet,
            preview_title="Formula help",
            explanation="I need a clearer formula request or a selected formula cell to help with.",
            risk_level="low",
            requires_confirmation=False,
            parameters={},
            impact=ActionImpact(summary="No workbook changes will be made."),
        )

    def _preview_formatting_workflow(
        self,
        command_lower: str,
        snapshot: WorkbookSnapshot,
        target_sheet: str,
        sheet: dict[str, Any],
        matched_header: str | None,
        numeric_header: str | None,
        text_header: str | None,
    ) -> ActionPlan | None:
        if any(token in command_lower for token in ["freeze header", "freeze top row", "freeze panes", "freeze row"]):
            return ActionPlan(
                action="freeze_header",
                target_sheet=target_sheet,
                preview_title="Freeze header row",
                explanation=f"I will freeze the top row of {target_sheet} so headers stay visible while scrolling.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range="A1",
                parameters={},
                impact=ActionImpact(summary="The header row will stay visible during scrolling.", estimated_rows=1, estimated_cells=0),
            )

        if any(token in command_lower for token in ["auto fit", "autofit", "fit columns", "adjust width", "resize columns"]):
            return ActionPlan(
                action="auto_fit_columns",
                target_sheet=target_sheet,
                preview_title="Auto-fit columns",
                explanation=f"I will resize columns on {target_sheet} to better fit the visible values.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=f"A1:{self._col_letter(sheet['max_column'])}{sheet['max_row']}",
                parameters={},
                impact=ActionImpact(summary="Column widths will be adjusted for readability.", estimated_rows=0, estimated_cells=0),
            )

        if any(token in command_lower for token in ["make table", "create table", "format as table", "table style"]):
            header_row = workbook_service.detect_header_row(snapshot.session_id, target_sheet)
            return ActionPlan(
                action="create_table",
                target_sheet=target_sheet,
                preview_title="Create Excel table",
                explanation=f"I will convert the used range on {target_sheet} into an Excel table with banded rows and filter controls.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=f"A{header_row}:{self._col_letter(sheet['max_column'])}{sheet['max_row']}",
                parameters={
                    "table_style": "TableStyleMedium2",
                    "table_name": f"{target_sheet.replace(' ', '')[:20]}Table",
                },
                impact=ActionImpact(summary="The range will become a structured Excel table.", estimated_rows=max(0, sheet["max_row"] - 1), estimated_cells=max(0, sheet["max_row"] * sheet["max_column"])),
            )

        if any(token in command_lower for token in ["bold header", "header style", "format header", "highlight header"]):
            header_row = workbook_service.detect_header_row(snapshot.session_id, target_sheet)
            return ActionPlan(
                action="format_header",
                target_sheet=target_sheet,
                preview_title="Format header row",
                explanation=f"I will apply a stronger header style to the top row on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=f"A{header_row}:{self._col_letter(sheet['max_column'])}{header_row}",
                parameters={
                    "fill": "107C41",
                    "font_color": "FFFFFF",
                    "bold": True,
                },
                impact=ActionImpact(summary="Header cells will get a clean bold style.", estimated_rows=1, estimated_cells=sheet["max_column"]),
            )

        if any(token in command_lower for token in ["currency format", "format as currency", "rupees format", "number format", "date format", "percentage format", "percent format"]):
            target_column = matched_header or numeric_header or text_header
            if target_column is None:
                return None
            number_format = self._detect_number_format(command_lower)
            if number_format is None:
                return None
            column_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, target_column)
            if column_letter is None:
                return None
            return ActionPlan(
                action="format_number",
                target_sheet=target_sheet,
                target_column=target_column,
                preview_title="Format numbers",
                explanation=f"I will format {target_column} on {target_sheet} using {number_format}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=f"{column_letter}2:{column_letter}{sheet['max_row']}",
                parameters={"number_format": number_format},
                impact=ActionImpact(summary="Number formatting will be applied to the selected column.", estimated_rows=max(0, sheet["max_row"] - 1), estimated_cells=max(0, sheet["max_row"] - 1)),
            )

        return None

    def _preview_editing_workflow(
        self,
        command_lower: str,
        snapshot: WorkbookSnapshot,
        target_sheet: str,
        sheet: dict[str, Any],
        selected_cell: str | None,
        matched_header: str | None,
        numeric_header: str | None,
        text_header: str | None,
    ) -> ActionPlan | None:
        rename_request = self._extract_rename_request(command_lower)
        if rename_request is not None:
            new_name = self._clean_sheet_name(rename_request)
            return ActionPlan(
                action="rename_sheet",
                target_sheet=target_sheet,
                preview_title="Rename sheet",
                explanation=f"I will rename {target_sheet} to {new_name}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=None,
                parameters={"new_name": new_name},
                impact=ActionImpact(summary=f"Sheet tab will be renamed to {new_name}."),
            )

        row_span = self._extract_row_span(command_lower)
        col_span = self._extract_column_span(command_lower)
        cell_range = self._extract_cell_range(command_lower)

        if re.search(r"\b(insert|add)\b.*\brows?\b", command_lower):
            amount_match = re.search(r"(?:insert|add)\s+(\d+)\s+rows?", command_lower)
            row_index_match = self._extract_row_anchor(command_lower)
            row_index = (
                int(row_index_match.group(1))
                if row_index_match
                else (row_span[0] if row_span else max(2, workbook_service.data_start_row(snapshot.session_id, target_sheet)))
            )
            amount = int(amount_match.group(1)) if amount_match else (row_span[1] - row_span[0] + 1 if row_span else 1)
            return ActionPlan(
                action="insert_rows",
                target_sheet=target_sheet,
                preview_title="Insert rows",
                explanation=f"I will insert {amount} row(s) near row {row_index} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=f"row {row_index}",
                parameters={"row_index": row_index, "amount": amount},
                impact=ActionImpact(summary=f"{amount} blank row(s) will be inserted.", estimated_rows=amount, estimated_cells=0),
            )

        if re.search(r"\b(delete|remove|trash)\b.*\brows?\b", command_lower):
            amount_match = re.search(r"(?:delete|remove)\s+(\d+)\s+rows?", command_lower)
            row_index_match = self._extract_row_anchor(command_lower)
            row_index = (
                int(row_index_match.group(1))
                if row_index_match
                else (row_span[0] if row_span else max(2, workbook_service.data_start_row(snapshot.session_id, target_sheet)))
            )
            amount = int(amount_match.group(1)) if amount_match else (row_span[1] - row_span[0] + 1 if row_span else 1)
            return ActionPlan(
                action="delete_rows",
                target_sheet=target_sheet,
                preview_title="Delete rows",
                explanation=f"I will delete {amount} row(s) starting at row {row_index} on {target_sheet}.",
                risk_level="high",
                requires_confirmation=True,
                impacted_range=f"row {row_index}",
                parameters={"row_index": row_index, "amount": amount},
                impact=ActionImpact(summary=f"{amount} row(s) will be removed.", estimated_rows=amount, estimated_cells=amount * max(1, sheet["max_column"]), warnings=["Row deletion is destructive."]),
            )

        if re.search(r"\b(insert|add)\b.*\bcolumns?\b", command_lower):
            amount_match = re.search(r"(?:insert|add)\s+(\d+)\s+columns?", command_lower)
            column_letter_match = self._extract_column_anchor(command_lower)
            column_letter = (
                column_letter_match.group(1).upper()
                if column_letter_match
                else (col_span[0] if col_span else (matched_header and workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, matched_header)) or "B")
            )
            amount = int(amount_match.group(1)) if amount_match else (self._column_span_width(col_span) if col_span else 1)
            return ActionPlan(
                action="insert_columns",
                target_sheet=target_sheet,
                preview_title="Insert columns",
                explanation=f"I will insert {amount} column(s) near {column_letter} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=column_letter,
                parameters={"column_letter": column_letter, "amount": amount},
                impact=ActionImpact(summary=f"{amount} blank column(s) will be inserted.", estimated_rows=0, estimated_cells=0),
            )

        if re.search(r"\b(delete|remove|trash)\b.*\bcolumns?\b", command_lower):
            amount_match = re.search(r"(?:delete|remove)\s+(\d+)\s+columns?", command_lower)
            column_letter_match = self._extract_column_anchor(command_lower)
            column_letter = (
                column_letter_match.group(1).upper()
                if column_letter_match
                else (col_span[0] if col_span else (matched_header and workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, matched_header)) or "B")
            )
            amount = int(amount_match.group(1)) if amount_match else (self._column_span_width(col_span) if col_span else 1)
            return ActionPlan(
                action="delete_columns",
                target_sheet=target_sheet,
                preview_title="Delete columns",
                explanation=f"I will delete {amount} column(s) starting at {column_letter} on {target_sheet}.",
                risk_level="high",
                requires_confirmation=True,
                impacted_range=column_letter,
                parameters={"column_letter": column_letter, "amount": amount},
                impact=ActionImpact(summary=f"{amount} column(s) will be removed.", estimated_rows=0, estimated_cells=0, warnings=["Column deletion is destructive."]),
            )

        if any(token in command_lower for token in ["clear cells", "clear contents", "clear values", "wipe cells", "erase cells"]):
            ref = cell_range or self._cell_range_from_selected(selected_cell) or f"A2:{self._col_letter(sheet['max_column'])}{sheet['max_row']}"
            return ActionPlan(
                action="clear_cells",
                target_sheet=target_sheet,
                preview_title="Clear cells",
                explanation=f"I will clear the contents of {ref} on {target_sheet}.",
                risk_level="medium",
                requires_confirmation=True,
                impacted_range=ref,
                parameters={"range": ref},
                impact=ActionImpact(summary="Cell contents will be cleared without deleting rows or columns.", estimated_rows=0, estimated_cells=0),
            )

        if any(token in command_lower for token in ["merge cells", "merge range", "merge and center", "combine cells", "merge "]):
            ref = cell_range or self._cell_range_from_selected(selected_cell)
            if not ref:
                return None
            return ActionPlan(
                action="merge_cells",
                target_sheet=target_sheet,
                preview_title="Merge cells",
                explanation=f"I will merge {ref} on {target_sheet}.",
                risk_level="medium",
                requires_confirmation=True,
                impacted_range=ref,
                parameters={"range": ref},
                impact=ActionImpact(summary=f"{ref} will be merged into one cell.", estimated_rows=0, estimated_cells=0, warnings=["Merging cells may affect formulas and layout."]),
            )

        if any(token in command_lower for token in ["unmerge cells", "unmerge range", "split cells"]):
            ref = cell_range or self._cell_range_from_selected(selected_cell)
            if not ref:
                return None
            return ActionPlan(
                action="unmerge_cells",
                target_sheet=target_sheet,
                preview_title="Unmerge cells",
                explanation=f"I will unmerge {ref} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=ref,
                parameters={"range": ref},
                impact=ActionImpact(summary=f"{ref} will be unmerged.", estimated_rows=0, estimated_cells=0),
            )

        if any(token in command_lower for token in ["hide rows", "hide row", "collapse rows"]):
            start_row, end_row = row_span or (workbook_service.data_start_row(snapshot.session_id, target_sheet), workbook_service.data_start_row(snapshot.session_id, target_sheet))
            return ActionPlan(
                action="hide_rows",
                target_sheet=target_sheet,
                preview_title="Hide rows",
                explanation=f"I will hide rows {start_row} to {end_row} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=f"rows {start_row}-{end_row}",
                parameters={"start_row": start_row, "end_row": end_row},
                impact=ActionImpact(summary=f"Rows {start_row} to {end_row} will be hidden.", estimated_rows=end_row - start_row + 1, estimated_cells=0),
            )

        if any(token in command_lower for token in ["unhide rows", "show rows", "expand rows"]):
            start_row, end_row = row_span or (workbook_service.data_start_row(snapshot.session_id, target_sheet), workbook_service.data_start_row(snapshot.session_id, target_sheet))
            return ActionPlan(
                action="unhide_rows",
                target_sheet=target_sheet,
                preview_title="Unhide rows",
                explanation=f"I will unhide rows {start_row} to {end_row} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=f"rows {start_row}-{end_row}",
                parameters={"start_row": start_row, "end_row": end_row},
                impact=ActionImpact(summary=f"Rows {start_row} to {end_row} will become visible.", estimated_rows=end_row - start_row + 1, estimated_cells=0),
            )

        if any(token in command_lower for token in ["hide columns", "hide column", "collapse columns"]):
            start_col, end_col = col_span or ("B", "B")
            return ActionPlan(
                action="hide_columns",
                target_sheet=target_sheet,
                preview_title="Hide columns",
                explanation=f"I will hide columns {start_col} to {end_col} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=f"{start_col}:{end_col}",
                parameters={"start_column": start_col, "end_column": end_col},
                impact=ActionImpact(summary=f"Columns {start_col} to {end_col} will be hidden."),
            )

        if any(token in command_lower for token in ["unhide columns", "show columns", "expand columns"]):
            start_col, end_col = col_span or ("B", "B")
            return ActionPlan(
                action="unhide_columns",
                target_sheet=target_sheet,
                preview_title="Unhide columns",
                explanation=f"I will unhide columns {start_col} to {end_col} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=f"{start_col}:{end_col}",
                parameters={"start_column": start_col, "end_column": end_col},
                impact=ActionImpact(summary=f"Columns {start_col} to {end_col} will be visible again."),
            )

        if any(token in command_lower for token in ["comment", "note", "add note", "attach note"]):
            ref = selected_cell or cell_range or self._extract_cell_reference(command_lower)
            text_match = re.search(r"(?:comment|note)(?:\s+on|\s+to|\s+for)?\s+[a-z0-9:]+(?:\s+as|\s*:)?\s*(.+)", command_lower)
            text = text_match.group(1).strip() if text_match else "Review this cell."
            if not ref:
                return None
            return ActionPlan(
                action="add_comment",
                target_sheet=target_sheet,
                target_cell=ref.split(":")[0],
                preview_title="Add comment",
                explanation=f"I will add a comment to {ref} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=ref,
                parameters={"cell": ref.split(":")[0], "text": text, "author": "Copilot"},
                impact=ActionImpact(summary="A comment will be attached to the selected cell."),
            )

        if any(token in command_lower for token in ["hyperlink", "link", "add link"]):
            ref = selected_cell or cell_range or self._extract_cell_reference(command_lower)
            url_match = re.search(r"(https?://\S+)", command_lower)
            if not ref or not url_match:
                return None
            return ActionPlan(
                action="add_hyperlink",
                target_sheet=target_sheet,
                target_cell=ref.split(":")[0],
                preview_title="Add hyperlink",
                explanation=f"I will add a hyperlink to {ref} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=ref,
                parameters={"cell": ref.split(":")[0], "url": url_match.group(1), "text": ref.split(":")[0]},
                impact=ActionImpact(summary="A hyperlink will be added to the selected cell."),
            )

        if any(token in command_lower for token in ["validation", "validate", "data validation", "dropdown", "list validation", "restrict input"]):
            ref = cell_range or self._cell_range_from_selected(selected_cell)
            if not ref:
                return None
            if any(token in command_lower for token in ["list", "dropdown", "choices", "options"]):
                source_match = re.search(r"(?:from|as)\s+(.+)", command_lower)
                source = source_match.group(1).strip() if source_match else '"Yes,No"'
                return ActionPlan(
                    action="add_validation",
                    target_sheet=target_sheet,
                    preview_title="Add validation",
                    explanation=f"I will add list validation to {ref} on {target_sheet}.",
                    risk_level="low",
                    requires_confirmation=False,
                    impacted_range=ref,
                    parameters={"range": ref, "validation_type": "list", "source": source},
                    impact=ActionImpact(summary="Data validation will restrict cell input."),
                )
            return ActionPlan(
                action="add_validation",
                target_sheet=target_sheet,
                preview_title="Add validation",
                explanation=f"I will add validation to {ref} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=ref,
                parameters={"range": ref, "validation_type": "whole", "operator": "between", "minimum": 0, "maximum": 999999},
                impact=ActionImpact(summary="Data validation will be added."),
            )

        if any(token in command_lower for token in ["conditional format", "conditional formatting", "highlight", "color code", "colour code"]):
            ref = cell_range or self._cell_range_from_selected(selected_cell)
            threshold_match = re.search(r"(?:>=|>|above|greater than)\s*(\d+(?:\.\d+)?)", command_lower)
            if not ref or not threshold_match:
                return None
            return ActionPlan(
                action="conditional_format_range",
                target_sheet=target_sheet,
                preview_title="Conditional format range",
                explanation=f"I will apply conditional formatting to {ref} on {target_sheet}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=ref,
                parameters={"range": ref, "threshold": float(threshold_match.group(1)), "operator": "greaterThan", "color": "C7F9CC"},
                impact=ActionImpact(summary="A conditional formatting rule will be added."),
            )

        return None

    def _preview_analysis_workflow(
        self,
        command_lower: str,
        snapshot: WorkbookSnapshot,
        target_sheet: str,
    ) -> ActionPlan | None:
        if not any(token in command_lower for token in ["analyze", "analysis", "insight", "insights", "anomaly", "anomalies", "outlier", "trend", "overview", "review", "what is happening", "what's happening"]):
            return None

        insights = snapshot.insights[:3]
        anomalies = snapshot.anomalies[:3]
        recommendations = snapshot.chart_recommendations[:3]
        summary_bits: list[str] = []
        if insights:
            summary_bits.append(insights[0].detail)
        if anomalies:
            summary_bits.append(f"{len(anomalies)} anomaly signal(s) detected.")
        if recommendations:
            summary_bits.append(f"Top chart suggestion: {recommendations[0].chart_type} chart for {recommendations[0].value_column}.")
        if not summary_bits:
            summary_bits.append("Workbook analysis is ready, but no strong signals were detected.")

        return ActionPlan(
            action="analyze_workbook",
            target_sheet=target_sheet,
            preview_title="Workbook analysis",
            explanation=" ".join(summary_bits),
            risk_level="low",
            requires_confirmation=False,
            impacted_range=None,
            parameters={
                "insights": [item.model_dump() for item in insights],
                "anomalies": [item.model_dump() for item in anomalies],
                "chart_recommendations": [item.model_dump() for item in recommendations],
            },
            impact=ActionImpact(
                summary="Workbook analysis completed.",
                estimated_rows=0,
                estimated_cells=0,
            ),
        )

    def _best_chart_recommendation(
        self,
        snapshot: WorkbookSnapshot,
        target_sheet: str,
        matched_header: str | None,
        numeric_header: str | None,
        text_header: str | None,
        command_lower: str,
    ) -> ChartRecommendation | None:
        recommendations = [item for item in snapshot.chart_recommendations if item.sheet_name == target_sheet]
        if matched_header:
            recommendations = [
                item
                for item in recommendations
                if item.value_column == matched_header or item.category_column == matched_header
            ] or recommendations
        if "line" in command_lower:
            for recommendation in recommendations:
                if recommendation.chart_type == "line":
                    return recommendation
        if "scatter" in command_lower:
            for recommendation in recommendations:
                if recommendation.chart_type == "scatter":
                    return recommendation
        if "pie" in command_lower:
            for recommendation in recommendations:
                if recommendation.chart_type == "pie":
                    return recommendation
        if "bar" in command_lower or "column" in command_lower:
            for recommendation in recommendations:
                if recommendation.chart_type in {"bar", "column"}:
                    return recommendation
        if recommendations:
            return recommendations[0]

        if numeric_header is None:
            return None
        category_column = text_header or matched_header
        return ChartRecommendation(
            title="Fallback chart",
            detail="Fallback chart recommendation.",
            sheet_name=target_sheet,
            chart_type="bar",
            category_column=category_column,
            value_column=numeric_header,
            confidence="low",
        )

    @staticmethod
    def _infer_chart_type(command_lower: str, recommendation: Any | None) -> str:
        if "line" in command_lower:
            return "line"
        if "scatter" in command_lower:
            return "scatter"
        if "pie" in command_lower:
            return "pie"
        if recommendation is not None:
            return str(getattr(recommendation, "chart_type", "bar")).lower()
        return "bar"

    def _preview_batch_workflow(
        self,
        command_lower: str,
        snapshot: WorkbookSnapshot,
        selected_cell: str | None,
        selected_value: Any,
    ) -> ActionPlan | None:
        if any(token in command_lower for token in ["analyze", "analysis", "insight", "insights", "anomaly", "anomalies", "outlier", "trend"]):
            return None
        segments = self._split_multi_step_command(command_lower)
        if len(segments) < 2:
            return None

        steps: list[ActionPlan] = []
        for segment in segments:
            plan = self._preview_with_heuristics(
                segment,
                snapshot,
                selected_cell=selected_cell,
                selected_value=selected_value,
                allow_batch=False,
            )
            if plan.action == "noop":
                continue
            steps.append(plan)

        if len(steps) < 2:
            return None

        risk_rank = {"low": 0, "medium": 1, "high": 2}
        combined_risk = max(steps, key=lambda item: risk_rank.get(item.risk_level, 0)).risk_level
        summary = "; ".join(step.preview_title for step in steps[:3])
        if len(steps) > 3:
            summary += f" (+{len(steps) - 3} more)"
        total_rows = sum(step.impact.estimated_rows for step in steps)
        total_cells = sum(step.impact.estimated_cells for step in steps)
        warnings: list[str] = []
        for step in steps:
            warnings.extend(step.impact.warnings)

        return ActionPlan(
            action="batch",
            target_sheet=steps[0].target_sheet,
            preview_title=f"Run {len(steps)}-step task",
            explanation=f"I will run these steps in order: {summary}.",
            risk_level=combined_risk,
            requires_confirmation=any(step.requires_confirmation for step in steps),
            impacted_range=", ".join(step.impacted_range for step in steps if step.impacted_range) or None,
            parameters={"steps": [step.model_dump() for step in steps]},
            impact=ActionImpact(
                summary=f"{len(steps)} actions will run in sequence.",
                estimated_rows=total_rows,
                estimated_cells=total_cells,
                warnings=list(dict.fromkeys(warnings))[:5],
            ),
        )

    def _preview_follow_up_workflow(
        self,
        command_lower: str,
        snapshot: WorkbookSnapshot,
        selected_cell: str | None,
        selected_value: Any,
    ) -> ActionPlan | None:
        memory = snapshot.memory
        if not memory.last_action or not memory.recent_steps:
            return None

        follow_up_tokens = [
            "same",
            "do the same",
            "repeat",
            "apply it",
            "apply same",
            "also do",
            "same for",
            "as well",
        ]
        if not any(token in command_lower for token in follow_up_tokens):
            return None

        target_sheet = self._find_target_sheet(command_lower, snapshot)
        latest_step = memory.recent_steps[0]
        if "all sheets" in command_lower or "every sheet" in command_lower:
            sheets = [sheet.name for sheet in snapshot.sheets if sheet.name != latest_step.target_sheet]
            if not sheets:
                return None
            steps = [self._clone_step_for_sheet(latest_step, snapshot, sheet_name) for sheet_name in sheets]
            steps = [step for step in steps if step is not None]
            if not steps:
                return None
            return self._batch_from_steps(steps, f"Run {len(steps)} follow-up steps")

        if target_sheet == latest_step.target_sheet and target_sheet == snapshot.active_sheet:
            return None

        cloned = self._clone_step_for_sheet(latest_step, snapshot, target_sheet)
        if cloned is None:
            return None

        cloned.title = f"{latest_step.title} on {target_sheet}"
        cloned.explanation = f"I will repeat the last task on {target_sheet}."
        return self._step_to_plan(cloned)

    def _batch_from_steps(self, steps: list[TaskStep], title: str) -> ActionPlan:
        plans = [self._step_to_plan(step) for step in steps]
        risk_rank = {"low": 0, "medium": 1, "high": 2}
        combined_risk = max(plans, key=lambda item: risk_rank.get(item.risk_level, 0)).risk_level
        return ActionPlan(
            action="batch",
            target_sheet=plans[0].target_sheet,
            preview_title=title,
            explanation="I will repeat the last task across multiple sheets in order.",
            risk_level=combined_risk,
            requires_confirmation=any(step.requires_confirmation for step in plans),
            impacted_range=", ".join(step.impacted_range for step in plans if step.impacted_range) or None,
            parameters={"steps": [step.model_dump() for step in plans]},
            impact=ActionImpact(
                summary=f"{len(plans)} follow-up actions will run in sequence.",
                estimated_rows=sum(step.impact.estimated_rows for step in plans),
                estimated_cells=sum(step.impact.estimated_cells for step in plans),
                warnings=list(dict.fromkeys(warning for step in plans for warning in step.impact.warnings))[:5],
            ),
        )

    def _clone_step_for_sheet(
        self,
        step: TaskStep,
        snapshot: WorkbookSnapshot,
        target_sheet: str,
    ) -> TaskStep | None:
        sheet = self._sheet_by_name(snapshot, target_sheet)
        cloned = step.model_copy(deep=True)
        cloned.target_sheet = target_sheet

        if cloned.target_column:
            target_column = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, cloned.target_column)
            if target_column is None and cloned.target_column in sheet.get("headers", []):
                target_column = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, cloned.target_column)
            if target_column is None and cloned.target_column and len(cloned.target_column) == 1:
                target_column = cloned.target_column
            if target_column:
                cloned.target_column = cloned.target_column

        if cloned.target_cell and cloned.target_cell[0].isalpha():
            column_part, row_part = self._coordinate_parts(cloned.target_cell)
            if column_part and row_part is not None:
                cloned.target_cell = f"{column_part}{row_part}"

        return cloned

    def _step_to_plan(self, step: TaskStep) -> ActionPlan:
        return ActionPlan(
            action=step.action,
            target_sheet=step.target_sheet,
            target_cell=step.target_cell,
            target_column=step.target_column,
            formula=step.formula,
            preview_title=step.title,
            explanation=step.explanation,
            risk_level=step.risk_level,
            requires_confirmation=step.requires_confirmation,
            impacted_range=step.target_cell or step.target_column,
            parameters=step.parameters,
            impact=ActionImpact(summary=step.explanation, estimated_rows=0, estimated_cells=0),
        )

    @staticmethod
    def _split_multi_step_command(command_lower: str) -> list[str]:
        text = command_lower.replace("\n", " ")
        separators = [";"]
        for separator in separators:
            if separator in text:
                parts = [part.strip() for part in text.split(separator) if part.strip()]
                if len(parts) > 1:
                    return parts

        multi_step_markers = [" then ", " and then ", " after that ", " afterwards "]
        for marker in multi_step_markers:
            if marker in text:
                parts = [part.strip() for part in text.split(marker) if part.strip()]
                if len(parts) > 1:
                    return parts

        if " and " in text:
            parts = [part.strip() for part in text.split(" and ") if part.strip()]
            if len(parts) > 1:
                return parts

        return [command_lower.strip()]

    @staticmethod
    def _clean_sheet_name(raw_name: str) -> str:
        cleaned = re.sub(r"[\\/*?:\[\]]", "_", raw_name).strip()
        cleaned = re.sub(r"\s+", " ", cleaned)
        cleaned = cleaned.title()
        return cleaned[:31] or "Sheet"

    @staticmethod
    def _extract_rename_request(command_lower: str) -> str | None:
        if not any(token in command_lower for token in ["rename sheet", "rename tab", "sheet name", "tab name", "rename this sheet", "current sheet"]):
            return None
        match = re.search(r"(?:to|as|named|called)\s+([a-z0-9 _-]+)$", command_lower)
        if match:
            return match.group(1).strip()
        match = re.search(r"(?:to|as|named|called)\s+([a-z0-9 _-]+?)(?:\s+(?:and|then|,|$))", command_lower)
        if match:
            return match.group(1).strip()
        return None

    @staticmethod
    def _extract_row_anchor(command_lower: str) -> re.Match[str] | None:
        patterns = [
            r"(?:above|before|at)\s+(?:row\s+)?(\d+)",
            r"row\s+(\d+)",
        ]
        for pattern in patterns:
            match = re.search(pattern, command_lower)
            if match:
                return match
        return None

    @staticmethod
    def _extract_column_anchor(command_lower: str) -> re.Match[str] | None:
        patterns = [
            r"(?:before|after|at)\s+(?:column\s+)?([a-z]{1,3})",
            r"column\s+([a-z]{1,3})",
        ]
        for pattern in patterns:
            match = re.search(pattern, command_lower)
            if match:
                return match
        return None

    @staticmethod
    def _extract_row_span(command_lower: str) -> tuple[int, int] | None:
        patterns = [
            r"rows?\s+(\d+)\s*(?:to|-|through)\s*(\d+)",
            r"row\s+(\d+)\s*(?:to|-|through)\s*(\d+)",
            r"above\s+row\s+(\d+)",
            r"before\s+row\s+(\d+)",
            r"below\s+row\s+(\d+)",
            r"after\s+row\s+(\d+)",
        ]
        for pattern in patterns:
            match = re.search(pattern, command_lower)
            if match:
                if "above" in pattern or "before" in pattern:
                    row = int(match.group(1))
                    return row, row
                if "below" in pattern or "after" in pattern:
                    row = int(match.group(1)) + 1
                    return row, row
                return int(match.group(1)), int(match.group(2))
        return None

    @staticmethod
    def _extract_column_span(command_lower: str) -> tuple[str, str] | None:
        patterns = [
            r"columns?\s+([a-z]{1,3})\s*(?:to|-|through)\s*([a-z]{1,3})",
            r"column\s+([a-z]{1,3})\s*(?:to|-|through)\s*([a-z]{1,3})",
            r"from\s+([a-z]{1,3})\s+to\s+([a-z]{1,3})",
        ]
        for pattern in patterns:
            match = re.search(pattern, command_lower)
            if match:
                return match.group(1).upper(), match.group(2).upper()
        return None

    @staticmethod
    def _extract_cell_range(command_lower: str) -> str | None:
        match = re.search(r"\b([a-z]{1,3}\d+:[a-z]{1,3}\d+)\b", command_lower)
        if match:
            return match.group(1).upper()
        return None

    @staticmethod
    def _extract_cell_reference(command_lower: str) -> str | None:
        match = re.search(r"\b([a-z]{1,3}\d+)\b", command_lower)
        if match:
            return match.group(1).upper()
        return None

    @staticmethod
    def _cell_range_from_selected(selected_cell: str | None) -> str | None:
        if not selected_cell:
            return None
        cleaned = selected_cell.strip().upper()
        if ":" in cleaned:
            return cleaned
        return cleaned

    @staticmethod
    def _column_span_width(col_span: tuple[str, str] | None) -> int:
        if not col_span:
            return 1
        return abs(column_index_from_string(col_span[1]) - column_index_from_string(col_span[0])) + 1

    @staticmethod
    def _detect_number_format(command_lower: str) -> str | None:
        if any(token in command_lower for token in ["currency", "rupees", "rs", "$", "₹"]):
            return '"₹"#,##0.00'
        if any(token in command_lower for token in ["percent", "percentage"]):
            return "0.00%"
        if any(token in command_lower for token in ["date format", "as date", "date"]):
            return "DD-MMM-YYYY"
        if any(token in command_lower for token in ["number format", "as number", "numeric", "number"]):
            return "#,##0.00"
        return None

    def _unsupported_plan(self, target_sheet: str) -> ActionPlan:
        return ActionPlan(
            action="noop",
            target_sheet=target_sheet,
            preview_title="Command not supported yet",
            explanation=(
                "Is advanced build me formulas, sorting, duplicate cleanup, replace, filters, highlights, charts, "
                "column conversion, aur cross-sheet lookup supported hain. Is specific command ko abhi aur training chahiye."
            ),
            risk_level="low",
            requires_confirmation=False,
            parameters={},
            impact=ActionImpact(summary="No workbook changes will be made."),
        )

    def _build_difference_plan(
        self,
        command_lower: str,
        snapshot: WorkbookSnapshot,
        target_sheet: str,
        sheet: dict[str, Any],
        matched_headers: list[str],
    ) -> ActionPlan | None:
        header_row = workbook_service.detect_header_row(snapshot.session_id, target_sheet)
        data_start_row = header_row + 1
        target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)

        explicit_match = re.search(
            r"\b([a-z0-9 _&()-]+?)\s*(?:minus|subtract(?:ed)? by|less)\s*([a-z0-9 _&()-]+)\b",
            command_lower,
        )
        left_column = None
        right_column = None
        if explicit_match:
            left_token = explicit_match.group(1).strip()
            right_token = explicit_match.group(2).strip()
            left_column = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, left_token)
            right_column = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, right_token)

        if left_column is None or right_column is None:
            grand_total_candidates = workbook_service.find_columns_with_keywords(
                snapshot.session_id,
                target_sheet,
                ["grand total"],
            )
            if len(grand_total_candidates) >= 2:
                left_column = grand_total_candidates[0]
                right_column = grand_total_candidates[-1]

        if left_column is None or right_column is None:
            income_keywords = [
                "grand total",
                "income",
                "collection",
                "received",
                "revenue",
                "sales",
                "earning",
                "total income",
                "total collection",
            ]
            expense_keywords = [
                "grand total",
                "expense",
                "spend",
                "spent",
                "payment",
                "cost",
                "outflow",
                "debit",
                "total expense",
            ]
            income_candidates = workbook_service.find_columns_with_keywords(snapshot.session_id, target_sheet, income_keywords)
            expense_candidates = workbook_service.find_columns_with_keywords(snapshot.session_id, target_sheet, expense_keywords)
            if income_candidates:
                left_column = income_candidates[0]
            if expense_candidates:
                right_column = expense_candidates[-1]

        if (left_column is None or right_column is None) and len(sheet.get("numeric_headers") or []) >= 2:
            left_header = next((header for header in matched_headers if header in sheet.get("numeric_headers", [])), None)
            numeric_headers = list(sheet.get("numeric_headers") or [])
            left_header = left_header or numeric_headers[0]
            right_header = next((header for header in reversed(numeric_headers) if header != left_header), None)
            left_column = left_column or workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, left_header)
            right_column = right_column or workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, right_header)

        if not left_column or not right_column or left_column == right_column:
            return None

        return ActionPlan(
            action="fill_formula_down",
            target_sheet=target_sheet,
            target_column=target_column_letter,
            preview_title="Insert profit formula",
            explanation=(
                f"I will calculate profit on {target_sheet} using {left_column} minus {right_column} "
                f"and fill it down from row {data_start_row}."
            ),
            risk_level="low",
            requires_confirmation=False,
            impacted_range=f"{target_column_letter}{header_row}:{target_column_letter}{sheet['max_row']}",
            formula=f"={left_column}{{row}}-{right_column}{{row}}",
            parameters={
                "start_row": data_start_row,
                "end_row": sheet["max_row"],
                "header_row": header_row,
                "column_header": "Profit",
            },
            impact=ActionImpact(
                summary=f"A Profit column will be created from {left_column} minus {right_column}.",
                estimated_rows=max(0, sheet["max_row"] - data_start_row + 1),
                estimated_cells=max(0, sheet["max_row"] - data_start_row + 1),
            ),
        )

    def _build_prompt(
        self,
        command: str,
        snapshot: WorkbookSnapshot,
        selected_cell: str | None = None,
        selected_value: Any = None,
    ) -> str:
        selection_context = (
            "No current cell selection."
            if not selected_cell
            else f"Selected cell: {selected_cell}. Selected value: {json.dumps(selected_value, default=str)}."
        )
        return (
            "You are the ExcelMind Autonomous AI Agent. You don't just answer questions; you execute complex workflows.\n\n"
            "### AGENT CAPABILITIES:\n"
            "1. REASONING: Analyze headers and sample data to understand the semantic meaning of columns (e.g., 'Dues' means income).\n"
            "2. MULTI-STEP PLANNING: If a user gives a complex task, break it into a 'batch' of actions. For example, 'Clean and Chart' = [delete_duplicates, format_number, create_chart].\n"
            "3. AUTONOMY: Proactively suggest data improvements (anomalies, formatting) when asked to 'analyze' or 'fix everything'.\n"
            "4. MEMORY: Refer to previous actions in 'Conversation memory' to handle follow-up intent correctly.\n\n"
            "### CRITICAL INSTRUCTION ON INTENT:\n"
            "1. ACTION vs QUESTION:\n"
            "   - If the user asks a QUESTION (e.g., 'Analysis', 'Batao', 'Kitna hai', 'What is', 'Explain'): You MUST set action to 'analyze_workbook'. Do NOT try to modify the sheet. Instead, calculate the answer in your head and write it clearly in the 'explanation' field in Hinglish/English.\n"
            "   - If the user wants an ACTION (e.g., 'Apply', 'Fix', 'Insert'): Choose the appropriate action(s). Use 'batch' if >1 step is needed.\n\n"
            "2. PROFIT CALCULATION:\n"
            "   - If asked for profit, look for columns like 'Income', 'Sales', 'Revenue', 'Received' and subtract 'Expense', 'Cost', 'Paid', 'Dues'. If you find them, calculate the sum and report it.\n\n"
            "### RESPONSE SCHEMA (JSON):\n"
            "{\n"
            "  \"action\": \"string\",\n"
            "  \"target_sheet\": \"string\",\n"
            "  \"preview_title\": \"string (clear & concise)\",\n"
            "  \"explanation\": \"string (Agent's reasoning and verbal results)\",\n"
            "  \"risk_level\": \"low | medium | high\",\n"
            "  \"requires_confirmation\": boolean,\n"
            "  \"parameters\": { ... },\n"
            "  \"impact\": { \"summary\": \"string\", \"estimated_rows\": number, \"estimated_cells\": number, \"warnings\": [] }\n"
            "}\n\n"
            "### CONTEXT:\n"
            f"Selection: {selection_context}\n"
            f"Memory: {json.dumps(snapshot.memory.model_dump(), default=str)}\n"
            f"Workbook: {json.dumps(snapshot.model_dump(), default=str)}\n\n"
            f"USER COMMAND: {command}\n"
        )

    def _sheet_by_name(self, snapshot: WorkbookSnapshot, sheet_name: str) -> dict[str, Any]:
        context_map = {sheet.name: sheet.model_dump() for sheet in snapshot.context}
        summary_map = {sheet.name: sheet.model_dump() for sheet in snapshot.sheets}
        if sheet_name not in summary_map:
            raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found in snapshot.")
        return {**summary_map[sheet_name], **context_map.get(sheet_name, {})}

    def _find_target_sheet(self, command: str, snapshot: WorkbookSnapshot) -> str:
        normalized_command = self._normalize(command).replace(" ", "")
        for sheet in snapshot.sheets:
            normalized_sheet = self._normalize(sheet.name).replace(" ", "")
            if sheet.name.lower() in command or (normalized_sheet and normalized_sheet in normalized_command):
                return sheet.name
        return snapshot.active_sheet

    def _find_header_in_command(self, command: str, headers: list[str]) -> str | None:
        normalized_command = self._normalize(command)
        # Try exact matches first
        for header in headers:
            normalized_header = self._normalize(header)
            if normalized_header and normalized_header in normalized_command:
                return header
        
        # Try synonym matches if no direct match
        synonyms = {
            "income": ["collection", "received", "tuition", "fee", "revenue", "sales", "earning", "profit"],
            "expense": ["spent", "payment", "cost", "outflow", "debit", "salary", "spend"],
            "date": ["time", "month", "year", "day", "period"],
        }
        for category, tokens in synonyms.items():
            if category in normalized_command or any(t in normalized_command for t in tokens):
                for header in headers:
                    norm_h = self._normalize(header)
                    if any(t in norm_h for t in tokens) or norm_h == category:
                        return header
        return None

    def _selected_header_for_cell(self, sheet: dict[str, Any], selected_cell: str | None) -> str | None:
        if not selected_cell:
            return None
        column_letter, row_number = self._coordinate_parts(selected_cell)
        if not column_letter or row_number is None:
            return None

        headers = list(sheet.get("headers") or [])
        if not headers:
            return None

        column_index = self._column_index_from_letter(column_letter) - 1
        if column_index < 0 or column_index >= len(headers):
            return None
        return headers[column_index] or None

    def _selected_header_from_value(self, sheet: dict[str, Any], selected_value: Any) -> str | None:
        headers = list(sheet.get("headers") or [])
        if not headers:
            return None
        normalized_value = self._normalize(self._stringify(selected_value))
        if not normalized_value:
            return None
        for header in headers:
            if self._normalize(header) == normalized_value:
                return header
        return None

    @staticmethod
    def _extract_formula_text(value: Any) -> str | None:
        if isinstance(value, str):
            text = value.strip()
            return text if text.startswith("=") else None
        return None

    def _explain_formula(
        self,
        formula: str,
        sheet: dict[str, Any],
        selected_header: str | None,
        selected_cell: str | None,
    ) -> str:
        normalized = formula.strip()
        body = normalized[1:] if normalized.startswith("=") else normalized
        match = re.match(r"([A-Z][A-Z0-9\.]*)\((.*)\)$", body, re.IGNORECASE)
        location = selected_cell or f"{sheet['name']}!A1"
        header_part = f" for {selected_header}" if selected_header else ""
        if not match:
            return f"{location}{header_part} uses the formula {formula}. I can explain a clearer breakdown if you select a standard Excel function."

        function_name = match.group(1).upper()
        args = [part.strip() for part in self._split_formula_args(match.group(2))]
        if function_name == "SUM" and args:
            return f"{location}{header_part}: adds every numeric value in {args[0]}."
        if function_name == "AVERAGE" and args:
            return f"{location}{header_part}: calculates the average of values in {args[0]}."
        if function_name == "SUMIFS" and len(args) >= 3:
            criteria = ", ".join(f"{args[idx]}={args[idx + 1]}" for idx in range(1, len(args), 2) if idx + 1 < len(args))
            return f"{location}{header_part}: sums {args[0]} when {criteria}."
        if function_name == "COUNTIFS" and len(args) >= 2:
            criteria = ", ".join(f"{args[idx]}={args[idx + 1]}" for idx in range(0, len(args), 2) if idx + 1 < len(args))
            return f"{location}{header_part}: counts rows when {criteria}."
        if function_name == "AVERAGEIFS" and len(args) >= 3:
            criteria = ", ".join(f"{args[idx]}={args[idx + 1]}" for idx in range(1, len(args), 2) if idx + 1 < len(args))
            return f"{location}{header_part}: averages {args[0]} when {criteria}."
        if function_name == "COUNTIF" and args:
            return f"{location}{header_part}: counts cells in {args[0]} that match {args[1] if len(args) > 1 else 'the condition'}."
        if function_name == "SUMIF" and len(args) >= 3:
            return f"{location}{header_part}: sums {args[2]} where {args[0]} matches {args[1]}."
        if function_name == "AVERAGEIF" and len(args) >= 3:
            return f"{location}{header_part}: averages {args[2]} where {args[0]} matches {args[1]}."
        if function_name == "XLOOKUP" and len(args) >= 3:
            fallback = args[3] if len(args) > 3 else 'an optional fallback'
            return (
                f"{location}{header_part}: searches for {args[0]} in {args[1]}, "
                f"returns the matching value from {args[2]}, and uses {fallback} when not found."
            )
        if function_name == "INDEX" and len(args) >= 2:
            return f"{location}{header_part}: returns the value at the position described by {', '.join(args[:2])}."
        if function_name == "MATCH" and len(args) >= 2:
            return f"{location}{header_part}: returns the position of {args[0]} inside {args[1]}."
        if function_name == "IFERROR" and args:
            return f"{location}{header_part}: returns the first expression, and shows the fallback value when it errors."
        if function_name == "IFNA" and args:
            return f"{location}{header_part}: returns the first expression, and shows the fallback value only when #N/A appears."
        if function_name == "VLOOKUP" and len(args) >= 4:
            return (
                f"{location}{header_part}: looks for {args[0]} in the leftmost column of {args[1]}, "
                f"returns column {args[2]}, and uses {args[3]} for exact match behavior."
            )
        if function_name == "IF" and len(args) >= 3:
            return f"{location}{header_part}: checks {args[0]}; if true it returns {args[1]}, otherwise {args[2]}."
        if function_name == "LET" and len(args) >= 3:
            return f"{location}{header_part}: defines named values, then evaluates the final expression more efficiently."
        if function_name == "FILTER" and len(args) >= 2:
            return f"{location}{header_part}: returns only the rows from {args[0]} that satisfy {args[1]}."
        if function_name == "UNIQUE" and args:
            return f"{location}{header_part}: returns distinct values from {args[0]}."
        if function_name == "SORT" and args:
            return f"{location}{header_part}: sorts the array in {args[0]}."
        if function_name == "TEXTJOIN" and len(args) >= 3:
            return f"{location}{header_part}: joins text values using {args[0]} while optionally ignoring blanks."
        if function_name == "CONCAT" and args:
            return f"{location}{header_part}: joins the supplied text values into one string."
        if function_name == "SUMPRODUCT" and args:
            return f"{location}{header_part}: multiplies matching array values and sums the results."
        if function_name in {"LEFT", "RIGHT", "MID"} and args:
            return f"{location}{header_part}: extracts text using {function_name} with arguments {', '.join(args)}."
        if any(operator in body for operator in ["+", "-", "*", "/"]):
            return f"{location}{header_part}: performs a direct calculation using {body}."
        return f"{location}{header_part} uses {function_name} with arguments {', '.join(args)}."

    @staticmethod
    def _split_formula_args(argument_text: str) -> list[str]:
        args: list[str] = []
        current: list[str] = []
        depth = 0
        in_quotes = False
        for char in argument_text:
            if char == '"':
                in_quotes = not in_quotes
            if char == "," and depth == 0 and not in_quotes:
                args.append("".join(current).strip())
                current = []
                continue
            if char == "(" and not in_quotes:
                depth += 1
            elif char == ")" and not in_quotes and depth > 0:
                depth -= 1
            current.append(char)
        if current:
            args.append("".join(current).strip())
        return args

    def _fix_formula(
        self,
        formula: str,
        sheet: dict[str, Any],
        selected_header: str | None,
    ) -> str:
        corrected = formula.strip().replace("“", '"').replace("”", '"').replace("’", "'")
        if not corrected.startswith("="):
            corrected = f"={corrected}"
        corrected = corrected.replace(";", ",")
        corrected = re.sub(r"==+", "=", corrected)
        open_count = corrected.count("(")
        close_count = corrected.count(")")
        if close_count < open_count:
            corrected += ")" * (open_count - close_count)
        match = re.match(r"=([A-Z][A-Z0-9\.]*)\((.*)\)$", corrected, re.IGNORECASE)
        if match:
            function_name = match.group(1).upper()
            args = self._split_formula_args(match.group(2))
            if function_name == "VLOOKUP" and len(args) == 3:
                corrected = f"=VLOOKUP({args[0]},{args[1]},{args[2]},FALSE)"
            elif function_name == "XLOOKUP" and len(args) == 3:
                corrected = f'=XLOOKUP({args[0]},{args[1]},{args[2]},"")'
            elif function_name == "MATCH" and len(args) == 2:
                corrected = f"=MATCH({args[0]},{args[1]},0)"
            elif function_name == "IFERROR" and len(args) == 1:
                corrected = f'=IFERROR({args[0]},"")'
            elif function_name == "IFNA" and len(args) == 1:
                corrected = f'=IFNA({args[0]},"")'
            elif function_name in {"SUMIFS", "COUNTIFS", "AVERAGEIFS"} and len(args) >= 3 and len(args) % 2 == 0:
                # Keep modern aggregate formulas syntactically safe if the command omitted a final criterion pair.
                corrected = corrected[:-1] if corrected.endswith(")") else corrected
        if "??" in corrected:
            corrected = corrected.replace("??", "")
        return corrected

    def _generate_formula_plan(
        self,
        command_lower: str,
        snapshot: WorkbookSnapshot,
        target_sheet: str,
        sheet: dict[str, Any],
        selected_cell: str | None,
        selected_header: str | None,
        current_formula: str | None,
    ) -> ActionPlan | None:
        matched_headers = self._find_matching_headers(command_lower, sheet["headers"])
        matched_header = self._find_header_in_command(command_lower, sheet["headers"]) or selected_header
        numeric_header = self._pick_numeric_header(sheet, matched_header)
        text_header = self._pick_text_header(sheet, matched_header)

        if any(token in command_lower for token in ["profit", "net profit", "difference", "minus", "subtract", "revenue minus", "income minus", "collection minus"]):
            difference_plan = self._build_difference_plan(command_lower, snapshot, target_sheet, sheet, matched_headers)
            if difference_plan is not None:
                return difference_plan

        if any(token in command_lower for token in ["sum", "total", "average", "avg", "mean"]) and numeric_header is not None:
            multi_criteria_plan = self._build_multi_criteria_formula_plan(
                command_lower,
                snapshot,
                target_sheet,
                sheet,
                matched_headers,
                numeric_header,
            )
            if multi_criteria_plan is not None:
                return multi_criteria_plan

        if "count" in command_lower:
            count_plan = self._build_multi_criteria_formula_plan(
                command_lower,
                snapshot,
                target_sheet,
                sheet,
                matched_headers,
                numeric_header,
            )
            if count_plan is not None:
                return count_plan

        if any(token in command_lower for token in ["sum", "total"]) and numeric_header is not None:
            column_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, numeric_header)
            target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
            target_cell = workbook_service.find_first_empty_cell(snapshot.session_id, target_sheet, target_column_letter)
            return ActionPlan(
                action="generate_formula",
                target_sheet=target_sheet,
                target_cell=target_cell,
                target_column=numeric_header,
                formula=f"=SUM({column_letter}2:{column_letter}{sheet['max_row']})",
                preview_title="Generate total formula",
                explanation=f"I will generate a SUM formula for {numeric_header} in {target_cell}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=target_cell,
                parameters={},
                impact=ActionImpact(summary="One formula cell will be inserted.", estimated_rows=1, estimated_cells=1),
            )

        if "average" in command_lower or "avg" in command_lower or "mean" in command_lower:
            if numeric_header is None:
                return None
            column_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, numeric_header)
            target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
            target_cell = workbook_service.find_first_empty_cell(snapshot.session_id, target_sheet, target_column_letter)
            return ActionPlan(
                action="generate_formula",
                target_sheet=target_sheet,
                target_cell=target_cell,
                target_column=numeric_header,
                formula=f"=AVERAGE({column_letter}2:{column_letter}{sheet['max_row']})",
                preview_title="Generate average formula",
                explanation=f"I will generate an AVERAGE formula for {numeric_header} in {target_cell}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=target_cell,
                parameters={},
                impact=ActionImpact(summary="One formula cell will be inserted.", estimated_rows=1, estimated_cells=1),
            )

        lookup_plan = self._build_lookup_formula_plan(
            command_lower,
            snapshot,
            target_sheet,
            sheet,
            selected_cell,
            selected_header,
            matched_header,
            text_header,
        )
        if lookup_plan is not None:
            return lookup_plan

        if "vlookup" in command_lower and len(snapshot.sheets) > 1:
            source_sheet = next(sheet_item for sheet_item in snapshot.sheets if sheet_item.name != target_sheet)
            target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
            target_cell = workbook_service.find_first_empty_cell(snapshot.session_id, target_sheet, target_column_letter)
            return ActionPlan(
                action="generate_formula",
                target_sheet=target_sheet,
                target_cell=target_cell,
                target_column=matched_header or text_header or sheet["headers"][0],
                formula=f'=IFERROR(VLOOKUP(A{{row}},\'{source_sheet.name}\'!A:B,2,FALSE),"")',
                preview_title="Generate VLOOKUP formula",
                explanation=f"I will generate a VLOOKUP formula on {target_sheet} using {source_sheet.name} as the source table.",
                risk_level="medium",
                requires_confirmation=True,
                impacted_range=target_cell,
                parameters={"target_column_letter": target_column_letter, "start_row": 2, "end_row": max(2, sheet["max_row"])},
                impact=ActionImpact(
                    summary=f"A lookup formula will be filled across {max(0, sheet['max_row'] - 1)} rows.",
                    estimated_rows=max(0, sheet["max_row"] - 1),
                    estimated_cells=max(0, sheet["max_row"] - 1),
                    warnings=["Cross-sheet formulas should be reviewed before applying on production files."],
                ),
            )

        if any(token in command_lower for token in ["unique", "distinct", "dedupe summary", "distinct list"]) and text_header is not None:
            source_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, text_header)
            if source_letter is not None:
                target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
                target_cell = workbook_service.find_first_empty_cell(snapshot.session_id, target_sheet, target_column_letter)
                return ActionPlan(
                    action="generate_formula",
                    target_sheet=target_sheet,
                    target_cell=target_cell,
                    target_column=text_header,
                    formula=f"=UNIQUE({source_letter}2:{source_letter}{sheet['max_row']})",
                    preview_title="Generate UNIQUE formula",
                    explanation=f"I will generate a UNIQUE formula for {text_header} in {target_cell}.",
                    risk_level="low",
                    requires_confirmation=False,
                    impacted_range=target_cell,
                    parameters={},
                    impact=ActionImpact(summary="A distinct-value formula will be inserted.", estimated_rows=1, estimated_cells=1),
                )

        if any(token in command_lower for token in ["sort", "sorted list"]) and numeric_header is not None:
            source_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, numeric_header)
            if source_letter is not None:
                target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
                target_cell = workbook_service.find_first_empty_cell(snapshot.session_id, target_sheet, target_column_letter)
                descending = "desc" in command_lower or "descending" in command_lower or "reverse" in command_lower
                sort_formula = f"=SORT({source_letter}2:{source_letter}{sheet['max_row']},{1 if descending else 1},{str(descending).upper()})"
                return ActionPlan(
                    action="generate_formula",
                    target_sheet=target_sheet,
                    target_cell=target_cell,
                    target_column=numeric_header,
                    formula=sort_formula,
                    preview_title="Generate SORT formula",
                    explanation=f"I will generate a SORT formula for {numeric_header} in {target_cell}.",
                    risk_level="low",
                    requires_confirmation=False,
                    impacted_range=target_cell,
                    parameters={},
                    impact=ActionImpact(summary="A dynamic sorting formula will be inserted.", estimated_rows=1, estimated_cells=1),
                )

        if "where" in command_lower and numeric_header is not None:
            operator, criterion = self._extract_filter_criterion(command_lower)
            value_header = next((header for header in matched_headers if header in sheet.get("numeric_headers", [])), numeric_header)
            condition_header = self._pick_condition_header(sheet, matched_headers, value_header, operator)
            condition_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, condition_header)
            value_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, value_header) if value_header else None
            if condition_letter is None:
                return None
            target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
            target_cell = workbook_service.find_first_empty_cell(snapshot.session_id, target_sheet, target_column_letter)
            criteria = self._excel_criteria(operator, criterion).replace('"', '""')
            condition_range = f"{condition_letter}2:{condition_letter}{sheet['max_row']}"
            formula = (
                f'=COUNTIF({condition_range},"{criteria}")'
                if value_letter is None
                else f'=SUMIF({condition_range},"{criteria}",{value_letter}2:{value_letter}{sheet["max_row"]})'
            )
            return ActionPlan(
                action="generate_formula",
                target_sheet=target_sheet,
                target_cell=target_cell,
                target_column=value_header or condition_header,
                formula=formula,
                preview_title="Generate conditional formula",
                explanation=f"I will generate a conditional formula for {target_sheet} using {condition_header} {operator} {criterion}.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=target_cell,
                parameters={},
                impact=ActionImpact(summary="One formula cell will be inserted.", estimated_rows=1, estimated_cells=1),
            )

        if current_formula and selected_cell:
            return ActionPlan(
                action="generate_formula",
                target_sheet=target_sheet,
                target_cell=selected_cell,
                target_column=selected_header,
                formula=current_formula,
                preview_title="Reuse formula",
                explanation="I can reuse the selected formula, but I need a more specific request to generate a new one.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range=selected_cell,
                parameters={},
                impact=ActionImpact(summary="No workbook changes will be made."),
            )

        return None

    def _build_multi_criteria_formula_plan(
        self,
        command_lower: str,
        snapshot: WorkbookSnapshot,
        target_sheet: str,
        sheet: dict[str, Any],
        matched_headers: list[str],
        numeric_header: str | None,
    ) -> ActionPlan | None:
        if " where " not in command_lower and " with " not in command_lower and " when " not in command_lower:
            return None

        clauses = self._parse_condition_clauses(command_lower, sheet.get("headers") or [], None)
        if not clauses:
            return None

        is_count = "count" in command_lower and not any(token in command_lower for token in ["sum", "average", "avg", "mean"])
        header_pool = list(sheet.get("headers") or [])
        if is_count:
            value_header = next(
                (header for header in matched_headers if header not in {clause_header for clause_header, _, _ in clauses}),
                header_pool[0] if header_pool else clauses[0][0],
            )
            formula_name = "COUNTIFS" if len(clauses) > 1 else "COUNTIF"
        else:
            value_header = next((header for header in matched_headers if header in sheet.get("numeric_headers", [])), numeric_header)
            if value_header is None:
                return None
            formula_name = "AVERAGEIFS" if "average" in command_lower or "avg" in command_lower or "mean" in command_lower else "SUMIFS"
            if len(clauses) == 1:
                formula_name = "AVERAGEIF" if formula_name == "AVERAGEIFS" else "SUMIF"

        target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
        target_cell = workbook_service.find_first_empty_cell(snapshot.session_id, target_sheet, target_column_letter)
        data_end_row = max(2, sheet["max_row"])

        value_range = None
        if not is_count:
            value_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, value_header)
            if value_letter is None:
                return None
            value_range = f"{value_letter}2:{value_letter}{data_end_row}"

        formula_parts: list[str] = []
        if formula_name in {"SUMIFS", "AVERAGEIFS"}:
            formula_parts.append(value_range)
        elif formula_name in {"SUMIF", "AVERAGEIF"}:
            formula_parts.append(value_range)

        for header, operator, criterion in clauses:
            header_letter = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, header)
            if header_letter is None:
                continue
            condition_range = f"{header_letter}2:{header_letter}{data_end_row}"
            criteria = self._excel_criteria(operator, criterion).replace('"', '""')
            formula_parts.append(condition_range)
            formula_parts.append(f'"{criteria}"')

        if formula_name == "COUNTIF" and len(formula_parts) < 2:
            return None
        if formula_name in {"SUMIF", "AVERAGEIF"} and len(formula_parts) < 3:
            return None
        if formula_name in {"COUNTIFS", "SUMIFS", "AVERAGEIFS"} and len(formula_parts) < 3:
            return None

        if formula_name == "SUMIF":
            formula = f"=SUMIF({formula_parts[1]},{formula_parts[2]},{value_range})"
        elif formula_name == "AVERAGEIF":
            formula = f"=AVERAGEIF({formula_parts[1]},{formula_parts[2]},{value_range})"
        elif formula_name == "COUNTIF":
            formula = f"=COUNTIF({formula_parts[1]},{formula_parts[2]})"
        elif formula_name == "COUNTIFS":
            formula = f"=COUNTIFS({', '.join(formula_parts)})"
        else:
            formula = f"={formula_name}({', '.join(formula_parts)})"

        readable_conditions = ", ".join(f"{header} {operator} {criterion}" for header, operator, criterion in clauses)
        return ActionPlan(
            action="generate_formula",
            target_sheet=target_sheet,
            target_cell=target_cell,
            target_column=value_header,
            formula=formula,
            preview_title=f"Generate {formula_name}",
            explanation=f"I will generate {formula_name} for {value_header} using {readable_conditions}.",
            risk_level="low" if len(clauses) == 1 else "medium",
            requires_confirmation=len(clauses) > 1,
            impacted_range=target_cell,
            parameters={},
            impact=ActionImpact(
                summary="A formula cell will be inserted.",
                estimated_rows=1,
                estimated_cells=1,
                warnings=["Multiple conditions were inferred from the command."] if len(clauses) > 1 else [],
            ),
        )

    def _build_lookup_formula_plan(
        self,
        command_lower: str,
        snapshot: WorkbookSnapshot,
        target_sheet: str,
        sheet: dict[str, Any],
        selected_cell: str | None,
        selected_header: str | None,
        matched_header: str | None,
        text_header: str | None,
    ) -> ActionPlan | None:
        if not any(token in command_lower for token in ["lookup", "match", "search", "find"]):
            return None
        if len(snapshot.sheets) <= 1:
            return None

        source_sheet = next(sheet_item for sheet_item in snapshot.sheets if sheet_item.name != target_sheet)
        source_headers = list(self._sheet_by_name(snapshot, source_sheet.name).get("headers") or [])
        if not source_headers:
            return None

        lookup_header = matched_header or selected_header or text_header or source_headers[0]
        source_lookup_header = self._find_header_in_command(command_lower, source_headers) or source_headers[0]
        return_header = None
        for header in source_headers:
            if header != source_lookup_header:
                return_header = header
                break
        if return_header is None:
            return None

        lookup_column = workbook_service.header_to_column_letter(snapshot.session_id, target_sheet, lookup_header)
        source_lookup_letter = workbook_service.header_to_column_letter(snapshot.session_id, source_sheet.name, source_lookup_header)
        source_return_letter = workbook_service.header_to_column_letter(snapshot.session_id, source_sheet.name, return_header)
        if lookup_column is None or source_lookup_letter is None or source_return_letter is None:
            return None

        target_column_letter = workbook_service.next_available_column_letter(snapshot.session_id, target_sheet)
        target_cell = workbook_service.find_first_empty_cell(snapshot.session_id, target_sheet, target_column_letter)
        use_xlookup = "xlookup" in command_lower or "modern lookup" in command_lower
        use_index_match = "index match" in command_lower or "index/match" in command_lower
        current_reference = f"{lookup_column}2"
        source_range_start = 2
        source_range_end = max(2, self._sheet_by_name(snapshot, source_sheet.name)["max_row"])
        source_lookup_range = f"'{source_sheet.name}'!{source_lookup_letter}{source_range_start}:{source_lookup_letter}{source_range_end}"
        source_return_range = f"'{source_sheet.name}'!{source_return_letter}{source_range_start}:{source_return_letter}{source_range_end}"

        if use_xlookup:
            formula = f'=XLOOKUP({current_reference},{source_lookup_range},{source_return_range},"")'
            preview_title = "Generate XLOOKUP formula"
            explanation = f"I will generate an XLOOKUP formula on {target_sheet} using {source_sheet.name}."
        elif use_index_match:
            formula = (
                f'=IFERROR(INDEX({source_return_range},MATCH({current_reference},{source_lookup_range},0)),"")'
            )
            preview_title = "Generate INDEX/MATCH formula"
            explanation = f"I will generate an INDEX/MATCH formula on {target_sheet} using {source_sheet.name}."
        else:
            formula = f'=IFERROR(VLOOKUP({current_reference},\'{source_sheet.name}\'!{source_lookup_letter}:{source_return_letter},2,FALSE),"")'
            preview_title = "Generate lookup formula"
            explanation = f"I will generate a lookup formula on {target_sheet} using {source_sheet.name}."

        return ActionPlan(
            action="generate_formula",
            target_sheet=target_sheet,
            target_cell=target_cell,
            target_column=lookup_header,
            formula=formula,
            preview_title=preview_title,
            explanation=explanation,
            risk_level="medium",
            requires_confirmation=True,
            impacted_range=target_cell,
            parameters={},
            impact=ActionImpact(
                summary="A cross-sheet lookup formula will be inserted.",
                estimated_rows=1,
                estimated_cells=1,
                warnings=["Cross-sheet formulas should be reviewed before applying on production files."],
            ),
        )

    def _parse_condition_clauses(
        self,
        command_lower: str,
        headers: list[str],
        value_header: str | None,
    ) -> list[tuple[str, str, Any]]:
        tail = command_lower
        for marker in [" where ", " with ", " when "]:
            if marker in tail:
                tail = tail.split(marker, 1)[1]
                break

        clauses_raw = re.split(r"\s+(?:and|aur|or)\s+", tail)
        clauses: list[tuple[str, str, Any]] = []
        for clause in clauses_raw:
            normalized_clause = self._normalize(clause)
            header = next((item for item in headers if self._normalize(item) in normalized_clause), None)
            if header is None or header == value_header:
                continue
            operator, criterion = self._extract_clause_details(clause, header)
            clauses.append((header, operator, criterion))
        return clauses

    def _extract_clause_details(self, clause: str, header: str) -> tuple[str, Any]:
        clause_lower = clause.lower().strip()
        header_norm = self._normalize(header)
        remainder = re.sub(rf"\b{re.escape(header_norm)}\b", "", self._normalize(clause_lower))
        if "contains" in clause_lower or "includes" in clause_lower:
            raw = self._extract_text_after_marker(clause_lower, ["contains", "includes", "has"])
            if raw:
                return "contains", raw
        numeric_match = re.search(r"(>=|<=|>|<|=)\s*(-?\d+(?:\.\d+)?)", clause_lower)
        if numeric_match:
            return self._map_operator(numeric_match.group(1)), float(numeric_match.group(2))
        greater_match = re.search(r"(above|greater than|more than)\s*(-?\d+(?:\.\d+)?)", clause_lower)
        if greater_match:
            return "greater_than", float(greater_match.group(2))
        less_match = re.search(r"(below|less than|under)\s*(-?\d+(?:\.\d+)?)", clause_lower)
        if less_match:
            return "less_than", float(less_match.group(2))
        if "today" in clause_lower or "aaj" in clause_lower:
            from datetime import date
            return "equals", date.today().strftime("%Y-%m-%d")
        if "yesterday" in clause_lower or "kal" in clause_lower:
            from datetime import date, timedelta
            return "equals", (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")
        text_match = re.search(r"(?:is|equals|=)\s+([a-z0-9 _-]+)$", clause_lower)
        if text_match:
            return "equals", self._clean_criterion_text(text_match.group(1))
        cleaned = self._clean_criterion_text(clause_lower.replace(header.lower(), "").strip())
        cleaned = cleaned or self._clean_criterion_text(remainder)
        return "equals", cleaned

    @staticmethod
    def _extract_text_after_marker(text: str, markers: list[str]) -> str | None:
        for marker in markers:
            pattern = rf"{re.escape(marker)}\s+([a-z0-9 _-]+)$"
            match = re.search(pattern, text)
            if match:
                return match.group(1).strip()
        return None

    def _find_matching_headers(self, command: str, headers: list[str]) -> list[str]:
        normalized_command = self._normalize(command)
        matches: list[str] = []
        for header in headers:
            normalized_header = self._normalize(header)
            if normalized_header and normalized_header in normalized_command:
                matches.append(header)
        return matches

    def _pick_numeric_header(self, sheet: dict[str, Any], preferred: str | None) -> str | None:
        numeric_headers = list(sheet.get("numeric_headers") or [])
        if preferred in numeric_headers:
            return preferred
        
        # Filter out obvious ID/Index/Serial columns if others exist
        index_keywords = ["id", "serial", "s.no", "no.", "index", "row", "sl no", "a"]
        better_numeric = [
            h for h in numeric_headers 
            if not any(k == h.lower() or k == self._normalize(h) for k in index_keywords)
        ]
        
        if better_numeric:
            return better_numeric[0]
        return numeric_headers[0] if numeric_headers else None

    def _pick_text_header(self, sheet: dict[str, Any], preferred: str | None) -> str | None:
        text_headers = list(sheet.get("text_headers") or [])
        if preferred in text_headers:
            return preferred
        return text_headers[0] if text_headers else (sheet["headers"][0] if sheet.get("headers") else None)

    def _pick_condition_header(
        self,
        sheet: dict[str, Any],
        matched_headers: list[str],
        value_header: str | None,
        operator: str,
    ) -> str | None:
        preferred_pool = (
            list(sheet.get("text_headers") or [])
            if operator in {"equals", "contains"}
            else list(sheet.get("numeric_headers") or [])
        )
        for header in matched_headers:
            if header != value_header:
                return header
        for header in preferred_pool:
            if header != value_header:
                return header
        for header in sheet.get("headers") or []:
            if header != value_header:
                return header
        return value_header

    def _extract_filter_criterion(self, command: str) -> tuple[str, Any]:
        command_lower = command.lower()
        
        # Special handling for relative dates
        if "today" in command_lower or "aaj" in command_lower:
            from datetime import date
            return "equals", date.today().strftime("%Y-%m-%d")
        if "yesterday" in command_lower or "kal" in command_lower:
            from datetime import date, timedelta
            return "equals", (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")

        contains_match = re.search(r"(?:contains|includes)\s+([a-z0-9 _-]+)$", command)
        if contains_match:
            return "contains", self._clean_criterion_text(contains_match.group(1))
        numeric_match = re.search(r"(>=|<=|>|<|=)\s*(-?\d+(?:\.\d+)?)", command)
        if numeric_match:
            return self._map_operator(numeric_match.group(1)), float(numeric_match.group(2))
        greater_match = re.search(r"(above|greater than)\s*(-?\d+(?:\.\d+)?)", command)
        if greater_match:
            return "greater_than", float(greater_match.group(2))
        less_match = re.search(r"(below|less than)\s*(-?\d+(?:\.\d+)?)", command)
        if less_match:
            return "less_than", float(less_match.group(2))
        text_match = re.search(r"(?:is|equals|=)\s+([a-z0-9 _-]+)$", command)
        if text_match:
            return "equals", self._clean_criterion_text(text_match.group(1))
        return "greater_than", 0

    @staticmethod
    def _extract_target_type(command: str) -> str | None:
        if "date" in command:
            return "date"
        if any(token in command for token in ["number", "numeric"]):
            return "number"
        if "text" in command:
            return "text"
        return None

    @staticmethod
    def _excel_criteria(operator: str, criterion: Any) -> str:
        mapping = {
            "greater_than": ">",
            "greater_or_equal": ">=",
            "less_than": "<",
            "less_or_equal": "<=",
            "equals": "",
            "contains": "*",
        }
        normalized_operator = operator.lower()
        if normalized_operator == "contains":
            return f"*{criterion}*"
        prefix = mapping.get(normalized_operator, "")
        return f"{prefix}{criterion}"

    @staticmethod
    def _clean_criterion_text(text: str) -> str:
        cleaned = text.strip()
        cleaned = re.sub(r"\b(rows?|records?|cells?|values?)\b.*$", "", cleaned).strip()
        cleaned = re.sub(
            r"\b(filter|show|only|remain|nikalo|karo|banao|apply|insert|formula|highlight|convert|me|column|ka|ki)\b.*$",
            "",
            cleaned,
        ).strip()
        return cleaned

    @staticmethod
    def _map_operator(symbol: str) -> str:
        return {
            ">": "greater_than",
            ">=": "greater_or_equal",
            "<": "less_than",
            "<=": "less_or_equal",
            "=": "equals",
        }.get(symbol, "equals")

    @staticmethod
    def _col_letter(index: int) -> str:
        result = ""
        current = index
        while current:
            current, remainder = divmod(current - 1, 26)
            result = chr(65 + remainder) + result
        return result

    @staticmethod
    def _column_index_from_letter(letter: str) -> int:
        current = 0
        for char in letter.upper():
            current = current * 26 + (ord(char) - 64)
        return current

    @staticmethod
    def _normalize(text: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", text.lower())


ai_service = AIService()
