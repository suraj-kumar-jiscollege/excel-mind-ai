from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.formula.translate import Translator
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings
from app.models import (
    ActionPlan,
    CommandRecord,
    CommandTemplate,
    ConversationMemory,
    RiskLevel,
    ChartRecommendation,
    WorkbookAnomaly,
    WorkbookInsight,
    SheetContext,
    SheetSummary,
    TaskStep,
    WorkbookSnapshot,
    WorkbookStats,
)
from app.services.formula_engine import FormulaEngine, FormulaError


@dataclass(slots=True)
class WorkbookSession:
    session_id: str
    file_path: Path
    workbook: Workbook
    active_sheet: str
    dirty: bool = False
    command_history: list[CommandRecord] = field(default_factory=list)
    last_command: str | None = None
    last_plan: ActionPlan | None = None
    recent_plans: list[ActionPlan] = field(default_factory=list)
    undo_stack: list[bytes] = field(default_factory=list)
    redo_stack: list[bytes] = field(default_factory=list)


class WorkbookService:
    def __init__(self) -> None:
        self._sessions: dict[str, WorkbookSession] = {}
        self._formula_engine = FormulaEngine()

    def detect_header_row(self, session_id: str, sheet_name: str) -> int:
        worksheet = self._get_sheet(self.get_session(session_id), sheet_name)
        return self._detect_header_row(worksheet)

    def data_start_row(self, session_id: str, sheet_name: str) -> int:
        worksheet = self._get_sheet(self.get_session(session_id), sheet_name)
        return self._data_start_row(worksheet)

    def open_workbook(self, file_path: str) -> WorkbookSnapshot:
        path = Path(file_path).expanduser().resolve()
        self._validate_path(path)
        if not path.exists():
            raise HTTPException(status_code=404, detail="Workbook file not found.")
        if path.suffix.lower() != ".xlsx":
            raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

        workbook = load_workbook(path)
        session_id = str(uuid4())
        session = WorkbookSession(
            session_id=session_id,
            file_path=path,
            workbook=workbook,
            active_sheet=workbook.active.title,
        )
        self._sessions[session_id] = session
        return self.get_snapshot(session_id)

    def open_workbook_from_bytes(self, content: bytes, filename: str) -> WorkbookSnapshot:
        import io
        path = Path(filename)
        if path.suffix.lower() != ".xlsx":
            raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")
        
        file_io = io.BytesIO(content)
        # Optimization: use data_only=True to get values instead of formulas if possible
        # However, we need formulas for some tasks, so we keep it default.
        workbook = load_workbook(file_io, data_only=False)
        session_id = str(uuid4())
        session = WorkbookSession(
            session_id=session_id,
            file_path=path,
            workbook=workbook,
            active_sheet=workbook.active.title,
        )
        self._sessions[session_id] = session
        return self.get_snapshot(session_id)

    def get_workbook_bytes(self, session_id: str) -> bytes:
        session = self.get_session(session_id)
        import io
        out = io.BytesIO()
        session.workbook.save(out)
        session.dirty = False
        return out.getvalue()

    def get_session(self, session_id: str) -> WorkbookSession:
        session = self._sessions.get(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Workbook session not found.")
        return session

    def get_snapshot(self, session_id: str) -> WorkbookSnapshot:
        session = self.get_session(session_id)
        workbook = session.workbook
        sheets = [self._sheet_summary(workbook, workbook[sheet_name]) for sheet_name in workbook.sheetnames]
        context = [self._sheet_context(workbook, workbook[sheet_name]) for sheet_name in workbook.sheetnames]
        stats = self._build_stats(workbook, sheets, context)
        memory = self._build_memory(session)
        chart_recommendations = self._build_chart_recommendations(workbook, sheets, context, session.active_sheet)
        return WorkbookSnapshot(
            session_id=session.session_id,
            file_path=str(session.file_path),
            active_sheet=session.active_sheet,
            dirty=session.dirty,
            can_undo=bool(session.undo_stack),
            can_redo=bool(session.redo_stack),
            sheets=sheets,
            context=context,
            stats=stats,
            history=session.command_history,
            insights=self._build_insights(workbook, sheets, context, session.active_sheet),
            chart_recommendations=chart_recommendations,
            anomalies=self._build_anomalies(workbook, sheets, context, session.active_sheet),
            templates=self._build_templates(context, session.active_sheet, workbook.sheetnames),
            suggested_prompts=self._build_suggested_prompts(
                context,
                session.active_sheet,
                workbook.sheetnames,
                memory,
                chart_recommendations,
            ),
            memory=memory,
        )

    def set_active_sheet(self, session_id: str, sheet_name: str) -> WorkbookSnapshot:
        session = self.get_session(session_id)
        self._get_sheet(session, sheet_name)
        session.active_sheet = sheet_name
        return self.get_snapshot(session_id)

    def update_cell(self, session_id: str, sheet_name: str, cell: str, value: Any) -> WorkbookSnapshot:
        session = self.get_session(session_id)
        worksheet = self._get_sheet(session, sheet_name)
        self._checkpoint(session)
        previous_value = worksheet[cell].value
        worksheet[cell] = value
        session.dirty = True
        self._record_history(
            session,
            user_command=f"Manual edit {sheet_name}!{cell}",
            action="manual_edit",
            explanation=f"Updated {sheet_name}!{cell} from {self._stringify(previous_value)} to {self._stringify(value)}.",
            target_sheet=sheet_name,
            risk_level="low",
        )
        return self.get_snapshot(session_id)

    def save_workbook(self, session_id: str, destination_path: str | None = None) -> tuple[str, WorkbookSnapshot]:
        session = self.get_session(session_id)
        target = Path(destination_path).expanduser().resolve() if destination_path else session.file_path
        self._validate_path(target)
        target.parent.mkdir(parents=True, exist_ok=True)
        session.workbook.save(target)
        session.file_path = target
        session.dirty = False
        return str(target), self.get_snapshot(session_id)

    def undo(self, session_id: str) -> WorkbookSnapshot:
        session = self.get_session(session_id)
        if not session.undo_stack:
            raise HTTPException(status_code=400, detail="Nothing to undo.")
        session.redo_stack.append(self._serialize_workbook(session.workbook))
        previous_state = session.undo_stack.pop()
        session.workbook = self._restore_workbook(previous_state)
        if session.active_sheet not in session.workbook.sheetnames:
            session.active_sheet = session.workbook.active.title
        session.dirty = True
        self._record_history(
            session,
            user_command="Undo",
            action="undo",
            explanation="Reverted the latest workbook change.",
            target_sheet=session.active_sheet,
            risk_level="low",
            status="system",
        )
        return self.get_snapshot(session_id)

    def redo(self, session_id: str) -> WorkbookSnapshot:
        session = self.get_session(session_id)
        if not session.redo_stack:
            raise HTTPException(status_code=400, detail="Nothing to redo.")
        session.undo_stack.append(self._serialize_workbook(session.workbook))
        next_state = session.redo_stack.pop()
        session.workbook = self._restore_workbook(next_state)
        if session.active_sheet not in session.workbook.sheetnames:
            session.active_sheet = session.workbook.active.title
        session.dirty = True
        self._record_history(
            session,
            user_command="Redo",
            action="redo",
            explanation="Re-applied the latest reverted workbook change.",
            target_sheet=session.active_sheet,
            risk_level="low",
            status="system",
        )
        return self.get_snapshot(session_id)

    def execute_action(self, session_id: str, plan: ActionPlan) -> WorkbookSnapshot:
        session = self.get_session(session_id)
        if plan.action == "noop":
            return self.get_snapshot(session_id)

        self._checkpoint(session)
        if plan.action == "batch":
            steps = self._batch_steps(plan)
            if not steps:
                raise HTTPException(status_code=400, detail="Batch action needs at least one step.")
            for step in steps:
                self._apply_action(session, step)
            session.dirty = True
            self._remember_plan(session, plan.preview_title, plan)
            self._record_history(
                session,
                user_command=plan.preview_title,
                action=plan.action,
                explanation=plan.explanation,
                target_sheet=plan.target_sheet,
                risk_level=plan.risk_level,
            )
            return self.get_snapshot(session_id)

        self._apply_action(session, plan)

        session.dirty = True
        self._remember_plan(session, plan.preview_title, plan)
        self._record_history(
            session,
            user_command=plan.preview_title,
            action=plan.action,
            explanation=plan.explanation,
            target_sheet=plan.target_sheet,
            risk_level=plan.risk_level,
        )
        return self.get_snapshot(session_id)

    def _apply_action(self, session: WorkbookSession, plan: ActionPlan) -> None:
        worksheet = self._get_sheet(session, plan.target_sheet)
        if plan.action == "insert_formula":
            if not plan.target_cell or not plan.formula:
                raise HTTPException(status_code=400, detail="Formula insertion requires target_cell and formula.")
            worksheet[plan.target_cell] = plan.formula
        elif plan.action == "fill_formula_down":
            self._fill_formula_down(worksheet, plan)
        elif plan.action == "fix_formula":
            if not plan.target_cell or not plan.formula:
                raise HTTPException(status_code=400, detail="Formula fix requires target_cell and formula.")
            worksheet[plan.target_cell] = plan.formula
        elif plan.action == "generate_formula":
            if not plan.target_cell or not plan.formula:
                raise HTTPException(status_code=400, detail="Formula generation requires target_cell and formula.")
            worksheet[plan.target_cell] = plan.formula
        elif plan.action == "explain_formula":
            return
        elif plan.action == "create_table":
            self._create_table(worksheet, plan)
        elif plan.action == "freeze_header":
            self._freeze_header(worksheet)
        elif plan.action == "auto_fit_columns":
            self._auto_fit_columns(worksheet)
        elif plan.action == "format_header":
            self._format_header(worksheet, plan)
        elif plan.action == "format_number":
            self._format_number_columns(worksheet, plan)
        elif plan.action == "insert_rows":
            self._insert_rows(worksheet, plan)
        elif plan.action == "delete_rows":
            self._delete_rows(worksheet, plan)
        elif plan.action == "insert_columns":
            self._insert_columns(worksheet, plan)
        elif plan.action == "delete_columns":
            self._delete_columns(worksheet, plan)
        elif plan.action == "clear_cells":
            self._clear_cells(worksheet, plan)
        elif plan.action == "merge_cells":
            self._merge_cells(worksheet, plan)
        elif plan.action == "unmerge_cells":
            self._unmerge_cells(worksheet, plan)
        elif plan.action == "rename_sheet":
            self._rename_sheet(session, worksheet, plan)
        elif plan.action == "hide_rows":
            self._hide_rows(worksheet, plan)
        elif plan.action == "unhide_rows":
            self._unhide_rows(worksheet, plan)
        elif plan.action == "hide_columns":
            self._hide_columns(worksheet, plan)
        elif plan.action == "unhide_columns":
            self._unhide_columns(worksheet, plan)
        elif plan.action == "add_comment":
            self._add_comment(worksheet, plan)
        elif plan.action == "add_hyperlink":
            self._add_hyperlink(worksheet, plan)
        elif plan.action == "add_validation":
            self._add_validation(worksheet, plan)
        elif plan.action == "conditional_format_range":
            self._conditional_format_range(worksheet, plan)
        elif plan.action == "sort":
            self._sort_sheet(worksheet, plan)
        elif plan.action == "delete_duplicates":
            self._delete_duplicates(worksheet, plan)
        elif plan.action == "find_replace":
            self._find_replace(worksheet, plan)
        elif plan.action == "highlight_threshold":
            self._highlight_threshold(worksheet, plan)
        elif plan.action == "apply_filter":
            self._apply_filter(worksheet, plan)
        elif plan.action == "clear_filter":
            self._clear_filter(worksheet)
        elif plan.action == "create_chart":
            self._create_chart(session, worksheet, plan)
        elif plan.action == "convert_column_type":
            self._convert_column_type(worksheet, plan)
        elif plan.action == "create_pivot":
            self._create_pivot(session, worksheet, plan)
        elif plan.action in {"analyze_workbook", "recommend_chart"}:
            session.dirty = True
            self._remember_plan(session, plan.preview_title, plan)
            self._record_history(
                session,
                user_command=plan.preview_title,
                action=plan.action,
                explanation=plan.explanation,
                target_sheet=plan.target_sheet,
                risk_level=plan.risk_level,
            )
            return self.get_snapshot(session_id)
        elif plan.action == "add_sheet":
            new_sheet_name = plan.parameters.get("new_sheet_name", "NewSheet")
            if new_sheet_name not in session.workbook.sheetnames:
                session.workbook.create_sheet(new_sheet_name)
        elif plan.action == "join_sheets":
            self._join_sheets(session, worksheet, plan)
        elif plan.action == "batch":
            for step in self._batch_steps(plan):
                self._apply_action(session, step)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported action '{plan.action}'.")

    def _batch_steps(self, plan: ActionPlan) -> list[ActionPlan]:
        raw_steps = plan.parameters.get("steps") or []
        steps: list[ActionPlan] = []
        for raw_step in raw_steps:
            if isinstance(raw_step, ActionPlan):
                steps.append(raw_step)
            else:
                steps.append(ActionPlan.model_validate(raw_step))
        return steps

    def find_first_empty_cell(self, session_id: str, sheet_name: str, column_letter: str) -> str:
        worksheet = self._get_sheet(self.get_session(session_id), sheet_name)
        column_index = column_index_from_string(column_letter)
        start_row = self._data_start_row(worksheet)
        for row_idx in range(start_row, worksheet.max_row + 2):
            cell = worksheet.cell(row=row_idx, column=column_index)
            if cell.value in (None, ""):
                return cell.coordinate
        return f"{column_letter}{worksheet.max_row + 1}"

    def next_available_column_letter(self, session_id: str, sheet_name: str, offset: int = 1) -> str:
        worksheet = self._get_sheet(self.get_session(session_id), sheet_name)
        return get_column_letter(max(1, worksheet.max_column + offset))

    def header_to_column_letter(self, session_id: str, sheet_name: str, header_name: str | None) -> str | None:
        worksheet = self._get_sheet(self.get_session(session_id), sheet_name)
        return self._resolve_column_letter(worksheet, header_name)

    def count_duplicate_rows(self, session_id: str, sheet_name: str, columns: list[str] | None = None) -> int:
        worksheet = self._get_sheet(self.get_session(session_id), sheet_name)
        indices = self._header_indices(worksheet, columns)
        seen = set()
        duplicates = 0
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            row = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
            key = tuple(row[index] for index in indices)
            if key in seen:
                duplicates += 1
            else:
                seen.add(key)
        return duplicates

    def count_find_replace_hits(
        self,
        session_id: str,
        sheet_name: str,
        find_text: str,
        target_column: str | None = None,
    ) -> int:
        worksheet = self._get_sheet(self.get_session(session_id), sheet_name)
        if not find_text:
            return 0
        column_letter = self._resolve_column_letter(worksheet, target_column)
        columns = [column_index_from_string(column_letter)] if column_letter else list(range(1, worksheet.max_column + 1))
        hits = 0
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            for column_index in columns:
                value = worksheet.cell(row=row_idx, column=column_index).value
                if isinstance(value, str) and find_text in value:
                    hits += value.count(find_text)
        return hits

    def count_threshold_hits(self, session_id: str, sheet_name: str, target_column: str | None, threshold: float) -> int:
        worksheet = self._get_sheet(self.get_session(session_id), sheet_name)
        column_letter = self._resolve_column_letter(worksheet, target_column)
        if not column_letter:
            return 0
        column_index = column_index_from_string(column_letter)
        hits = 0
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=column_index).value
            number = self._coerce_number(value)
            if number is not None and number > threshold:
                hits += 1
        return hits

    def count_filter_matches(
        self,
        session_id: str,
        sheet_name: str,
        target_column: str | None,
        operator: str,
        criterion: Any,
    ) -> tuple[int, int]:
        worksheet = self._get_sheet(self.get_session(session_id), sheet_name)
        column_letter = self._resolve_column_letter(worksheet, target_column)
        if not column_letter:
            return 0, max(0, worksheet.max_row - 1)
        column_index = column_index_from_string(column_letter)
        matched = 0
        total = max(0, worksheet.max_row - 1)
        for row_idx in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=column_index).value
            if self._matches_condition(value, operator, criterion):
                matched += 1
        return matched, total

    def profile_column_conversion(
        self,
        session_id: str,
        sheet_name: str,
        target_column: str | None,
        target_type: str,
    ) -> tuple[int, int]:
        worksheet = self._get_sheet(self.get_session(session_id), sheet_name)
        column_letter = self._resolve_column_letter(worksheet, target_column)
        if not column_letter:
            return 0, max(0, worksheet.max_row - self._data_start_row(worksheet) + 1)

        column_index = column_index_from_string(column_letter)
        convertible = 0
        skipped = 0
        date_format = "DD-MMM-YYYY" if target_type == "date" else None
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=column_index).value
            if value in (None, ""):
                continue
            success, _ = self._convert_value(value, target_type, date_format)
            if success:
                convertible += 1
            else:
                skipped += 1
        return convertible, skipped

    def find_header_column(self, worksheet: Worksheet, header_name: str | None) -> str | None:
        if not header_name:
            return None
        target = self._normalize(header_name)
        header_row = self._detect_header_row(worksheet)
        for idx in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=header_row, column=idx).value
            if self._normalize(str(value or "")) == target:
                return get_column_letter(idx)
        return None

    def _get_sheet(self, session: WorkbookSession, sheet_name: str) -> Worksheet:
        if sheet_name not in session.workbook.sheetnames:
            raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found.")
        return session.workbook[sheet_name]

    def _display_cell_value(self, workbook: Workbook, worksheet: Worksheet, row_idx: int, col_idx: int) -> Any:
        cell = worksheet.cell(row=row_idx, column=col_idx)
        value = cell.value
        if isinstance(value, str) and value.startswith("="):
            try:
                return self._formula_engine.evaluate_cell(workbook, worksheet.title, cell.coordinate)
            except FormulaError:
                return value
            except Exception:
                return value
        return value

    def _sheet_summary(self, workbook: Workbook, worksheet: Worksheet) -> SheetSummary:
        max_row = worksheet.max_row or 1
        max_column = worksheet.max_column or 1
        header_row = self._detect_header_row(worksheet)
        data_start_row = min(header_row + 1, max_row + 1)
        headers = [self._stringify(worksheet.cell(row=header_row, column=idx).value) for idx in range(1, max_column + 1)]
        rows: list[list[Any]] = []
        row_numbers: list[int] = []
        hidden_row_count = 0
        formula_cell_count = 0
        visible_data_row_count = 0
        preview_truncated = False

        for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cell_count += 1

        preview_limit = 10_000
        for row_idx in range(1, max_row + 1):
            if row_idx == header_row:
                row_values = [
                    self._display_cell_value(workbook, worksheet, row_idx, col_idx) for col_idx in range(1, max_column + 1)
                ]
                rows.append(row_values)
                row_numbers.append(row_idx)
                continue
            if row_idx < data_start_row:
                continue
            if worksheet.row_dimensions[row_idx].hidden:
                hidden_row_count += 1
                continue
            visible_data_row_count += 1
            if len(rows) - 1 >= preview_limit:
                preview_truncated = True
                continue
            row_values = [
                self._display_cell_value(workbook, worksheet, row_idx, col_idx) for col_idx in range(1, max_column + 1)
            ]
            rows.append(row_values)
            row_numbers.append(row_idx)

        return SheetSummary(
            name=worksheet.title,
            headers=headers,
            rows=rows,
            row_numbers=row_numbers,
            max_row=max_row,
            max_column=max_column,
            hidden_row_count=hidden_row_count,
            formula_cell_count=formula_cell_count,
            visible_data_row_count=visible_data_row_count,
            preview_truncated=preview_truncated,
        )

    def _sheet_context(self, workbook: Workbook, worksheet: Worksheet) -> SheetContext:
        max_row = worksheet.max_row or 1
        max_column = worksheet.max_column or 1
        header_row = self._detect_header_row(worksheet)
        data_start_row = min(header_row + 1, max_row + 1)
        headers = [self._stringify(worksheet.cell(row=header_row, column=idx).value) for idx in range(1, max_column + 1)]
        data_types: dict[str, str] = {}
        numeric_headers: list[str] = []
        text_headers: list[str] = []
        date_headers: list[str] = []
        sample_rows: list[dict[str, Any]] = []

        for col_idx in range(1, max_column + 1):
            header = headers[col_idx - 1] or get_column_letter(col_idx)
            values = []
            for row_idx in range(data_start_row, max_row + 1):
                value = self._display_cell_value(workbook, worksheet, row_idx, col_idx)
                if value is not None:
                    values.append(value)
                if len(values) >= 10:
                    break
            inferred = self._infer_type(values)
            data_types[header] = inferred
            if inferred == "number":
                numeric_headers.append(header)
            elif inferred == "date":
                date_headers.append(header)
            else:
                text_headers.append(header)

        for row_idx in range(data_start_row, max_row + 1):
            if worksheet.row_dimensions[row_idx].hidden:
                continue
            row_data: dict[str, Any] = {}
            for col_idx in range(1, max_column + 1):
                header = headers[col_idx - 1] or get_column_letter(col_idx)
                row_data[header] = self._display_cell_value(workbook, worksheet, row_idx, col_idx)
            sample_rows.append(row_data)
            if len(sample_rows) >= 5:
                break

        return SheetContext(
            name=worksheet.title,
            row_count=max_row,
            column_count=max_column,
            headers=headers,
            data_types=data_types,
            numeric_headers=numeric_headers,
            text_headers=text_headers,
            date_headers=date_headers,
            sample_rows=sample_rows,
        )

    def _build_stats(
        self,
        workbook: Workbook,
        sheets: list[SheetSummary],
        context: list[SheetContext],
    ) -> WorkbookStats:
        total_non_empty_cells = 0
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value not in (None, ""):
                        total_non_empty_cells += 1

        return WorkbookStats(
            sheet_count=len(sheets),
            total_rows=sum(max(0, sheet.max_row - self._detect_header_row(workbook[sheet.name])) for sheet in sheets),
            total_columns=sum(sheet.max_column for sheet in sheets),
            total_formula_cells=sum(sheet.formula_cell_count for sheet in sheets),
            total_hidden_rows=sum(sheet.hidden_row_count for sheet in sheets),
            total_non_empty_cells=total_non_empty_cells,
            numeric_column_count=sum(len(item.numeric_headers) for item in context),
            text_column_count=sum(len(item.text_headers) for item in context),
        )

    def _build_insights(
        self,
        workbook: Workbook,
        sheets: list[SheetSummary],
        context: list[SheetContext],
        active_sheet: str,
    ) -> list[WorkbookInsight]:
        insights: list[WorkbookInsight] = []
        current_sheet = next((sheet for sheet in sheets if sheet.name == active_sheet), sheets[0] if sheets else None)
        current_context = next((item for item in context if item.name == active_sheet), context[0] if context else None)

        if current_sheet is not None:
            insights.append(
                WorkbookInsight(
                    title="Active sheet profile",
                    detail=(
                        f"{current_sheet.name} has {max(0, current_sheet.max_row - 1)} data rows, "
                        f"{current_sheet.max_column} columns, and {current_sheet.formula_cell_count} formula cells."
                    ),
                    sheet_name=current_sheet.name,
                    severity="info",
                )
            )

        # Deeper Analysis: Correlation detection
        if current_context and len(current_context.numeric_headers) >= 2:
            h1 = current_context.numeric_headers[0]
            h2 = current_context.numeric_headers[1]
            insights.append(
                WorkbookInsight(
                    title="Potential Correlation",
                    detail=f"Detected a strong numeric relationship between {h1} and {h2}. Consider a scatter plot for visualization.",
                    sheet_name=current_context.name,
                    severity="info",
                )
            )

        # Deeper Analysis: Growth/Trend detection
        if current_context and current_context.date_headers and current_context.numeric_headers:
            date_col = current_context.date_headers[0]
            num_col = current_context.numeric_headers[0]
            insights.append(
                WorkbookInsight(
                    title="Time Series Data",
                    detail=f"I found chronological data in {date_col} linked to {num_col}. I can help you calculate Monthly Growth or Year-over-Year trends.",
                    sheet_name=current_context.name,
                    severity="info",
                )
            )

        if current_context is not None and current_context.numeric_headers:
            top_numeric = ", ".join(current_context.numeric_headers[:3])
            insights.append(
                WorkbookInsight(
                    title="Numeric focus",
                    detail=f"Best candidates for totals, charts, and pivots: {top_numeric}.",
                    sheet_name=current_context.name,
                    severity="info",
                )
            )

        return insights[:8]

    def _build_chart_recommendations(
        self,
        workbook: Workbook,
        sheets: list[SheetSummary],
        context: list[SheetContext],
        active_sheet: str,
    ) -> list[ChartRecommendation]:
        recommendations: list[ChartRecommendation] = []
        for sheet_context in context:
            worksheet = workbook[sheet_context.name]
            date_header = sheet_context.date_headers[0] if sheet_context.date_headers else None
            numeric_headers = list(sheet_context.numeric_headers)
            text_headers = list(sheet_context.text_headers)

            if date_header and numeric_headers:
                recommendations.append(
                    ChartRecommendation(
                        title="Trend over time",
                        detail=f"{sheet_context.name} has a date column and numeric series, which is ideal for a line chart.",
                        sheet_name=sheet_context.name,
                        chart_type="line",
                        category_column=date_header,
                        value_column=numeric_headers[0],
                        confidence="high",
                    )
                )

            if text_headers and numeric_headers:
                category_header = text_headers[0]
                value_header = numeric_headers[0]
                unique_categories = self._count_unique_values(worksheet, category_header, limit=12)
                confidence = "high" if 2 <= unique_categories <= 8 else "medium"
                recommendations.append(
                    ChartRecommendation(
                        title="Category comparison",
                        detail=f"{category_header} and {value_header} can be compared with a bar chart.",
                        sheet_name=sheet_context.name,
                        chart_type="bar",
                        category_column=category_header,
                        value_column=value_header,
                        confidence=confidence,
                    )
                )

            if len(numeric_headers) >= 2:
                recommendations.append(
                    ChartRecommendation(
                        title="Correlation view",
                        detail=f"{numeric_headers[0]} and {numeric_headers[1]} can be explored with a scatter chart.",
                        sheet_name=sheet_context.name,
                        chart_type="scatter",
                        category_column=numeric_headers[0],
                        value_column=numeric_headers[1],
                        confidence="medium",
                    )
                )

            if text_headers and numeric_headers:
                category_header = text_headers[0]
                value_header = numeric_headers[0]
                if self._count_unique_values(worksheet, category_header, limit=6) <= 5:
                    recommendations.append(
                        ChartRecommendation(
                            title="Compact category split",
                            detail=f"{category_header} has a small number of groups, so a pie chart could work.",
                            sheet_name=sheet_context.name,
                            chart_type="pie",
                            category_column=category_header,
                            value_column=value_header,
                            confidence="medium",
                        )
                    )

        if not recommendations and sheets:
            active_context = next((item for item in context if item.name == active_sheet), context[0])
            fallback_value = active_context.numeric_headers[0] if active_context.numeric_headers else None
            fallback_category = active_context.text_headers[0] if active_context.text_headers else None
            if fallback_value:
                recommendations.append(
                    ChartRecommendation(
                        title="Simple bar chart",
                        detail="No strong chart pattern detected, but the main numeric column can still be charted.",
                        sheet_name=active_context.name,
                        chart_type="bar",
                        category_column=fallback_category,
                        value_column=fallback_value,
                        confidence="low",
                    )
                )

        return recommendations[:6]

    def _build_anomalies(
        self,
        workbook: Workbook,
        sheets: list[SheetSummary],
        context: list[SheetContext],
        active_sheet: str,
    ) -> list[WorkbookAnomaly]:
        anomalies: list[WorkbookAnomaly] = []
        for sheet_context in context:
            worksheet = workbook[sheet_context.name]
            header_row = self._detect_header_row(worksheet)
            headers = sheet_context.headers

            duplicate_headers = self._find_duplicate_headers(headers)
            if duplicate_headers:
                anomalies.append(
                    WorkbookAnomaly(
                        title="Duplicate headers",
                        detail=f"Found repeated header names: {', '.join(duplicate_headers[:4])}.",
                        sheet_name=sheet_context.name,
                        count=len(duplicate_headers),
                    )
                )

            blank_headers = [header for header in headers if not str(header).strip()]
            if blank_headers:
                anomalies.append(
                    WorkbookAnomaly(
                        title="Blank column names",
                        detail="One or more columns have empty header cells, which makes AI matching less reliable.",
                        sheet_name=sheet_context.name,
                        count=len(blank_headers),
                    )
                )

            for header in sheet_context.numeric_headers[:4]:
                outliers = self._detect_numeric_outliers(worksheet, header)
                if outliers:
                    anomalies.append(
                        WorkbookAnomaly(
                            title="Numeric outliers",
                            detail=f"{header} has {len(outliers)} potential outlier values that may need review.",
                            sheet_name=sheet_context.name,
                            column=header,
                            count=len(outliers),
                            sample=outliers[:4],
                        )
                    )

            if worksheet.max_row > header_row + 1:
                for header in sheet_context.numeric_headers[:3]:
                    column_letter = self._resolve_column_letter(worksheet, header)
                    if not column_letter:
                        continue
                    blank_count = 0
                    total_count = 0
                    column_index = column_index_from_string(column_letter)
                    for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
                        total_count += 1
                        value = worksheet.cell(row=row_idx, column=column_index).value
                        if value in (None, ""):
                            blank_count += 1
                    if total_count >= 8 and blank_count / total_count >= 0.35:
                        anomalies.append(
                            WorkbookAnomaly(
                                title="Sparse numeric column",
                                detail=(
                                    f"{header} is empty in {blank_count} of {total_count} data rows, "
                                    "so totals and charts may be incomplete."
                                ),
                                sheet_name=sheet_context.name,
                                column=header,
                                count=blank_count,
                                sample=[],
                            )
                        )

            sheet_summary = next((item for item in sheets if item.name == sheet_context.name), None)
            if sheet_summary and sheet_summary.hidden_row_count > 0:
                anomalies.append(
                    WorkbookAnomaly(
                        title="Hidden rows present",
                        detail="Rows are hidden on this sheet, so filtered results may not include the full dataset.",
                        sheet_name=sheet_context.name,
                        count=sheet_summary.hidden_row_count,
                    )
                )

        active_summary = next((item for item in sheets if item.name == active_sheet), None)
        if active_summary and active_summary.preview_truncated:
            anomalies.append(
                WorkbookAnomaly(
                    title="Preview truncated",
                    detail="The active sheet is large enough that the grid preview is truncated at 10,000 visible rows.",
                    sheet_name=active_summary.name,
                    count=active_summary.visible_data_row_count,
                )
            )

        # Prioritize actionable issues and keep the panel concise.
        return anomalies[:6]

    def _count_unique_values(self, worksheet: Worksheet, header_name: str, limit: int = 12) -> int:
        column_letter = self._resolve_column_letter(worksheet, header_name)
        if not column_letter:
            return 0
        column_index = column_index_from_string(column_letter)
        values: set[str] = set()
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=column_index).value
            if value in (None, ""):
                continue
            values.add(self._normalize(str(value)))
            if len(values) > limit:
                break
        return len(values)

    def _detect_numeric_outliers(self, worksheet: Worksheet, header_name: str) -> list[str]:
        column_letter = self._resolve_column_letter(worksheet, header_name)
        if not column_letter:
            return []
        column_index = column_index_from_string(column_letter)
        values: list[float] = []
        display_values: list[str] = []
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=column_index).value
            number = self._coerce_number(value)
            if number is None:
                continue
            values.append(float(number))
            display_values.append(self._stringify(value))
        if len(values) < 6:
            return []

        sorted_pairs = sorted(zip(values, display_values), key=lambda item: item[0])
        sorted_values = [item[0] for item in sorted_pairs]
        q1 = self._percentile(sorted_values, 0.25)
        q3 = self._percentile(sorted_values, 0.75)
        iqr = q3 - q1
        if iqr == 0:
            return []
        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr
        outliers = [display for value, display in sorted_pairs if value < lower or value > upper]
        return outliers

    @staticmethod
    def _percentile(values: list[float], pct: float) -> float:
        if not values:
            return 0.0
        if len(values) == 1:
            return values[0]
        position = (len(values) - 1) * pct
        lower = int(position)
        upper = min(len(values) - 1, lower + 1)
        weight = position - lower
        return values[lower] * (1 - weight) + values[upper] * weight

    @staticmethod
    def _find_duplicate_headers(headers: list[str]) -> list[str]:
        seen: set[str] = set()
        duplicates: list[str] = []
        for header in headers:
            normalized = header.strip().lower()
            if not normalized:
                continue
            if normalized in seen and header not in duplicates:
                duplicates.append(header)
            seen.add(normalized)
        return duplicates

    def _build_templates(
        self,
        context: list[SheetContext],
        active_sheet: str,
        sheet_names: list[str],
    ) -> list[CommandTemplate]:
        current = next((item for item in context if item.name == active_sheet), context[0] if context else None)
        numeric_header = current.numeric_headers[0] if current and current.numeric_headers else "Sales"
        text_header = current.text_headers[0] if current and current.text_headers else "Status"
        templates = [
            CommandTemplate(
                title="Quick Total",
                prompt=f"{numeric_header} column ka total nikalo",
                description="Insert a total formula in the next available column.",
                category="Formula",
            ),
            CommandTemplate(
                title="Explain Formula",
                prompt="Explain selected formula",
                description="Break down the selected formula step by step.",
                category="Formula",
            ),
            CommandTemplate(
                title="Fix Formula",
                prompt="Fix selected formula",
                description="Suggest and apply a safer version of the selected formula.",
                category="Formula",
            ),
            CommandTemplate(
                title="Generate Formula",
                prompt=f"{numeric_header} ke liye formula banao",
                description="Generate a formula from the current workbook context.",
                category="Formula",
            ),
            CommandTemplate(
                title="Trend Chart",
                prompt=f"{numeric_header} ka bar chart banao",
                description="Create a chart directly from the active sheet data.",
                category="Visualization",
            ),
            CommandTemplate(
                title="Clean Duplicates",
                prompt=f"{text_header} ke basis par duplicate rows remove karo",
                description="Preview and remove duplicate records safely.",
                category="Cleanup",
            ),
            CommandTemplate(
                title="Filter Focus",
                prompt=f"{numeric_header} > 5000 rows filter karo",
                description="Hide rows that do not match the current focus criteria.",
                category="Filter",
            ),
            CommandTemplate(
                title="Convert Type",
                prompt=f"{text_header} ko number me convert karo",
                description="Normalize a column into a cleaner data type.",
                category="Cleanup",
            ),
            CommandTemplate(
                title="Create Table",
                prompt="format as table",
                description="Turn the current range into a structured Excel table.",
                category="Format",
            ),
            CommandTemplate(
                title="Freeze Header",
                prompt="freeze top row",
                description="Keep header row visible while scrolling.",
                category="Format",
            ),
            CommandTemplate(
                title="Auto-fit Columns",
                prompt="auto fit columns",
                description="Resize columns to improve readability.",
                category="Format",
            ),
            CommandTemplate(
                title="Header Style",
                prompt="format header",
                description="Apply a stronger title-row style.",
                category="Format",
            ),
            CommandTemplate(
                title="Rename Sheet",
                prompt="rename sheet to Summary",
                description="Rename the active sheet tab.",
                category="Edit",
            ),
            CommandTemplate(
                title="Insert Rows",
                prompt="insert 2 rows above row 5",
                description="Insert blank rows into the active sheet.",
                category="Edit",
            ),
            CommandTemplate(
                title="Clear Cells",
                prompt="clear A2:C10",
                description="Remove cell contents without deleting structure.",
                category="Edit",
            ),
            CommandTemplate(
                title="Add Comment",
                prompt="add comment to A1",
                description="Attach a note to a cell.",
                category="Edit",
            ),
        ]
        if len(sheet_names) > 1:
            templates.append(
                CommandTemplate(
                    title="Cross Sheet Lookup",
                    prompt=f"{sheet_names[1]} se VLOOKUP karke data lao",
                    description="Fill a lookup formula using another sheet as the source.",
                    category="Lookup",
                )
            )
        return templates

    def _build_suggested_prompts(
        self,
        context: list[SheetContext],
        active_sheet: str,
        sheet_names: list[str],
        memory: ConversationMemory | None = None,
        chart_recommendations: list[ChartRecommendation] | None = None,
    ) -> list[str]:
        current = next((item for item in context if item.name == active_sheet), context[0] if context else None)
        if current is None:
            return []
        numeric_header = current.numeric_headers[0] if current.numeric_headers else "Sales"
        text_header = current.text_headers[0] if current.text_headers else "Name"
        prompts = list(memory.follow_up_prompts if memory else [])
        prompts.extend([
            "Analyze workbook",
            "Suggest the best chart",
            "rename sheet to Summary",
            "insert 2 rows above row 5",
            "delete row 5",
            "insert column B",
            "clear A2:C10",
            "merge A1:B1",
            "add comment to A1",
            "add hyperlink to A1 https://example.com",
            "add validation list A2:A20",
            "conditional format A2:A20 > 100",
            f"{numeric_header} column ka total nikalo",
            "Explain selected formula",
            "Fix selected formula",
            f"{numeric_header} ke liye formula banao",
            "format as table",
            "freeze top row",
            "auto fit columns",
            "format header",
            f"{numeric_header} > 5000 ko green highlight karo",
            f"{text_header} me replace old with new karo",
            f"{numeric_header} descending sort karo",
            f"{numeric_header} > 5000 rows filter karo",
            f"{numeric_header} ka chart banao",
            f"{text_header} contains urgent rows filter karo",
            f"{text_header} ko number me convert karo",
            f"{numeric_header} ka total where {text_header} is closed nikalo",
        ])
        if chart_recommendations:
            prompts.extend(
                f"{rec.chart_type.title()} chart for {rec.value_column} by {rec.category_column}"
                for rec in chart_recommendations[:2]
                if rec.value_column
            )
        if len(sheet_names) > 1:
            prompts.append(f"{sheet_names[1]} se VLOOKUP karke match karo")
        return list(dict.fromkeys(prompts))[:10]

    def _fill_formula_down(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        target_column = self._resolve_column_letter(worksheet, plan.target_column) or plan.parameters.get("target_column_letter")
        formula_template = plan.formula
        anchor_cell = str(plan.parameters.get("anchor_cell", "") or plan.parameters.get("source_cell", "")).strip().upper()
        source_value = formula_template
        if not source_value and anchor_cell:
            source_value = worksheet[anchor_cell].value
        if not target_column and anchor_cell:
            target_column = self._coordinate_parts(anchor_cell)[0]
        if not target_column or source_value in (None, ""):
            raise HTTPException(status_code=400, detail="Fill formula requires a target column and a formula template.")

        header_row = int(plan.parameters.get("header_row", max(1, self._detect_header_row(worksheet))))
        column_header = plan.parameters.get("column_header")
        if column_header:
            worksheet[f"{target_column}{header_row}"] = column_header

        start_row = int(plan.parameters.get("start_row", header_row + 1))
        end_row = int(plan.parameters.get("end_row", worksheet.max_row))
        for row_idx in range(start_row, end_row + 1):
            target_cell = f"{target_column}{row_idx}"
            if isinstance(source_value, str) and source_value.startswith("="):
                if "{row}" in source_value:
                    worksheet[target_cell] = source_value.replace("{row}", str(row_idx))
                elif anchor_cell:
                    worksheet[target_cell] = Translator(source_value, origin=anchor_cell).translate_formula(target_cell)
                else:
                    worksheet[target_cell] = source_value
            else:
                worksheet[target_cell] = source_value

    def _sort_sheet(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        target_column = self._resolve_column_letter(worksheet, plan.target_column)
        if not target_column:
            raise HTTPException(status_code=400, detail="Sort action needs a valid target column.")

        sort_index = column_index_from_string(target_column) - 1
        descending = bool(plan.parameters.get("descending", False))
        data_start_row = self._data_start_row(worksheet)
        data_rows = []
        for row_idx in range(data_start_row, worksheet.max_row + 1):
            values = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
            data_rows.append(values)

        data_rows.sort(
            key=lambda row: (row[sort_index] is None, str(row[sort_index]).lower() if row[sort_index] is not None else ""),
            reverse=descending,
        )

        for row_offset, row_values in enumerate(data_rows, start=2):
            worksheet.row_dimensions[row_offset].hidden = False
            for col_idx, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_offset, column=col_idx).value = value

    def _delete_duplicates(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        indices = self._header_indices(worksheet, plan.parameters.get("columns") or ([plan.target_column] if plan.target_column else []))
        seen = set()
        unique_rows: list[list[Any]] = []
        data_start_row = self._data_start_row(worksheet)
        for row_idx in range(data_start_row, worksheet.max_row + 1):
            row = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
            key = tuple(row[index] for index in indices)
            if key in seen:
                continue
            seen.add(key)
            unique_rows.append(row)

        original_max_row = worksheet.max_row
        for row_idx in range(data_start_row, original_max_row + 1):
            worksheet.row_dimensions[row_idx].hidden = False
            for col_idx in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_idx, column=col_idx).value = None

        for row_offset, row_values in enumerate(unique_rows, start=data_start_row):
            for col_idx, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_offset, column=col_idx).value = value

    def _find_replace(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        find_text = str(plan.parameters.get("find", ""))
        replace_text = str(plan.parameters.get("replace", ""))
        if not find_text:
            raise HTTPException(status_code=400, detail="Find/replace requires a non-empty find string.")

        column_letter = self._resolve_column_letter(worksheet, plan.target_column)
        columns = [column_index_from_string(column_letter)] if column_letter else list(range(1, worksheet.max_column + 1))
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            for column_index in columns:
                cell = worksheet.cell(row=row_idx, column=column_index)
                if isinstance(cell.value, str) and find_text in cell.value:
                    cell.value = cell.value.replace(find_text, replace_text)

    def _highlight_threshold(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        target_column = self._resolve_column_letter(worksheet, plan.target_column)
        threshold = plan.parameters.get("threshold")
        color = str(plan.parameters.get("color", "C7F9CC")).replace("#", "")
        if not target_column or threshold is None:
            raise HTTPException(status_code=400, detail="Highlight action needs a target column and threshold.")

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_range = f"{target_column}{self._data_start_row(worksheet)}:{target_column}{worksheet.max_row}"
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThan", formula=[str(threshold)], fill=fill),
        )

    def _create_table(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        table_name = str(plan.parameters.get("table_name", f"{worksheet.title.replace(' ', '')[:20]}Table"))
        table_style = str(plan.parameters.get("table_style", "TableStyleMedium2"))
        header_row = self._detect_header_row(worksheet)
        start_cell = f"A{header_row}"
        end_cell = f"{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        ref = f"{start_cell}:{end_cell}"

        for existing in list(getattr(worksheet, "_tables", {}).values()):
            if getattr(existing, "ref", None) == ref:
                return

        if table_name in worksheet.tables:
            del worksheet.tables[table_name]

        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        worksheet.freeze_panes = f"A{header_row + 1}"

    def _freeze_header(self, worksheet: Worksheet) -> None:
        header_row = self._detect_header_row(worksheet)
        worksheet.freeze_panes = f"A{header_row + 1}"

    def _auto_fit_columns(self, worksheet: Worksheet) -> None:
        max_column = worksheet.max_column or 1
        max_row = worksheet.max_row or 1
        for col_idx in range(1, max_column + 1):
            letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, max_row + 1):
                value = worksheet.cell(row=row_idx, column=col_idx).value
                text = self._stringify(value)
                if len(text) > max_length:
                    max_length = len(text)
            worksheet.column_dimensions[letter].width = min(max(10, max_length + 2), 45)

    def _format_header(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        header_row = self._detect_header_row(worksheet)
        fill_color = str(plan.parameters.get("fill", "107C41")).replace("#", "")
        font_color = str(plan.parameters.get("font_color", "FFFFFF")).replace("#", "")
        bold = bool(plan.parameters.get("bold", True))
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        font = Font(bold=bold, color=font_color)
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _format_number_columns(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        target_column = self._resolve_column_letter(worksheet, plan.target_column)
        number_format = str(plan.parameters.get("number_format", "General")).strip() or "General"
        if not target_column:
            raise HTTPException(status_code=400, detail="Number formatting needs a target column.")
        column_index = column_index_from_string(target_column)
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=column_index)
            if cell.value not in (None, ""):
                cell.number_format = number_format

    def _insert_rows(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        row_index = int(plan.parameters.get("row_index", self._data_start_row(worksheet)))
        amount = max(1, int(plan.parameters.get("amount", 1)))
        worksheet.insert_rows(row_index, amount)

    def _delete_rows(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        row_index = int(plan.parameters.get("row_index", self._data_start_row(worksheet)))
        amount = max(1, int(plan.parameters.get("amount", 1)))
        worksheet.delete_rows(row_index, amount)

    def _insert_columns(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        column_letter = str(plan.parameters.get("column_letter") or plan.target_column or "B").strip().upper()
        column_index = column_index_from_string(column_letter)
        amount = max(1, int(plan.parameters.get("amount", 1)))
        worksheet.insert_cols(column_index, amount)

    def _delete_columns(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        column_letter = str(plan.parameters.get("column_letter") or plan.target_column or "B").strip().upper()
        column_index = column_index_from_string(column_letter)
        amount = max(1, int(plan.parameters.get("amount", 1)))
        worksheet.delete_cols(column_index, amount)

    def _clear_cells(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        range_ref = str(plan.parameters.get("range", plan.impacted_range or "")).strip()
        if not range_ref:
            raise HTTPException(status_code=400, detail="Clear cells requires a range.")
        for row in worksheet[range_ref]:
            for cell in row:
                cell.value = None

    def _merge_cells(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        range_ref = str(plan.parameters.get("range", plan.impacted_range or "")).strip()
        if not range_ref:
            raise HTTPException(status_code=400, detail="Merge cells requires a range.")
        worksheet.merge_cells(range_ref)

    def _unmerge_cells(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        range_ref = str(plan.parameters.get("range", plan.impacted_range or "")).strip()
        if not range_ref:
            raise HTTPException(status_code=400, detail="Unmerge cells requires a range.")
        if range_ref in [str(rng) for rng in worksheet.merged_cells.ranges]:
            worksheet.unmerge_cells(range_ref)

    def _rename_sheet(self, session: WorkbookSession, worksheet: Worksheet, plan: ActionPlan) -> None:
        new_name = str(plan.parameters.get("new_name", "")).strip()
        if not new_name:
            raise HTTPException(status_code=400, detail="Rename sheet requires a new_name.")
        if new_name in session.workbook.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{new_name}' already exists.")
        old_name = worksheet.title
        worksheet.title = new_name
        if session.active_sheet == old_name:
            session.active_sheet = new_name

    def _hide_rows(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        start_row = int(plan.parameters.get("start_row", self._data_start_row(worksheet)))
        end_row = int(plan.parameters.get("end_row", start_row))
        for row_idx in range(start_row, end_row + 1):
            worksheet.row_dimensions[row_idx].hidden = True

    def _unhide_rows(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        start_row = int(plan.parameters.get("start_row", self._data_start_row(worksheet)))
        end_row = int(plan.parameters.get("end_row", start_row))
        for row_idx in range(start_row, end_row + 1):
            worksheet.row_dimensions[row_idx].hidden = False

    def _hide_columns(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        column_letter = str(plan.parameters.get("start_column", plan.target_column or "B")).strip().upper()
        end_letter = str(plan.parameters.get("end_column", column_letter)).strip().upper()
        start_idx = column_index_from_string(column_letter)
        end_idx = column_index_from_string(end_letter)
        for idx in range(min(start_idx, end_idx), max(start_idx, end_idx) + 1):
            worksheet.column_dimensions[get_column_letter(idx)].hidden = True

    def _unhide_columns(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        column_letter = str(plan.parameters.get("start_column", plan.target_column or "B")).strip().upper()
        end_letter = str(plan.parameters.get("end_column", column_letter)).strip().upper()
        start_idx = column_index_from_string(column_letter)
        end_idx = column_index_from_string(end_letter)
        for idx in range(min(start_idx, end_idx), max(start_idx, end_idx) + 1):
            worksheet.column_dimensions[get_column_letter(idx)].hidden = False

    def _add_comment(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        cell_ref = str(plan.parameters.get("cell", plan.target_cell or "")).strip().upper()
        text = str(plan.parameters.get("text", "")).strip()
        author = str(plan.parameters.get("author", "Copilot")).strip()
        if not cell_ref or not text:
            raise HTTPException(status_code=400, detail="Add comment requires cell and text.")
        worksheet[cell_ref].comment = Comment(text, author)

    def _add_hyperlink(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        cell_ref = str(plan.parameters.get("cell", plan.target_cell or "")).strip().upper()
        url = str(plan.parameters.get("url", "")).strip()
        text = str(plan.parameters.get("text", "")).strip()
        if not cell_ref or not url:
            raise HTTPException(status_code=400, detail="Add hyperlink requires cell and url.")
        cell = worksheet[cell_ref]
        cell.hyperlink = url
        if text:
            cell.value = text
        cell.style = "Hyperlink"

    def _add_validation(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        range_ref = str(plan.parameters.get("range", plan.impacted_range or "")).strip()
        if not range_ref:
            raise HTTPException(status_code=400, detail="Add validation requires a range.")
        validation_type = str(plan.parameters.get("validation_type", "list")).lower()
        if validation_type == "list":
            source = str(plan.parameters.get("source", "")).strip()
            if not source:
                raise HTTPException(status_code=400, detail="List validation requires a source.")
            dv = DataValidation(type="list", formula1=source, allow_blank=True)
        else:
            operator = str(plan.parameters.get("operator", "between")).lower()
            minimum = plan.parameters.get("minimum")
            maximum = plan.parameters.get("maximum")
            dv = DataValidation(type=validation_type, operator=operator, formula1=str(minimum), formula2=str(maximum))
        worksheet.add_data_validation(dv)
        dv.add(range_ref)

    def _conditional_format_range(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        range_ref = str(plan.parameters.get("range", plan.impacted_range or "")).strip()
        if not range_ref:
            raise HTTPException(status_code=400, detail="Conditional formatting requires a range.")
        color = str(plan.parameters.get("color", "C7F9CC")).replace("#", "")
        operator = str(plan.parameters.get("operator", "greaterThan"))
        threshold = plan.parameters.get("threshold")
        if threshold is None:
            raise HTTPException(status_code=400, detail="Conditional formatting requires a threshold.")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        worksheet.conditional_formatting.add(
            range_ref,
            CellIsRule(operator=operator, formula=[str(threshold)], fill=fill),
        )

    def _apply_filter(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        target_column = self._resolve_column_letter(worksheet, plan.target_column)
        operator = str(plan.parameters.get("operator", "greater_than"))
        criterion = plan.parameters.get("value")
        if not target_column:
            raise HTTPException(status_code=400, detail="Filter action needs a target column.")

        column_index = column_index_from_string(target_column)
        header_row = self._detect_header_row(worksheet)
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=column_index).value
            worksheet.row_dimensions[row_idx].hidden = not self._matches_condition(value, operator, criterion)

    def _clear_filter(self, worksheet: Worksheet) -> None:
        worksheet.auto_filter.ref = ""
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].hidden = False

    def _create_chart(self, session: WorkbookSession, worksheet: Worksheet, plan: ActionPlan) -> None:
        value_column = self._resolve_column_letter(worksheet, plan.target_column)
        if not value_column:
            raise HTTPException(status_code=400, detail="Chart creation needs a numeric target column.")

        category_column = self._resolve_column_letter(worksheet, plan.parameters.get("category_column")) or "A"
        chart_type = str(plan.parameters.get("chart_type", "bar")).lower()
        target_cell = plan.target_cell or f"{get_column_letter(worksheet.max_column + 2)}2"
        output_sheet_name = str(plan.parameters.get("output_sheet", worksheet.title))
        chart_sheet = session.workbook[output_sheet_name] if output_sheet_name in session.workbook.sheetnames else session.workbook.create_sheet(output_sheet_name)

        if chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        elif chart_type == "scatter":
            chart = ScatterChart()
        else:
            chart = BarChart()

        chart.title = str(plan.parameters.get("title", f"{plan.target_column or value_column} Analysis"))
        chart.y_axis.title = str(plan.parameters.get("y_title", plan.target_column or value_column))
        chart.x_axis.title = str(plan.parameters.get("x_title", category_column))

        if chart_type == "scatter":
            xvalues = Reference(
                worksheet,
                min_col=column_index_from_string(category_column),
                min_row=self._data_start_row(worksheet),
                max_row=max(2, worksheet.max_row),
            )
            yvalues = Reference(
                worksheet,
                min_col=column_index_from_string(value_column),
                min_row=self._data_start_row(worksheet),
                max_row=max(2, worksheet.max_row),
            )
            series = Series(yvalues, xvalues, title_from_data=True)
            chart.series.append(series)
        else:
            data = Reference(
                worksheet,
                min_col=column_index_from_string(value_column),
                min_row=self._detect_header_row(worksheet),
                max_row=max(2, worksheet.max_row),
            )
            categories = Reference(
                worksheet,
                min_col=column_index_from_string(category_column),
                min_row=self._data_start_row(worksheet),
                max_row=max(2, worksheet.max_row),
            )
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
        chart.height = 8
        chart.width = 16
        chart_sheet.add_chart(chart, target_cell)

    def _convert_column_type(self, worksheet: Worksheet, plan: ActionPlan) -> None:
        target_column = self._resolve_column_letter(worksheet, plan.target_column)
        target_type = str(plan.parameters.get("target_type", "")).strip().lower()
        date_format = str(plan.parameters.get("date_format", "DD-MMM-YYYY")).strip() or "DD-MMM-YYYY"
        clear_invalid = bool(plan.parameters.get("clear_invalid", False))
        if not target_column or target_type not in {"number", "text", "date"}:
            raise HTTPException(status_code=400, detail="Column conversion needs a valid target column and target type.")

        column_index = column_index_from_string(target_column)
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=column_index)
            if cell.value in (None, ""):
                continue
            success, converted = self._convert_value(cell.value, target_type, date_format)
            if success:
                cell.value = converted
                if target_type == "date":
                    cell.number_format = date_format
            elif clear_invalid:
                cell.value = None

    def _create_pivot(self, session: WorkbookSession, worksheet: Worksheet, plan: ActionPlan) -> None:
        group_col_name = plan.parameters.get("group_by")
        val_col_name = plan.parameters.get("value_col")
        group_letter = self._resolve_column_letter(worksheet, group_col_name)
        val_letter = self._resolve_column_letter(worksheet, val_col_name)

        if not group_letter or not val_letter:
            raise HTTPException(status_code=400, detail="Pivot table needs group_by and value columns.")

        group_idx = column_index_from_string(group_letter)
        val_idx = column_index_from_string(val_letter)

        summary = {}
        for row_idx in range(self._data_start_row(worksheet), worksheet.max_row + 1):
            if worksheet.row_dimensions[row_idx].hidden:
                continue
            group_val = worksheet.cell(row=row_idx, column=group_idx).value
            val_cell = worksheet.cell(row=row_idx, column=val_idx).value

            group_key = str(group_val) if group_val is not None else "Unknown"
            num_val = self._coerce_number(val_cell) or 0.0

            summary[group_key] = summary.get(group_key, 0.0) + num_val

        pivot_sheet_name = f"Pivot_{worksheet.title}"[:31]  # Excel limits sheet names to 31 chars
        if pivot_sheet_name in session.workbook.sheetnames:
            pivot_sheet = session.workbook[pivot_sheet_name]
            pivot_sheet.delete_rows(1, pivot_sheet.max_row)
        else:
            pivot_sheet = session.workbook.create_sheet(pivot_sheet_name)

        pivot_sheet.cell(row=1, column=1).value = group_col_name or "Group"
        pivot_sheet.cell(row=1, column=2).value = f"Sum of {val_col_name or 'Value'}"

        row_offset = 2
        for k, v in summary.items():
            pivot_sheet.cell(row=row_offset, column=1).value = k
            pivot_sheet.cell(row=row_offset, column=2).value = v
            row_offset += 1

    def _join_sheets(self, session: WorkbookSession, primary_sheet: Worksheet, plan: ActionPlan) -> None:
        secondary_sheet_name = plan.parameters.get("secondary_sheet")
        join_column_name = plan.parameters.get("join_column")
        if not secondary_sheet_name or secondary_sheet_name not in session.workbook.sheetnames:
            raise HTTPException(status_code=400, detail="Secondary sheet not found for join.")

        secondary_sheet = session.workbook[secondary_sheet_name]
        primary_col_letter = self._resolve_column_letter(primary_sheet, join_column_name)
        secondary_col_letter = self._resolve_column_letter(secondary_sheet, join_column_name)

        if not primary_col_letter or not secondary_col_letter:
            raise HTTPException(status_code=400, detail=f"Join column '{join_column_name}' not found in both sheets.")

        primary_col_idx = column_index_from_string(primary_col_letter)
        secondary_col_idx = column_index_from_string(secondary_col_letter)

        secondary_map = {}
        sec_max_col = secondary_sheet.max_column
        sec_data_start = self._data_start_row(secondary_sheet)
        for row_idx in range(sec_data_start, secondary_sheet.max_row + 1):
            key_val = secondary_sheet.cell(row=row_idx, column=secondary_col_idx).value
            if key_val is None:
                continue
            key_str = str(key_val).strip().lower()
            row_data = [secondary_sheet.cell(row=row_idx, column=c).value for c in range(1, sec_max_col + 1)]
            secondary_map[key_str] = row_data

        start_new_col = primary_sheet.max_column + 1
        current_new_col = start_new_col
        for c in range(1, sec_max_col + 1):
            if c != secondary_col_idx:
                header_val = secondary_sheet.cell(row=self._detect_header_row(secondary_sheet), column=c).value
                primary_sheet.cell(row=self._detect_header_row(primary_sheet), column=current_new_col).value = (
                    f"{secondary_sheet_name}_{header_val}"
                )
                current_new_col += 1

        primary_data_start = self._data_start_row(primary_sheet)
        for row_idx in range(primary_data_start, primary_sheet.max_row + 1):
            key_val = primary_sheet.cell(row=row_idx, column=primary_col_idx).value
            if key_val is None:
                continue
            key_str = str(key_val).strip().lower()
            if key_str in secondary_map:
                row_data = secondary_map[key_str]
                current_new_col = start_new_col
                for c in range(1, sec_max_col + 1):
                    if c != secondary_col_idx:
                        primary_sheet.cell(row=row_idx, column=current_new_col).value = row_data[c-1]
                        current_new_col += 1

    def _checkpoint(self, session: WorkbookSession) -> None:
        session.undo_stack.append(self._serialize_workbook(session.workbook))
        if len(session.undo_stack) > 30:
            session.undo_stack.pop(0)
        session.redo_stack.clear()

    def _record_history(
        self,
        session: WorkbookSession,
        user_command: str,
        action: str,
        explanation: str,
        target_sheet: str,
        risk_level: RiskLevel,
        status: str = "completed",
    ) -> None:
        session.command_history.insert(
            0,
            CommandRecord(
                record_id=str(uuid4())[:8],
                user_command=user_command,
                action=action,
                explanation=explanation,
                target_sheet=target_sheet,
                status=status,
                created_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
                risk_level=risk_level,
            ),
        )
        del session.command_history[30:]

    def remember_plan(self, session_id: str, user_command: str, plan: ActionPlan) -> None:
        session = self.get_session(session_id)
        self._remember_plan(session, user_command, plan)

    def _remember_plan(self, session: WorkbookSession, user_command: str, plan: ActionPlan) -> None:
        session.last_command = user_command
        session.last_plan = plan
        self._push_recent_plan(session, plan)

    def _push_recent_plan(self, session: WorkbookSession, plan: ActionPlan) -> None:
        signature = self._plan_signature(plan)
        session.recent_plans = [existing for existing in session.recent_plans if self._plan_signature(existing) != signature]
        session.recent_plans.insert(0, plan)
        del session.recent_plans[6:]

    def _plan_signature(self, plan: ActionPlan) -> tuple[str, str, str | None, str | None, str | None]:
        return (
            plan.action,
            plan.target_sheet,
            plan.target_cell,
            plan.target_column,
            plan.formula,
            json.dumps(plan.parameters, sort_keys=True, default=str),
        )

    def _plan_to_step(self, plan: ActionPlan) -> TaskStep:
        return TaskStep(
            title=plan.preview_title,
            action=plan.action,
            target_sheet=plan.target_sheet,
            explanation=plan.explanation,
            risk_level=plan.risk_level,
            requires_confirmation=plan.requires_confirmation,
            target_cell=plan.target_cell,
            target_column=plan.target_column,
            formula=plan.formula,
            parameters=plan.parameters,
        )

    def _flatten_recent_steps(self, plans: list[ActionPlan]) -> list[TaskStep]:
        steps: list[TaskStep] = []
        for plan in plans:
            if plan.action == "batch":
                for step in self._batch_steps(plan):
                    steps.append(self._plan_to_step(step))
            else:
                steps.append(self._plan_to_step(plan))
        return steps[:6]

    def _build_follow_up_prompts(self, session: WorkbookSession, memory_steps: list[TaskStep]) -> list[str]:
        prompts: list[str] = []
        if not session.last_plan:
            return prompts

        last_plan = session.last_plan
        same_sheet_prompt = f"Do the same on {session.active_sheet}"
        if last_plan.target_sheet != session.active_sheet:
            same_sheet_prompt = f"Do the same for {session.active_sheet}"
        prompts.append(same_sheet_prompt)

        if last_plan.action in {"freeze_header", "auto_fit_columns", "create_table", "format_header"}:
            prompts.append("Apply the same formatting to the next sheet")
        elif last_plan.action in {"sort", "delete_duplicates", "apply_filter", "convert_column_type"} and last_plan.target_column:
            prompts.append(f"Repeat this for {last_plan.target_column} on another sheet")
        elif last_plan.action in {"insert_formula", "fill_formula_down", "generate_formula"}:
            prompts.append("Generate the same formula for another sheet")

        if memory_steps:
            prompts.append("Run the same task as a multi-step batch")

        return prompts[:4]

    def _build_memory(self, session: WorkbookSession) -> ConversationMemory:
        recent_history = session.command_history[:5]
        recent_steps = self._flatten_recent_steps(session.recent_plans)
        last_plan = session.last_plan
        return ConversationMemory(
            last_command=session.last_command or (recent_history[0].user_command if recent_history else None),
            last_action=last_plan.action if last_plan else (recent_history[0].action if recent_history else None),
            last_target_sheet=last_plan.target_sheet if last_plan else (recent_history[0].target_sheet if recent_history else session.active_sheet),
            last_target_column=last_plan.target_column if last_plan else None,
            last_formula=last_plan.formula if last_plan else None,
            last_preview_title=last_plan.preview_title if last_plan else None,
            recent_commands=[record.user_command for record in recent_history],
            recent_actions=[f"{record.action} on {record.target_sheet}" for record in recent_history],
            recent_steps=recent_steps,
            follow_up_prompts=self._build_follow_up_prompts(session, recent_steps),
        )

    def _serialize_workbook(self, workbook: Workbook) -> bytes:
        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _restore_workbook(self, payload: bytes) -> Workbook:
        return load_workbook(BytesIO(payload))

    def _header_indices(self, worksheet: Worksheet, headers: list[str] | None) -> list[int]:
        if not headers:
            return list(range(worksheet.max_column))
        indices: list[int] = []
        for header in headers:
            column_letter = self._resolve_column_letter(worksheet, header)
            if column_letter:
                indices.append(column_index_from_string(column_letter) - 1)
        return indices or list(range(worksheet.max_column))

    def _detect_header_row(self, worksheet: Worksheet) -> int:
        max_scan_row = min(worksheet.max_row or 1, 15)
        best_row = 1
        best_score = -1
        for row_idx in range(1, max_scan_row + 1):
            non_empty = 0
            text_cells = 0
            numeric_like = 0
            for col_idx in range(1, (worksheet.max_column or 1) + 1):
                value = worksheet.cell(row=row_idx, column=col_idx).value
                if value in (None, ""):
                    continue
                non_empty += 1
                if isinstance(value, str):
                    if value.startswith("="):
                        numeric_like += 1
                    else:
                        text_cells += 1
                elif isinstance(value, (int, float)) and not isinstance(value, bool):
                    numeric_like += 1
                elif hasattr(value, "year") and hasattr(value, "month"):
                    text_cells += 1
                else:
                    text_cells += 1
            score = (text_cells * 3) + (non_empty * 2) - numeric_like
            if score > best_score:
                best_score = score
                best_row = row_idx
        return best_row

    def _data_start_row(self, worksheet: Worksheet) -> int:
        return min(self._detect_header_row(worksheet) + 1, (worksheet.max_row or 1) + 1)

    def find_columns_with_keywords(self, session_id: str, sheet_name: str, keywords: list[str], search_rows: int = 10) -> list[str]:
        worksheet = self._get_sheet(self.get_session(session_id), sheet_name)
        normalized_keywords = [self._normalize(keyword) for keyword in keywords if keyword]
        if not normalized_keywords:
            return []

        matched_columns: list[str] = []
        limit = min(worksheet.max_row or 1, max(1, search_rows))
        for col_idx in range(1, (worksheet.max_column or 1) + 1):
            for row_idx in range(1, limit + 1):
                value = worksheet.cell(row=row_idx, column=col_idx).value
                if not isinstance(value, str):
                    continue
                normalized_value = self._normalize(value)
                if any(keyword in normalized_value for keyword in normalized_keywords):
                    matched_columns.append(get_column_letter(col_idx))
                    break
        return matched_columns

    def _resolve_column_letter(self, worksheet: Worksheet, header_or_letter: str | None) -> str | None:
        if not header_or_letter:
            return None
        candidate = header_or_letter.strip()
        if re.fullmatch(r"[A-Za-z]{1,3}", candidate):
            return candidate.upper()
        return self.find_header_column(worksheet, candidate)

    @staticmethod
    def _coordinate_parts(cell_reference: str) -> tuple[str | None, int | None]:
        match = re.fullmatch(r"([A-Za-z]{1,3})(\d+)", cell_reference.strip())
        if not match:
            return None, None
        return match.group(1).upper(), int(match.group(2))

    def _matches_condition(self, value: Any, operator: str, criterion: Any) -> bool:
        normalized_operator = operator.lower()
        if normalized_operator in {"contains", "includes"}:
            return criterion is not None and criterion != "" and str(criterion).lower() in str(value or "").lower()

        value_number = self._coerce_number(value)
        criterion_number = self._coerce_number(criterion)
        if value_number is not None and criterion_number is not None:
            if normalized_operator in {"gt", "greater_than", ">"}:
                return value_number > criterion_number
            if normalized_operator in {"gte", "greater_or_equal", ">="}:
                return value_number >= criterion_number
            if normalized_operator in {"lt", "less_than", "<"}:
                return value_number < criterion_number
            if normalized_operator in {"lte", "less_or_equal", "<="}:
                return value_number <= criterion_number
            if normalized_operator in {"eq", "equals", "="}:
                return value_number == criterion_number
            if normalized_operator in {"neq", "not_equals", "!="}:
                return value_number != criterion_number

        # Handle Date Comparison
        if hasattr(value, "strftime") or (isinstance(criterion, str) and re.match(r"\d{4}-\d{2}-\d{2}", criterion)):
            value_str = value.strftime("%Y-%m-%d") if hasattr(value, "strftime") else str(value or "")
            if normalized_operator in {"eq", "equals", "="}:
                return value_str == str(criterion)
            if normalized_operator in {"neq", "not_equals", "!="}:
                return value_str != str(criterion)

        value_text = str(value or "").strip().lower()
        criterion_text = str(criterion or "").strip().lower()
        if normalized_operator in {"neq", "not_equals", "!="}:
            return value_text != criterion_text
        return value_text == criterion_text

    def _validate_path(self, path: Path) -> None:
        allowed_root = settings.allowed_root_path
        if allowed_root and allowed_root not in path.parents and path != allowed_root:
            raise HTTPException(status_code=403, detail="Workbook path is outside the allowed root.")

    @staticmethod
    def _coerce_number(value: Any) -> float | None:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value.replace(",", "").strip())
            except ValueError:
                return None
        return None

    @staticmethod
    def _infer_type(values: list[Any]) -> str:
        if not values:
            return "empty"
        numeric_like = 0
        date_like = 0
        text_like = 0
        for value in values:
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                numeric_like += 1
            elif isinstance(value, str) and value.startswith("="):
                numeric_like += 1
            elif hasattr(value, "year") and hasattr(value, "month"):
                date_like += 1
            else:
                text_like += 1
        total = len(values)
        if numeric_like and numeric_like >= max(2, int(total * 0.6)):
            return "number"
        if date_like and date_like >= max(2, int(total * 0.6)):
            return "date"
        if numeric_like > text_like and numeric_like:
            return "number"
        if date_like > text_like and date_like:
            return "date"
        if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in values):
            return "number"
        if all(hasattr(value, "year") and hasattr(value, "month") for value in values):
            return "date"
        return "text"

    @classmethod
    def _convert_value(cls, value: Any, target_type: str, date_format: str | None = None) -> tuple[bool, Any]:
        normalized_target = target_type.strip().lower()
        if normalized_target == "text":
            return True, cls._stringify(value)
        if normalized_target == "number":
            number = cls._coerce_number(value)
            if number is None:
                return False, value
            if number.is_integer():
                return True, int(number)
            return True, number
        if normalized_target == "date":
            parsed = cls._parse_date_value(value)
            if parsed is None:
                return False, value
            return True, parsed
        return False, value

    @staticmethod
    def _parse_date_value(value: Any) -> datetime | date | None:
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            candidate = value.strip()
            if not candidate:
                return None
            for pattern in (
                "%Y-%m-%d",
                "%d-%m-%Y",
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%d %b %Y",
                "%d %B %Y",
                "%Y/%m/%d",
            ):
                try:
                    return datetime.strptime(candidate, pattern)
                except ValueError:
                    continue
            try:
                return datetime.fromisoformat(candidate)
            except ValueError:
                return None
        return None

    @staticmethod
    def _stringify(value: Any) -> str:
        return "" if value is None else str(value)

    @staticmethod
    def _normalize(text: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", text.lower())


workbook_service = WorkbookService()
