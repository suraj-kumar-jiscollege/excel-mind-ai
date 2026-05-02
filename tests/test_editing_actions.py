from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys
import unittest

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.models import ActionImpact, ActionPlan
from app.services.workbook_service import workbook_service


def _make_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value", "Status"])
    sheet.append(["Alpha", 10, "Open"])
    sheet.append(["Beta", 20, "Closed"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class EditingActionTests(unittest.TestCase):
    def setUp(self) -> None:
        snapshot = workbook_service.open_workbook_from_bytes(_make_workbook_bytes(), "editing-test.xlsx")
        self.session_id = snapshot.session_id

    def test_insert_and_delete_rows(self) -> None:
        workbook_service.execute_action(
            self.session_id,
            ActionPlan(
                action="insert_rows",
                target_sheet="Data",
                preview_title="Insert rows",
                explanation="Insert one row.",
                risk_level="low",
                requires_confirmation=False,
                parameters={"row_index": 2, "amount": 1},
                impact=ActionImpact(summary="Insert row"),
            ),
        )
        snapshot = workbook_service.get_snapshot(self.session_id)
        self.assertEqual(snapshot.sheets[0].max_row, 4)

        workbook_service.execute_action(
            self.session_id,
            ActionPlan(
                action="delete_rows",
                target_sheet="Data",
                preview_title="Delete rows",
                explanation="Delete one row.",
                risk_level="high",
                requires_confirmation=True,
                parameters={"row_index": 2, "amount": 1},
                impact=ActionImpact(summary="Delete row"),
            ),
        )
        snapshot = workbook_service.get_snapshot(self.session_id)
        self.assertEqual(snapshot.sheets[0].max_row, 3)

    def test_rename_merge_comment_and_link(self) -> None:
        workbook_service.execute_action(
            self.session_id,
            ActionPlan(
                action="rename_sheet",
                target_sheet="Data",
                preview_title="Rename sheet",
                explanation="Rename the sheet.",
                risk_level="low",
                requires_confirmation=False,
                parameters={"new_name": "Summary"},
                impact=ActionImpact(summary="Rename"),
            ),
        )
        snapshot = workbook_service.get_snapshot(self.session_id)
        self.assertEqual(snapshot.active_sheet, "Summary")
        self.assertIn("Summary", [sheet.name for sheet in snapshot.sheets])

        workbook_service.execute_action(
            self.session_id,
            ActionPlan(
                action="merge_cells",
                target_sheet="Summary",
                preview_title="Merge cells",
                explanation="Merge title cells.",
                risk_level="medium",
                requires_confirmation=True,
                impacted_range="A1:B1",
                parameters={"range": "A1:B1"},
                impact=ActionImpact(summary="Merge"),
            ),
        )
        session = workbook_service.get_session(self.session_id)
        self.assertIn("A1:B1", [str(rng) for rng in session.workbook["Summary"].merged_cells.ranges])

        workbook_service.execute_action(
            self.session_id,
            ActionPlan(
                action="add_comment",
                target_sheet="Summary",
                target_cell="A1",
                preview_title="Add comment",
                explanation="Add a note.",
                risk_level="low",
                requires_confirmation=False,
                parameters={"cell": "A1", "text": "Review this value", "author": "Copilot"},
                impact=ActionImpact(summary="Comment"),
            ),
        )
        self.assertIsNotNone(session.workbook["Summary"]["A1"].comment)

        workbook_service.execute_action(
            self.session_id,
            ActionPlan(
                action="add_hyperlink",
                target_sheet="Summary",
                target_cell="C1",
                preview_title="Add hyperlink",
                explanation="Add a link.",
                risk_level="low",
                requires_confirmation=False,
                parameters={"cell": "C1", "url": "https://example.com", "text": "Docs"},
                impact=ActionImpact(summary="Hyperlink"),
            ),
        )
        self.assertEqual(session.workbook["Summary"]["C1"].hyperlink.target, "https://example.com")

    def test_validation_and_conditional_formatting(self) -> None:
        workbook_service.execute_action(
            self.session_id,
            ActionPlan(
                action="add_validation",
                target_sheet="Data",
                preview_title="Add validation",
                explanation="Add dropdown.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range="C2:C10",
                parameters={"range": "C2:C10", "validation_type": "list", "source": '"Open,Closed"'},
                impact=ActionImpact(summary="Validation"),
            ),
        )
        session = workbook_service.get_session(self.session_id)
        self.assertTrue(session.workbook["Data"].data_validations.dataValidation)

        workbook_service.execute_action(
            self.session_id,
            ActionPlan(
                action="conditional_format_range",
                target_sheet="Data",
                preview_title="Conditional formatting",
                explanation="Highlight high values.",
                risk_level="low",
                requires_confirmation=False,
                impacted_range="B2:B10",
                parameters={"range": "B2:B10", "threshold": 15, "operator": "greaterThan", "color": "C7F9CC"},
                impact=ActionImpact(summary="CF"),
            ),
        )
        self.assertTrue(session.workbook["Data"].conditional_formatting)


if __name__ == "__main__":
    unittest.main()
