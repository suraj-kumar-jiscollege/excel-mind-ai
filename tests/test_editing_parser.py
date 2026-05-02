from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys
import unittest

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.ai_service import ai_service
from app.services.workbook_service import workbook_service


def _make_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Name", "Value", "Status"])
    sheet.append(["Alpha", 10, "Open"])
    sheet.append(["Beta", 20, "Closed"])
    sheet.append(["Gamma", 30, "Open"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class EditingParserTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self) -> None:
        snapshot = workbook_service.open_workbook_from_bytes(_make_workbook_bytes(), "parser-test.xlsx")
        self.session_id = snapshot.session_id

    async def test_row_and_column_phrasing(self) -> None:
        insert_plan = await ai_service.preview_command(self.session_id, "please insert 2 rows above row 5 on Data")
        self.assertEqual(insert_plan.action, "insert_rows")
        self.assertEqual(insert_plan.parameters["row_index"], 5)
        self.assertEqual(insert_plan.parameters["amount"], 2)

        delete_plan = await ai_service.preview_command(self.session_id, "remove 1 column at B on Data")
        self.assertEqual(delete_plan.action, "delete_columns")
        self.assertEqual(delete_plan.parameters["column_letter"], "B")
        self.assertEqual(delete_plan.parameters["amount"], 1)

    async def test_sheet_and_cell_edit_phrasing(self) -> None:
        rename_plan = await ai_service.preview_command(self.session_id, "rename current sheet to Summary")
        self.assertEqual(rename_plan.action, "rename_sheet")
        self.assertEqual(rename_plan.parameters["new_name"], "Summary")

        clear_plan = await ai_service.preview_command(self.session_id, "clear contents in B2:C3")
        self.assertEqual(clear_plan.action, "clear_cells")
        self.assertEqual(clear_plan.impacted_range, "B2:C3")

        merge_plan = await ai_service.preview_command(self.session_id, "merge A1:B1")
        self.assertEqual(merge_plan.action, "merge_cells")
        self.assertEqual(merge_plan.impacted_range, "A1:B1")

        unmerge_plan = await ai_service.preview_command(self.session_id, "split cells A1:B1")
        self.assertEqual(unmerge_plan.action, "unmerge_cells")
        self.assertEqual(unmerge_plan.impacted_range, "A1:B1")

    async def test_validation_comment_link_and_highlight(self) -> None:
        validation_plan = await ai_service.preview_command(self.session_id, 'add dropdown validation to C2:C10 from "Open,Closed"')
        self.assertEqual(validation_plan.action, "add_validation")
        self.assertEqual(validation_plan.parameters["validation_type"], "list")

        comment_plan = await ai_service.preview_command(self.session_id, "add note to A1: review this")
        self.assertEqual(comment_plan.action, "add_comment")
        self.assertEqual(comment_plan.target_cell, "A1")

        link_plan = await ai_service.preview_command(self.session_id, "add link to C1 https://example.com")
        self.assertEqual(link_plan.action, "add_hyperlink")
        self.assertEqual(link_plan.target_cell, "C1")

        highlight_plan = await ai_service.preview_command(self.session_id, "highlight B2:B10 above 15")
        self.assertEqual(highlight_plan.action, "conditional_format_range")
        self.assertEqual(highlight_plan.impacted_range, "B2:B10")


if __name__ == "__main__":
    unittest.main()
